from __future__ import annotations

import asyncio
import json
import sqlite3
import importlib.util
import re
import subprocess
import sys
import threading
import time
import uuid
import shutil
from contextlib import asynccontextmanager
from datetime import datetime
from hashlib import sha256
from pathlib import Path
from typing import Any, Iterator, Literal
from urllib.parse import quote, urlparse
from urllib.error import URLError
from urllib.request import urlopen

import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import yaml
from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response, StreamingResponse
from pydantic import AliasChoices, BaseModel, Field, SecretStr

from icm_platform.ai_service import AIConfigurationError, AIProviderError, AIService
from icm_platform.generation_control import GenerationCancellationRegistry
from icm_platform.assets import latest_screenshots, list_batch_child_reports, list_cases, list_reports, parse_report, read_report
from icm_platform.db import (
    connect,
    create_project_profile,
    delete_project_profile,
    get_ai_settings,
    get_platform_settings,
    get_project_profile,
    init_db,
    list_project_profiles,
    rows_to_dicts,
    save_ai_settings,
    save_platform_settings,
    update_project_profile,
    utc_now,
)
from icm_platform.paths import DB_PATH, DRAFT_RUN_DIR, OBSERVED_ASSET_DIR, REPORT_DIR, ROOT, SCREENSHOTS_LATEST_DIR, SCREENSHOTS_RUNS_DIR, SPEC_FILE, TEST_CASE_DIR
from icm_platform.run_views import summarize_run_task
from icm_platform import recorder
from icm_platform.recorder_runtime import RecorderRuntime
from icm_platform.codegen_experiment_runtime import CodegenExperimentError, CodegenExperimentRuntime
from icm_platform.worker import RunnerWorker
from runner.evidence_recorder import EVIDENCE_ROOT, TRACE_ROOT, evidence_summary, read_jsonl
from runner.element_knowledge_refresh import refresh_element_knowledge, refresh_library_file
from runner.element_library_validator import validate_element_library
from runner.element_route_discovery import discover_routes
from runner.element_knowledge_report import DEFAULT_LIBRARY_PATH as ELEMENT_LIBRARY_PATH, DEFAULT_SUMMARY_PATH as ELEMENT_SUMMARY_PATH, build_report_model, load_json_file
from runner.element_scanner import is_login_url
from runner.browser import attach_browser_over_cdp, close_browser, launch_browser
from runner.environment_config import build_scan_targets_from_profile, list_environment_profiles, resolve_scan_settings, with_account_credentials
from runner.login_manager import ensure_storage_state_for_profile, existing_storage_state_for_profile, is_login_state_valid, resolve_storage_state_path
from runner.step_details import load_step_details

ai_service = AIService()
worker = RunnerWorker()
generation_control = GenerationCancellationRegistry()
recorder_runtime = RecorderRuntime()
codegen_experiment_runtime = CodegenExperimentRuntime()

_DEDICATED_CDP_DEFAULT_PORT = 9222
_DEDICATED_CDP_PROFILE_ROOT = ROOT / "platform-data" / "chrome-element-scan"

# AI测试任务队列只允许展示测试执行类任务。
# 其他平台维护任务（例如元素知识库刷新）必须进入自己的业务模块。
AI_TEST_TASK_MODES = {
    "run-case",
    "run-batch",
    "run-draft",
    "agent-explore",
}


@asynccontextmanager
async def lifespan(_app: FastAPI):
    init_db()
    worker.start()
    yield


app = FastAPI(title="ICM AI Automation Platform", version="0.2.0", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://127.0.0.1:5175", "http://127.0.0.1:5176", "http://localhost:5175"],
    allow_origin_regex=r"^https?://[^/]+:(5175|5176)$",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)


class AISettingsRequest(BaseModel):
    provider: str | None = None
    api_key: str | None = None
    base_url: str | None = None
    model: str | None = None


class PlatformSettingsRequest(BaseModel):
    runner: dict | None = None
    asset_policy: dict | None = None
    environment: dict | None = None
    accounts: dict | None = None


class RequirementRequest(BaseModel):
    title: str = "未命名需求"
    document: str
    context_info: dict | None = None
    project_id: str | None = None
    generation_id: str | None = None


class RequirementPatchRequest(BaseModel):
    title: str | None = None
    document: str | None = None
    status: str | None = None
    project_id: str | None = None


class GenerateCasesRequest(BaseModel):
    test_points: list[dict]


class RequirementSpecRequest(BaseModel):
    title: str = "未命名需求"
    document: str


class TestPointPatchRequest(BaseModel):
    parent_id: int | None = None
    name: str | None = None
    category: str | None = None
    priority: str | None = None
    status: str | None = None
    description: str | None = None
    module: str | None = None
    source: str | None = None
    sort_order: int | None = None


class TestPointCreateRequest(BaseModel):
    requirement_id: int | None = None
    parent_id: int | None = None
    name: str
    category: str = "功能"
    priority: str = "P1"
    status: str = "已确认"
    description: str = ""
    module: str = ""
    source: str = "manual"
    sort_order: int | None = None


class TestPointReorderUpdate(BaseModel):
    id: int
    parent_id: int | None = None
    sort_order: int
    module: str | None = None
    category: str | None = None


class TestPointReorderRequest(BaseModel):
    updates: list[TestPointReorderUpdate]


class SelectedTestPointsRequest(BaseModel):
    test_point_ids: list[int]
    template: str = "functional"
    title: str | None = None
    generator: Literal["rule", "ai"] = "rule"


class CaseDraftPatchRequest(BaseModel):
    title: str | None = None
    yaml: str | None = None
    status: str | None = None


class CaseDraftCreateRequest(BaseModel):
    requirement_id: int | None = None
    title: str = "新增用例草稿"
    yaml: str | None = None
    template: str = "manual"


class CaseDraftBatchDeleteRequest(BaseModel):
    draft_ids: list[int]


class PromoteDraftRequest(BaseModel):
    case_id: str
    filename: str | None = None


class ValidateDraftRequest(BaseModel):
    yaml: str | None = None
    case_id: str | None = None


class AnalyzeReportRequest(BaseModel):
    force: bool = False


class ReportBatchDeleteRequest(BaseModel):
    run_ids: list[str]


class ElementKnowledgeRefreshRequest(BaseModel):
    no_scan: bool = True
    min_healing_failures: int = 1
    base_url: str | None = None
    environment_id: str | None = None
    target_url: str | None = None
    target_page_id: str | None = None
    target_name: str | None = None
    include_states: bool = False
    headless: bool = True


class ElementKnowledgeRefreshTaskRequest(BaseModel):
    no_scan: bool = True
    min_healing_failures: int = 1
    base_url: str | None = None
    environment_id: str | None = None
    target_url: str | None = None
    target_page_id: str | None = None
    target_name: str | None = None
    include_states: bool = False
    headless: bool = True


class ElementKnowledgeValidationTaskRequest(BaseModel):
    environment_id: str


class RunRequest(BaseModel):
    mode: Literal["run-case", "run-batch", "run-draft", "agent-explore"]
    case_id: str | None = None
    case_ids: list[str] = Field(default_factory=list)
    draft_id: int | None = None


class RecorderStartRequest(BaseModel):
    start_url: str = Field(validation_alias=AliasChoices("start_url", "entry_url"))


class CodegenExperimentStartRequest(BaseModel):
    start_url: str = Field(validation_alias=AliasChoices("start_url", "entry_url"))


class CodegenExperimentRunRequest(BaseModel):
    variables: dict[str, SecretStr] = Field(default_factory=dict)


def _recorder_allowed_origins() -> list[str]:
    environment = get_platform_settings(mask_secrets=False).get("environment") or {}
    origins: set[str] = set()
    for value in environment.values():
        parsed = urlparse(str(value or ""))
        if parsed.scheme in {"http", "https"} and parsed.netloc and not parsed.username and not parsed.password:
            origins.add(f"{parsed.scheme}://{parsed.netloc}")
    return sorted(origins)


def _codegen_experiment_view(session_id: str) -> dict[str, Any]:
    state = codegen_experiment_runtime.get(session_id)
    running = bool(state.process and state.process.poll() is None and not state.stopped)
    script = codegen_experiment_runtime.read_script(session_id)
    status = "failed" if state.error else "recording" if running else "stopped" if state.stopped else "failed"
    warnings = ["仅用于比较 Playwright Codegen 输出；不会写入 Recorder、候选脚本或回归集。"]
    if script:
        warnings.append("敏感字段在预览中已脱敏；原始临时脚本不会被发布。")
    return {
        "id": session_id,
        "mode": "codegen-experiment",
        "status": status,
        "start_url": state.start_url,
        "script": script,
        "inputs": getattr(state, "input_variables", []),
        "error": state.error,
        "run": {
            "status": getattr(state, "run_status", "not_started"),
            "error": getattr(state, "run_error", None),
        },
        "warnings": warnings,
    }


@app.post("/api/codegen-experiments")
def create_codegen_experiment(payload: CodegenExperimentStartRequest) -> dict:
    if urlparse(payload.start_url).scheme + "://" + urlparse(payload.start_url).netloc not in _recorder_allowed_origins():
        raise HTTPException(status_code=400, detail="Codegen experiment entry URL is not allowlisted")
    try:
        state = codegen_experiment_runtime.start(payload.start_url)
        return _codegen_experiment_view(state.session_id)
    except CodegenExperimentError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc


@app.get("/api/codegen-experiments/{session_id}")
def get_codegen_experiment(session_id: str) -> dict:
    try:
        return _codegen_experiment_view(session_id)
    except CodegenExperimentError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.post("/api/codegen-experiments/{session_id}/stop")
def stop_codegen_experiment(session_id: str) -> dict:
    try:
        codegen_experiment_runtime.stop(session_id)
        return _codegen_experiment_view(session_id)
    except CodegenExperimentError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/codegen-experiments/{session_id}/run")
def run_codegen_experiment(session_id: str, payload: CodegenExperimentRunRequest) -> dict:
    try:
        codegen_experiment_runtime.run(session_id, {name: value.get_secret_value() for name, value in payload.variables.items()})
        return _codegen_experiment_view(session_id)
    except CodegenExperimentError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


def _recorder_step(session_id: str, event: dict[str, Any]) -> dict[str, Any]:
    action = event["action"]
    selector = action.get("selector") or {}
    return {
        "id": f"{session_id}-{event['sequence']}",
        "sequence": event["sequence"],
        "action": action.get("type", "unknown"),
        "locator": selector.get("value") or None,
        "url": action.get("url") or None,
        "value": "[redacted]" if action.get("redacted") else action.get("value") or None,
        "status": "blocked" if action.get("review_required") else "recorded",
        "warning": selector.get("reason") or None,
        "created_at": event.get("created_at"),
    }


def _recorder_session_view(session_id: str) -> dict[str, Any]:
    with connect() as conn:
        session = recorder.get_session(conn, session_id)
        events = recorder.list_events(conn, session_id)
    state = recorder_runtime.get(session_id)
    candidate = None
    warnings = [str(event["action"].get("selector", {}).get("reason")) for event in events if event["action"].get("selector", {}).get("reason")]
    if session.get("candidate_yaml") and session.get("candidate_python"):
        try:
            dsl = yaml.safe_load(session["candidate_yaml"]) or {}
        except yaml.YAMLError:
            dsl = {}
        warnings.append("Candidate scripts require manual assertions and approval before publication.")
        candidate = {
            "yaml": session["candidate_yaml"],
            "playwright_python": session["candidate_python"],
            "publishable": bool(dsl.get("publishable")),
            "blocking_warnings": list(dict.fromkeys(warnings)),
        }
    return {
        "id": session_id,
        "status": session["status"],
        "start_url": session["start_url"],
        "current_url": state.current_url if state else session["start_url"],
        "stream_url": f"/api/recordings/{session_id}/events",
        "steps": [_recorder_step(session_id, event) for event in events],
        "candidate": candidate,
        "error": state.error if state and state.error else session.get("failure_reason"),
    }


@app.post("/api/recordings")
@app.post("/api/recorder/sessions", include_in_schema=False)
def create_recorder_session(payload: RecorderStartRequest) -> dict:
    allowed_origins = _recorder_allowed_origins()
    try:
        with connect() as conn:
            session = recorder.create_session(conn, start_url=payload.start_url, allowed_origins=allowed_origins)
        recorder_runtime.start(session["id"], session["start_url"])
    except recorder.RecorderError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"unable to start isolated recorder browser: {exc}") from exc
    return _recorder_session_view(session["id"])


@app.get("/api/recordings/{session_id}")
@app.get("/api/recorder/sessions/{session_id}", include_in_schema=False)
def get_recorder_session(session_id: str) -> dict:
    try:
        return _recorder_session_view(session_id)
    except recorder.RecorderError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.post("/api/recordings/{session_id}/stop")
@app.post("/api/recorder/sessions/{session_id}/stop", include_in_schema=False)
def stop_recorder_session(session_id: str) -> dict:
    try:
        with connect() as conn:
            recorder.get_session(conn, session_id)
        recorder_runtime.stop(session_id)
        with connect() as conn:
            recorder.stop_session(conn, session_id)
        return _recorder_session_view(session_id)
    except recorder.RecorderError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/recordings/{session_id}/candidate")
@app.post("/api/recorder/sessions/{session_id}/candidate", include_in_schema=False)
def create_recorder_candidate(session_id: str) -> dict:
    try:
        with connect() as conn:
            session = recorder.get_session(conn, session_id)
            if session["status"] != "stopped":
                raise HTTPException(status_code=409, detail="stop recording before generating a candidate")
            if not session.get("candidate_yaml"):
                recorder.stop_session(conn, session_id)
        return _recorder_session_view(session_id)
    except recorder.RecorderError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.get("/api/recordings/{session_id}/events")
@app.get("/api/recordings/{session_id}/stream", include_in_schema=False)
@app.get("/api/recorder/sessions/{session_id}/events", include_in_schema=False)
def recorder_events_stream(session_id: str) -> StreamingResponse:
    try:
        with connect() as conn:
            recorder.get_session(conn, session_id)
    except recorder.RecorderError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    def stream() -> Iterator[str]:
        after = 0
        while True:
            with connect() as conn:
                events = recorder.list_events(conn, session_id, after=after)
                session = recorder.get_session(conn, session_id)
            for event in events:
                after = event["sequence"]
                state = recorder_runtime.get(session_id)
                payload = {
                    "id": session_id,
                    "status": session["status"],
                    "current_url": state.current_url if state else session["start_url"],
                    "step": _recorder_step(session_id, event),
                }
                yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"
            if session["status"] != "recording":
                break
            time.sleep(0.4)

    return StreamingResponse(stream(), media_type="text/event-stream", headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# -----------------------------------------------------------------------------
# P0 · 所属项目下拉化（增量 2026-06-10）
# -----------------------------------------------------------------------------


class ProjectCreateRequest(BaseModel):
    """T2 · POST /api/projects 入参。name 必填；base_url / description 可空。"""

    name: str
    base_url: str | None = None
    description: str | None = None


class ProjectPatchRequest(BaseModel):
    """T2 · PATCH /api/projects/{id} 入参。name / base_url / description 三选其一。"""

    name: str | None = None
    base_url: str | None = None
    description: str | None = None


@app.get("/api/projects")
def list_projects() -> list[dict]:
    """T2 · GET /api/projects：返回 project_profiles 全部记录（按 created_at, name 排序）。"""
    return list_project_profiles()


@app.post("/api/projects")
def create_project(payload: ProjectCreateRequest) -> dict:
    """T2 · POST /api/projects：name 必填；UNIQUE 冲突 → 409。"""
    try:
        return create_project_profile(payload.model_dump(exclude_unset=True))
    except ValueError as exc:
        msg = str(exc)
        if msg.startswith("conflict:"):
            raise HTTPException(status_code=409, detail=msg.split(":", 1)[1].strip()) from exc
        if msg.startswith("invalid:"):
            raise HTTPException(status_code=400, detail=msg.split(":", 1)[1].strip()) from exc
        raise


@app.patch("/api/projects/{project_id}")
def update_project(project_id: str, payload: ProjectPatchRequest) -> dict:
    """T2 · PATCH /api/projects/{id}：改名 / 补 description / 改 base_url；404 / 409。"""
    try:
        result = update_project_profile(project_id, payload.model_dump(exclude_unset=True))
    except ValueError as exc:
        msg = str(exc)
        if msg.startswith("conflict:"):
            raise HTTPException(status_code=409, detail=msg.split(":", 1)[1].strip()) from exc
        if msg.startswith("invalid:"):
            raise HTTPException(status_code=400, detail=msg.split(":", 1)[1].strip()) from exc
        raise
    if result is None:
        raise HTTPException(status_code=404, detail="project not found")
    return result


@app.delete("/api/projects/{project_id}")
def delete_project(project_id: str) -> dict:
    """T2 · DELETE /api/projects/{id}：返回 {ok: true}；404。"""
    ok = delete_project_profile(project_id)
    if not ok:
        raise HTTPException(status_code=404, detail="project not found")
    return {"ok": True}


@app.get("/api/health")
def health() -> dict[str, str]:
    return {"status": "ok", "runner": "python -m runner.main", "api_version": "0.2.0"}


@app.get("/api/system/health")
def system_health() -> dict:
    runner_entry = ROOT / "runner" / "main.py"
    chrome_candidates = [
        Path("C:/Program Files/Google/Chrome/Application/chrome.exe"),
        Path("C:/Program Files (x86)/Google/Chrome/Application/chrome.exe"),
    ]
    chrome_path = next((path for path in chrome_candidates if path.exists()), None)
    return {
        "api": {"status": "ok", "version": app.version},
        "runner": {
            "status": "ok" if runner_entry.exists() else "missing",
            "entry": "python -m runner.main",
            "path": str(runner_entry),
        },
        "playwright": {
            "available": importlib.util.find_spec("playwright") is not None,
            "chrome_available": chrome_path is not None,
            "chrome_path": str(chrome_path) if chrome_path else "",
        },
        "paths": {
            "reports": path_health(REPORT_DIR),
            "screenshots_latest": path_health(SCREENSHOTS_LATEST_DIR),
            "screenshots_runs": path_health(SCREENSHOTS_RUNS_DIR),
            "observed_assets": path_health(OBSERVED_ASSET_DIR),
        },
        "sqlite": sqlite_health(),
    }


@app.get("/api/ai/settings")
def ai_settings() -> dict:
    return get_ai_settings(mask_key=True)


@app.put("/api/ai/settings")
def update_ai_settings(payload: AISettingsRequest) -> dict:
    return save_ai_settings(payload.model_dump(exclude_unset=True))


@app.get("/api/platform/settings")
def platform_settings() -> dict:
    return get_platform_settings()


@app.put("/api/platform/settings")
def update_platform_settings(payload: PlatformSettingsRequest) -> dict:
    return save_platform_settings(payload.model_dump(exclude_unset=True))


@app.post("/api/ai/test-connection")
def test_ai_connection() -> dict[str, str]:
    try:
        return ai_service.test_connection(get_ai_settings(mask_key=False))
    except AIConfigurationError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except AIProviderError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@app.get("/api/ai/ollama/models")
def ollama_models(base_url: str | None = Query(default=None)) -> dict:
    settings = get_ai_settings(mask_key=False)
    target_base_url = base_url or settings.get("base_url") or "http://192.168.12.38:11434/v1"
    try:
        return ai_service.list_ollama_models(target_base_url)
    except AIConfigurationError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except AIProviderError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@app.get("/api/cases")
def cases() -> list[dict]:
    return list_cases()


def resolve_requirement_project_id(project_id: str | None = None) -> str | None:
    if project_id and get_project_profile(project_id):
        return project_id
    if get_project_profile("proj-icm-default"):
        return "proj-icm-default"
    projects = list_project_profiles()
    return str(projects[0]["id"]) if projects else None


@app.get("/api/cases/{case_id}")
def case_detail(case_id: str) -> dict:
    case_id_norm = normalize_case_id(case_id)
    case_path = find_case_yaml(case_id_norm)
    yaml_text = case_path.read_text(encoding="utf-8")
    try:
        data = yaml.safe_load(yaml_text) or {}
    except yaml.YAMLError as exc:
        raise HTTPException(status_code=400, detail=f"case YAML parse failed: {exc}") from exc
    if not isinstance(data, dict):
        raise HTTPException(status_code=400, detail="case YAML root must be an object")
    automation_asset = data.get("automation_asset")
    return {
        "id": str(data.get("id") or case_id_norm),
        "title": str(data.get("title") or case_id_norm),
        "status": str(data.get("status") or ""),
        "path": str(case_path),
        "has_automation_asset": isinstance(automation_asset, dict) and bool(automation_asset),
        "yaml": yaml_text,
    }


@app.get("/api/requirements")
def requirements() -> list[dict]:
    with connect() as conn:
        rows = conn.execute(
            """
            select r.*,
              (select count(*) from test_points tp where tp.requirement_id = r.id) as test_point_count,
              (select count(*) from case_drafts cd where cd.requirement_id = r.id) as draft_count
            from requirements r
            order by r.created_at desc
            limit 50
            """
        ).fetchall()
    return rows_to_dicts(rows)


@app.post("/api/requirements")
def create_requirement(payload: RequirementRequest) -> dict:
    title = payload.title.strip()
    document = payload.document.strip()
    if not title:
        raise HTTPException(status_code=400, detail="title cannot be empty")
    if not document:
        raise HTTPException(status_code=400, detail="document cannot be empty")
    now = utc_now()
    project_id = resolve_requirement_project_id(payload.project_id)
    with connect() as conn:
        cur = conn.execute(
            """
            insert into requirements(title, document, status, project_id, created_at, updated_at)
            values (?, ?, 'draft', ?, ?, ?)
            """,
            (title, document, project_id, now, now),
        )
        requirement_id = cur.lastrowid
    detail = load_requirement_detail(int(requirement_id))
    if not detail:
        raise HTTPException(status_code=500, detail="requirement created but not found")
    return detail


@app.get("/api/requirements/{requirement_id}")
def requirement_detail(requirement_id: int) -> dict:
    detail = load_requirement_detail(requirement_id)
    if not detail:
        raise HTTPException(status_code=404, detail="requirement not found")
    return detail


@app.patch("/api/requirements/{requirement_id}")
def update_requirement(requirement_id: int, payload: RequirementPatchRequest) -> dict:
    updates: dict[str, str] = {}
    if payload.title is not None:
        title = payload.title.strip()
        if not title:
            raise HTTPException(status_code=400, detail="title cannot be empty")
        updates["title"] = title
    if payload.document is not None:
        document = payload.document.strip()
        if not document:
            raise HTTPException(status_code=400, detail="document cannot be empty")
        updates["document"] = document
    if payload.status is not None:
        status = payload.status.strip()
        if not status:
            raise HTTPException(status_code=400, detail="status cannot be empty")
        updates["status"] = status
    if payload.project_id is not None:
        updates["project_id"] = resolve_requirement_project_id(payload.project_id)
    if not updates:
        detail = load_requirement_detail(requirement_id)
        if not detail:
            raise HTTPException(status_code=404, detail="requirement not found")
        return detail
    updates["updated_at"] = utc_now()
    keys = list(updates)
    with connect() as conn:
        existing = conn.execute("select id from requirements where id = ?", (requirement_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="requirement not found")
        conn.execute(
            f"update requirements set {', '.join(f'{key} = ?' for key in keys)} where id = ?",
            [updates[key] for key in keys] + [requirement_id],
        )
    detail = load_requirement_detail(requirement_id)
    if not detail:
        raise HTTPException(status_code=404, detail="requirement not found")
    return detail


@app.delete("/api/requirements/{requirement_id}")
def delete_requirement(requirement_id: int) -> dict:
    with connect() as conn:
        existing = conn.execute("select id, title from requirements where id = ?", (requirement_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="requirement not found")
        test_point_count = conn.execute(
            "delete from test_points where requirement_id = ?", (requirement_id,)
        ).rowcount
        draft_count = conn.execute(
            "delete from case_drafts where requirement_id = ?", (requirement_id,)
        ).rowcount
        conn.execute("delete from requirements where id = ?", (requirement_id,))
    return {
        "status": "deleted",
        "id": requirement_id,
        "title": existing["title"],
        "deleted_test_points": test_point_count,
        "deleted_case_drafts": draft_count,
    }


@app.get("/api/test-points")
def test_points(status: str | None = None) -> list[dict]:
    with connect() as conn:
        rows = conn.execute(
            """
            select tp.*, r.title as requirement_title
            from test_points tp
            left join requirements r on r.id = tp.requirement_id
            order by coalesce(tp.sort_order, tp.id), tp.id
            """
        ).fetchall()
    points = rows_to_dicts(rows)
    if status == "confirmed":
        return [point for point in points if is_confirmed_status(str(point.get("status", "")))]
    return points


@app.post("/api/test-points")
def create_test_point(payload: TestPointCreateRequest) -> dict:
    if not payload.name.strip():
        raise HTTPException(status_code=400, detail="test point name is required")
    now = utc_now()
    requirement_id = payload.requirement_id or ensure_manual_requirement()
    module = payload.module.strip() or load_requirement_title(requirement_id) or "未归属模块"
    sort_order = payload.sort_order if payload.sort_order is not None else next_test_point_sort_order(payload.parent_id)
    with connect() as conn:
        cur = conn.execute(
            """
            insert into test_points(
              requirement_id, parent_id, name, category, priority, status, description,
              module, source, sort_order, created_at, updated_at
            )
            values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                requirement_id,
                payload.parent_id,
                payload.name.strip(),
                payload.category.strip() or "功能",
                payload.priority.strip() or "P1",
                payload.status.strip() or "已确认",
                payload.description.strip(),
                module,
                payload.source.strip() or "manual",
                sort_order,
                now,
                now,
            ),
        )
        test_point_id = cur.lastrowid
    return {"id": test_point_id, "requirement_id": requirement_id}


@app.post("/api/requirements/analyze")
def analyze_requirement(payload: RequirementRequest) -> dict:
    try:
        result = ai_service.generate_test_points(payload.document, get_ai_settings(mask_key=False))
    except AIConfigurationError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except AIProviderError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc

    now = utc_now()
    project_id = resolve_requirement_project_id(payload.project_id)
    with connect() as conn:
        cur = conn.execute(
            """
            insert into requirements(title, document, status, analysis_summary, risk_summary, case_count, project_id, created_at, updated_at)
            values (?, ?, 'analyzed', ?, ?, ?, ?, ?, ?)
            """,
            (
                payload.title,
                payload.document,
                result["analysis_summary"],
                result["risk_summary"],
                result["case_count"],
                project_id,
                now,
                now,
            ),
        )
        requirement_id = cur.lastrowid
        for index, point in enumerate(result["test_points"], start=1):
            conn.execute(
                """
                insert into test_points(
                  requirement_id, parent_id, name, category, priority, status, description,
                  module, source, sort_order, created_at, updated_at
                )
                values (?, null, ?, ?, ?, ?, ?, ?, 'ai_generated', ?, ?, ?)
                """,
                (
                    requirement_id,
                    point.name,
                    point.category,
                    point.priority,
                    point.status,
                    point.description,
                    payload.title,
                    index,
                    now,
                    now,
                ),
            )

    return load_requirement_detail(requirement_id) | {"provider": get_ai_settings(mask_key=False).get("provider", ai_service.provider)}


def _spec_case_summary(cases: list[dict]) -> tuple[str, str]:
    count = len(cases)
    priority_count: dict[str, int] = {}
    type_count: dict[str, int] = {}
    for case in cases:
        priority = str(case.get("priority") or "P?").strip() or "P?"
        case_type = str(case.get("type") or "未分类").strip() or "未分类"
        priority_count[priority] = priority_count.get(priority, 0) + 1
        type_count[case_type] = type_count.get(case_type, 0) + 1
    summary = f"已按《功能测试用例规范》生成 {count} 条测试用例草稿"
    if priority_count:
        summary += "，" + "，".join(f"{key} {value} 条" for key, value in sorted(priority_count.items()))
    risk = "请重点复核前置条件、测试数据和断言字段是否完整。"
    if type_count:
        risk += " 当前覆盖类型：" + "，".join(f"{key} {value} 条" for key, value in sorted(type_count.items()))
    return summary, risk


def _save_spec_case_drafts(requirement_id: int, cases: list[dict], context_info: dict | None = None) -> list[dict]:
    now = utc_now()
    saved: list[dict] = []
    with connect() as conn:
        for case in cases:
            title = str(case.get("title") or case.get("id") or "未命名用例").strip()[:200]
            # P1 · 增量：context_info 写到 case_drafts.yaml 顶层（ARCH §3.2）
            case_payload: dict = dict(case)
            if context_info:
                case_payload["context_info"] = context_info
            yaml_text = yaml.safe_dump(case_payload, allow_unicode=True, sort_keys=False)
            cur = conn.execute(
                """
                insert into case_drafts(requirement_id, title, yaml, status, created_at, updated_at, template)
                values (?, ?, ?, 'draft', ?, ?, 'spec')
                """,
                (requirement_id, title, yaml_text, now, now),
            )
            saved.append({"draft_id": cur.lastrowid, "title": title})
    return saved


@app.patch("/api/test-points/reorder")
def update_test_points_order(payload: TestPointReorderRequest) -> dict[str, int]:
    if not payload.updates:
        raise HTTPException(status_code=400, detail="no test point order updates")

    now = utc_now()
    with connect() as conn:
        existing_ids = {
            row["id"]
            for row in conn.execute(
                f"select id from test_points where id in ({','.join('?' for _ in payload.updates)})",
                [item.id for item in payload.updates],
            ).fetchall()
        }
        missing = [item.id for item in payload.updates if item.id not in existing_ids]
        if missing:
            raise HTTPException(status_code=404, detail=f"test points not found: {missing}")

        for item in payload.updates:
            fields = ["parent_id = ?", "sort_order = ?", "updated_at = ?"]
            values: list[object] = [item.parent_id, item.sort_order, now]
            if item.module is not None:
                fields.append("module = ?")
                values.append(item.module)
            if item.category is not None:
                fields.append("category = ?")
                values.append(item.category)
            values.append(item.id)
            conn.execute(f"update test_points set {', '.join(fields)} where id = ?", values)
    return {"updated": len(payload.updates)}


@app.patch("/api/test-points/{test_point_id}")
def update_test_point(test_point_id: int, payload: TestPointPatchRequest) -> dict:
    values = payload.model_dump(exclude_unset=True)
    allowed = ["parent_id", "name", "category", "priority", "status", "description", "module", "source", "sort_order"]
    keys = [key for key in allowed if key in values]
    if not keys:
        raise HTTPException(status_code=400, detail="no fields to update")

    with connect() as conn:
        row = conn.execute("select requirement_id from test_points where id = ?", (test_point_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="test point not found")
        conn.execute(
            f"update test_points set {', '.join(f'{key} = ?' for key in keys)}, updated_at = ? where id = ?",
            [values[key] for key in keys] + [utc_now(), test_point_id],
        )
        requirement_id = row["requirement_id"]
    return requirement_detail(requirement_id)


@app.delete("/api/test-points/{test_point_id}")
def delete_test_point(test_point_id: int) -> dict[str, str]:
    with connect() as conn:
        row = conn.execute("select id from test_points where id = ?", (test_point_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="test point not found")
        conn.execute("delete from test_points where id = ?", (test_point_id,))
    return {"status": "deleted"}


@app.post("/api/test-points/generate-cases")
def generate_cases_from_test_points(payload: SelectedTestPointsRequest) -> dict:
    ids = sorted(set(payload.test_point_ids))
    if not ids:
        raise HTTPException(status_code=400, detail="please select at least one test point")
    placeholders = ",".join("?" for _ in ids)
    with connect() as conn:
        rows = conn.execute(f"select * from test_points where id in ({placeholders}) order by coalesce(sort_order, id), id", ids).fetchall()
    points = rows_to_dicts(rows)
    if len(points) != len(ids):
        raise HTTPException(status_code=404, detail="some test points were not found")

    try:
        yaml_text = ai_service.generate_cases(
            points,
            template=payload.template,
            title=payload.title,
            settings=get_ai_settings(mask_key=False),
            generator=payload.generator,
        )
    except AIConfigurationError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except AIProviderError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    now = utc_now()
    requirement_id = ensure_generated_requirement(points)
    title = payload.title or draft_title_for_template(payload.template)
    with connect() as conn:
        cur = conn.execute(
            """
            insert into case_drafts(
              requirement_id, title, yaml, status, created_at, updated_at,
              template, source_test_point_ids
            )
            values (?, ?, ?, 'draft', ?, ?, ?, ?)
            """,
            (requirement_id, title, yaml_text, now, now, payload.template, json.dumps(ids)),
        )
        draft_id = cur.lastrowid
    return {"draft_id": draft_id, "yaml": yaml_text, "status": "draft"}


@app.post("/api/requirements/{requirement_id}/generate-cases")
def generate_requirement_cases(requirement_id: int) -> dict:
    detail = load_requirement_detail(requirement_id)
    if not detail:
        raise HTTPException(status_code=404, detail="requirement not found")

    try:
        yaml_text = ai_service.generate_cases(detail["test_points"], settings=get_ai_settings(mask_key=False), generator="rule")
    except AIProviderError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    now = utc_now()
    with connect() as conn:
        cur = conn.execute(
            """
            insert into case_drafts(requirement_id, title, yaml, status, created_at, updated_at, template, source_test_point_ids)
            values (?, ?, ?, 'draft', ?, ?, 'functional', ?)
            """,
            (requirement_id, f"{detail['requirement']['title']} - YAML 草稿", yaml_text, now, now, json.dumps([point["id"] for point in detail["test_points"]])),
        )
        draft_id = cur.lastrowid
    return {"draft_id": draft_id, "yaml": yaml_text, "status": "draft"}


def _load_spec_text() -> str:
    if not SPEC_FILE.exists():
        raise HTTPException(status_code=500, detail=f"spec file not found: {SPEC_FILE}")
    return SPEC_FILE.read_text(encoding="utf-8")


@app.post("/api/requirements/analyze-spec")
def analyze_requirement_spec(payload: RequirementRequest) -> dict:
    generation_id = (payload.generation_id or "").strip()
    if generation_id and not generation_control.begin(generation_id):
        generation_control.finish(generation_id)
        raise HTTPException(status_code=409, detail="generation stopped before completion")
    standard_text = _load_spec_text()
    # P0 · 增量：按 project_id 查 project_profiles（已有函数 db.get_project_profile），拿不到用空 dict
    project_id = resolve_requirement_project_id(payload.project_id)
    project = get_project_profile(project_id) if project_id else None
    project_for_prompt: dict = project or {}
    context_info = payload.context_info or None
    try:
        try:
            result = ai_service.generate_test_cases_spec(
                payload.document,
                standard_text,
                get_ai_settings(mask_key=False),
                project=project_for_prompt or None,
                context_info=context_info,
            )
        except AIConfigurationError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except AIProviderError as exc:
            raise HTTPException(status_code=502, detail=str(exc)) from exc

        if generation_id and not generation_control.claim_persistence(generation_id):
            raise HTTPException(status_code=409, detail="generation stopped before completion")

        cases = result.get("cases") or []
        analysis_summary, risk_summary = _spec_case_summary(cases)
        now = utc_now()
        with connect() as conn:
            cur = conn.execute(
                """
                insert into requirements(title, document, status, analysis_summary, risk_summary, case_count, project_id, created_at, updated_at)
                values (?, ?, 'analyzed', ?, ?, ?, ?, ?, ?)
                """,
                (
                    payload.title,
                    payload.document,
                    analysis_summary,
                    risk_summary,
                    len(cases),
                    project_id,
                    now,
                    now,
                ),
            )
            requirement_id = cur.lastrowid
        _save_spec_case_drafts(requirement_id, cases, context_info=context_info)
        detail = load_requirement_detail(requirement_id)
        if not detail:
            raise HTTPException(status_code=500, detail="failed to load generated requirement")
        detail["provider"] = get_ai_settings(mask_key=False).get("provider")
        detail["requirement"]["analysis_summary"] = analysis_summary
        detail["requirement"]["risk_summary"] = risk_summary
        detail["requirement"]["case_count"] = len(cases)
        detail["generated_cases"] = len(cases)
        return detail
    finally:
        if generation_id:
            generation_control.finish(generation_id)


@app.post("/api/requirements/generations/{generation_id}/stop")
def stop_requirement_generation(generation_id: str) -> dict:
    cancelled = generation_control.request_stop(generation_id)
    return {
        "status": "cancellation_requested" if cancelled else "persistence_started",
        "generation_id": generation_id,
    }


@app.post("/api/requirements/{requirement_id}/generate-spec-cases")
def generate_spec_cases(requirement_id: int) -> dict:
    detail = load_requirement_detail(requirement_id)
    if not detail:
        raise HTTPException(status_code=404, detail="requirement not found")
    standard_text = _load_spec_text()
    try:
        result = ai_service.generate_test_cases_spec(
            detail["requirement"]["document"],
            standard_text,
            get_ai_settings(mask_key=False),
        )
    except AIConfigurationError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except AIProviderError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    cases = result.get("cases") or []
    now = utc_now()
    saved: list[dict] = []
    with connect() as conn:
        for case in cases:
            title = str(case.get("title") or case.get("id") or "未命名用例").strip()[:200]
            yaml_text = yaml.safe_dump(case, allow_unicode=True, sort_keys=False)
            cur = conn.execute(
                """
                insert into case_drafts(requirement_id, title, yaml, status, created_at, updated_at, template)
                values (?, ?, ?, 'draft', ?, ?, 'spec')
                """,
                (requirement_id, title, yaml_text, now, now),
            )
            saved.append({"draft_id": cur.lastrowid, "title": title})
    return {
        "requirement_id": requirement_id,
        "count": len(saved),
        "drafts": saved,
    }


def _parse_case_draft_payload(draft: dict) -> dict:
    try:
        data = yaml.safe_load(draft.get("yaml") or "") or {}
    except yaml.YAMLError:
        data = {}
    if not isinstance(data, dict):
        data = {}
    return data


def _case_yaml_text(requirement: dict, drafts: list[dict]) -> str:
    out = {
        "requirement": {
            "id": requirement.get("id"),
            "title": requirement.get("title"),
            "document": requirement.get("document"),
        },
        "cases": [_parse_case_draft_payload(d) for d in drafts],
    }
    return yaml.safe_dump(out, allow_unicode=True, sort_keys=False)


def _case_xlsx_bytes(requirement: dict, drafts: list[dict]) -> bytes:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "TestCases"
    headers = [
        "id", "title", "module", "priority", "type",
        "precondition", "test_data", "steps", "expected",
        "requirement_id", "automation", "author/date", "note",
    ]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="E6F0FF")
    header_font = Font(bold=True, color="1F2A44")
    for col_idx in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")
    for draft in drafts:
        case = _parse_case_draft_payload(draft)
        author = str(case.get("author") or "").strip()
        date_ = str(case.get("date") or "").strip()
        author_date = f"{author} / {date_}" if author or date_ else ""
        steps = case.get("steps")
        expected = case.get("expected")
        steps_text = "\n".join(steps) if isinstance(steps, list) else (str(steps) if steps is not None else "")
        expected_text = "\n".join(expected) if isinstance(expected, list) else (str(expected) if expected is not None else "")
        row = [
            str(case.get("id") or ""),
            str(case.get("title") or draft.get("title") or ""),
            str(case.get("module") or ""),
            str(case.get("priority") or ""),
            str(case.get("type") or ""),
            str(case.get("precondition") or ""),
            str(case.get("test_data") or ""),
            steps_text,
            expected_text,
            str(case.get("requirement_id") or ""),
            str(case.get("automation") or ""),
            author_date,
            str(case.get("note") or ""),
        ]
        sheet.append(row)
    widths = [16, 36, 14, 8, 8, 30, 26, 36, 36, 14, 10, 18, 24]
    for col_idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A" + chr(64 + col_idx - 26)].width = width
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = wrap
    summary = wb.create_sheet("Summary")
    summary.append(["field", "value"])
    summary.append(["requirement_id", requirement.get("id")])
    summary.append(["requirement_title", requirement.get("title")])
    summary.append(["case_count", len(drafts)])
    priority_count: dict[str, int] = {}
    type_count: dict[str, int] = {}
    for draft in drafts:
        case = _parse_case_draft_payload(draft)
        p = str(case.get("priority") or "P?").strip() or "P?"
        t = str(case.get("type") or "").strip() or "未分类"
        priority_count[p] = priority_count.get(p, 0) + 1
        type_count[t] = type_count.get(t, 0) + 1
    summary.append([])
    summary.append(["priority", "count"])
    for p, c in sorted(priority_count.items()):
        summary.append([p, c])
    summary.append([])
    summary.append(["type", "count"])
    for t, c in sorted(type_count.items()):
        summary.append([t, c])
    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 36
    for col_idx in (1, 2):
        cell = summary.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _safe_filename(prefix: str, requirement: dict) -> str:
    title = str(requirement.get("title") or "requirement").strip()
    slug = re.sub(r"[^\w\-\u4e00-\u9fa5]+", "_", title)[:40] or "requirement"
    ts = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    return f"{prefix}-{slug}-{requirement.get('id')}-{ts}"


def _download_headers(filename: str) -> dict[str, str]:
    ascii_name = re.sub(r"[^A-Za-z0-9._-]+", "_", filename).strip("._") or "download"
    return {
        "Content-Disposition": f"attachment; filename={ascii_name}; filename*=UTF-8''{quote(filename)}",
    }


@app.get("/api/requirements/{requirement_id}/export")
def export_requirement_cases(requirement_id: int, format: str = "xlsx"):
    detail = load_requirement_detail(requirement_id)
    if not detail:
        raise HTTPException(status_code=404, detail="requirement not found")
    drafts = list(detail.get("drafts") or [])
    if not drafts:
        raise HTTPException(status_code=400, detail="requirement has no case drafts to export")
    requirement = detail["requirement"]
    if format == "yaml":
        content = _case_yaml_text(requirement, drafts).encode("utf-8")
        filename = f"{_safe_filename('cases', requirement)}.yaml"
        return Response(
            content=content,
            media_type="application/x-yaml; charset=utf-8",
            headers=_download_headers(filename),
        )
    if format == "xlsx":
        content = _case_xlsx_bytes(requirement, drafts)
        filename = f"{_safe_filename('cases', requirement)}.xlsx"
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=_download_headers(filename),
        )
    raise HTTPException(status_code=400, detail=f"unsupported format: {format}")


@app.post("/api/cases/generate")
def generate_cases(payload: GenerateCasesRequest) -> dict[str, str]:
    return {"yaml": ai_service.generate_cases(payload.test_points)}


@app.get("/api/case-drafts")
def case_drafts() -> list[dict]:
    with connect() as conn:
        rows = conn.execute(
            """
            select cd.*, r.title as requirement_title
            from case_drafts cd
            left join requirements r on r.id = cd.requirement_id
            order by cd.created_at desc, cd.id desc
            limit 100
            """
        ).fetchall()
    return [normalize_case_draft(row) for row in rows_to_dicts(rows)]


def build_blank_case_yaml(title: str) -> str:
    payload = {
        "id": "TC-ICM-DRAFT",
        "title": title or "新增用例草稿",
        "status": "draft",
        "type": "功能",
        "priority": "P1",
        "author": "AI",
        "precondition": ["请补充前置条件"],
        "steps": ["请补充执行步骤"],
        "expected_results": ["请补充预期结果"],
        "automation_asset": {
            "operation_steps": ["请补充自动化操作步骤"],
            "selectors": {"todo": ["请补充选择器"]},
            "input_values": {},
            "assertions": ["请补充断言"],
        },
    }
    return yaml.safe_dump(payload, allow_unicode=True, sort_keys=False)


@app.post("/api/case-drafts")
def create_case_draft(payload: CaseDraftCreateRequest) -> dict:
    title = (payload.title or "新增用例草稿").strip() or "新增用例草稿"
    requirement_id = payload.requirement_id or ensure_manual_requirement()
    yaml_text = payload.yaml or build_blank_case_yaml(title)
    now = utc_now()
    with connect() as conn:
        requirement = conn.execute("select id from requirements where id = ?", (requirement_id,)).fetchone()
        if not requirement:
            raise HTTPException(status_code=404, detail="requirement not found")
        cur = conn.execute(
            """
            insert into case_drafts(requirement_id, title, yaml, status, created_at, updated_at, template)
            values (?, ?, ?, 'draft', ?, ?, ?)
            """,
            (requirement_id, title, yaml_text, now, now, payload.template or "manual"),
        )
        draft_id = int(cur.lastrowid)
    return case_draft_detail(draft_id)


def _safe_unlink(path: Path, allowed_root: Path) -> bool:
    target = path.resolve()
    root = allowed_root.resolve()
    if not target.exists():
        return False
    if root not in target.parents and target != root:
        raise HTTPException(status_code=400, detail=f"refuse to delete outside allowed root: {path}")
    if target.is_dir():
        raise HTTPException(status_code=400, detail=f"refuse to delete directory: {path}")
    target.unlink()
    return True


def delete_case_draft_assets(draft_id: int) -> dict:
    draft = load_case_draft(draft_id)
    if not draft:
        raise HTTPException(status_code=404, detail="case draft not found")
    deleted_files = 0
    promoted_case_id = str(draft.get("promoted_case_id") or "").strip()
    promoted_path = str(draft.get("promoted_path") or "").strip()
    if promoted_path:
        deleted_files += int(_safe_unlink(Path(promoted_path), TEST_CASE_DIR))
    elif promoted_case_id:
        try:
            deleted_files += int(_safe_unlink(find_case_yaml(promoted_case_id), TEST_CASE_DIR))
        except HTTPException as exc:
            if exc.status_code != 404:
                raise
    if promoted_case_id:
        deleted_files += int(_safe_unlink(compute_target_path(promoted_case_id), ROOT / "runner" / "flows"))
    with connect() as conn:
        deleted_rows = conn.execute("delete from case_drafts where id = ?", (draft_id,)).rowcount
    return {
        "draft_id": draft_id,
        "promoted_case_id": promoted_case_id or None,
        "deleted_files": deleted_files,
        "deleted_rows": deleted_rows,
    }


@app.post("/api/case-drafts/batch-delete")
def batch_delete_case_drafts(payload: CaseDraftBatchDeleteRequest) -> dict:
    results = []
    for draft_id in dict.fromkeys(payload.draft_ids):
        try:
            results.append({"ok": True, **delete_case_draft_assets(int(draft_id))})
        except HTTPException as exc:
            results.append({"ok": False, "draft_id": draft_id, "error": exc.detail})
    return {
        "deleted": sum(1 for item in results if item["ok"]),
        "failed": sum(1 for item in results if not item["ok"]),
        "results": results,
    }


@app.delete("/api/case-drafts/{draft_id}")
def delete_case_draft(draft_id: int) -> dict:
    return {"ok": True, **delete_case_draft_assets(draft_id)}


@app.get("/api/case-drafts/{draft_id}")
def case_draft_detail(draft_id: int) -> dict:
    draft = load_case_draft(draft_id)
    if not draft:
        raise HTTPException(status_code=404, detail="case draft not found")
    return draft


@app.patch("/api/case-drafts/{draft_id}")
def update_case_draft(draft_id: int, payload: CaseDraftPatchRequest) -> dict:
    values = payload.model_dump(exclude_unset=True)
    allowed = ["title", "yaml", "status"]
    keys = [key for key in allowed if key in values]
    if not keys:
        raise HTTPException(status_code=400, detail="no fields to update")
    with connect() as conn:
        row = conn.execute("select id from case_drafts where id = ?", (draft_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="case draft not found")
        conn.execute(
            f"update case_drafts set {', '.join(f'{key} = ?' for key in keys)}, updated_at = ? where id = ?",
            [values[key] for key in keys] + [utc_now(), draft_id],
        )
    return case_draft_detail(draft_id)


@app.post("/api/case-drafts/{draft_id}/validate")
def validate_case_draft(draft_id: int, payload: ValidateDraftRequest | None = None) -> dict:
    draft = load_case_draft(draft_id)
    if not draft:
        raise HTTPException(status_code=404, detail="case draft not found")
    yaml_text = str(payload.yaml if payload and payload.yaml is not None else draft["yaml"])
    if payload and payload.case_id:
        yaml_text = replace_yaml_case_id(yaml_text, normalize_case_id(payload.case_id))
    return {"draft_id": draft_id, **validate_case_yaml(yaml_text)}


@app.post("/api/case-drafts/{draft_id}/promote")
def promote_case_draft(draft_id: int, payload: PromoteDraftRequest) -> dict:
    draft = load_case_draft(draft_id)
    if not draft:
        raise HTTPException(status_code=404, detail="case draft not found")
    case_id = normalize_case_id(payload.case_id)
    filename = normalize_case_filename(payload.filename, case_id)
    target = TEST_CASE_DIR / filename
    if target.exists():
        raise HTTPException(status_code=409, detail=f"case file already exists: {filename}")
    yaml_text = replace_yaml_case_id(str(draft["yaml"]), case_id)
    validation = validate_case_yaml(yaml_text)
    if not validation["valid"]:
        raise HTTPException(status_code=400, detail=f"YAML validation failed: {'; '.join(validation['errors'])}")
    TEST_CASE_DIR.mkdir(parents=True, exist_ok=True)
    target.write_text(yaml_text, encoding="utf-8")
    with connect() as conn:
        conn.execute(
            """
            update case_drafts
            set status = 'promoted', promoted_case_id = ?, promoted_path = ?, updated_at = ?
            where id = ?
            """,
            (case_id, str(target), utc_now(), draft_id),
        )
    return case_draft_detail(draft_id)


@app.post("/api/case-drafts/{draft_id}/run")
def run_case_draft(draft_id: int) -> dict[str, str | None]:
    try:
        return worker.enqueue("run-draft", draft_id=draft_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/runs")
def create_run(payload: RunRequest) -> dict[str, str | None]:
    try:
        return worker.enqueue(payload.mode, payload.case_id, payload.draft_id, payload.case_ids)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/runs")
def runs() -> list[dict]:
    with connect() as conn:
        # AI测试任务队列只展示测试执行类任务。
        # 元素知识库刷新属于平台维护任务，走独立的
        # /api/element-knowledge/refresh-tasks 接口，不应该污染测试队列。
        placeholders = ",".join("?" for _ in AI_TEST_TASK_MODES)
        rows = conn.execute(
            f"""
            select *
            from run_tasks
            where mode in ({placeholders})
            order by created_at desc
            limit 50
            """,
            tuple(AI_TEST_TASK_MODES),
        ).fetchall()
    tasks = rows_to_dicts(rows)
    return [{**task, "summary": summarize_run_task(task)} for task in tasks]


@app.post("/api/runs/{run_id}/stop")
def stop_run(run_id: str) -> dict[str, str]:
    try:
        return worker.stop(run_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


def _delete_run_artifacts(run_id: str, task_dict: dict) -> dict[str, int]:
    deleted_files = 0
    deleted_dirs = 0
    deleted_rows = 0
    report_path = str(task_dict.get("report_path") or "").strip()
    file_candidates = [
        Path(report_path) if report_path else None,
        REPORT_DIR / f"{run_id}.md",
        (ROOT / "reports" / "step-details" / f"{run_id}.json"),
    ]
    dir_candidates = [
        ROOT / "reports" / "agent-explore" / run_id,
        DRAFT_RUN_DIR / run_id,
        SCREENSHOTS_RUNS_DIR / run_id,
        EVIDENCE_ROOT / run_id,
        TRACE_ROOT / run_id,
    ]
    for path in file_candidates:
        if path and path.exists() and path.is_file():
            path.unlink()
            deleted_files += 1
    for path in dir_candidates:
        if path.exists() and path.is_dir():
            shutil.rmtree(path)
            deleted_dirs += 1
    with connect() as conn:
        deleted_rows += conn.execute("delete from run_logs where run_id = ?", (run_id,)).rowcount
        deleted_rows += conn.execute("delete from run_tasks where id = ?", (run_id,)).rowcount
    return {"deleted_files": deleted_files, "deleted_dirs": deleted_dirs, "deleted_rows": deleted_rows}


@app.delete("/api/runs/{run_id}")
def delete_run(run_id: str) -> dict:
    with connect() as conn:
        task = conn.execute("select * from run_tasks where id = ?", (run_id,)).fetchone()
    if not task:
        raise HTTPException(status_code=404, detail="run not found")
    result = _delete_run_artifacts(run_id, dict(task))
    return {"ok": True, "run_id": run_id, **result}


@app.get("/api/runs/{run_id}")
def run_detail(run_id: str) -> dict:
    with connect() as conn:
        task = conn.execute("select * from run_tasks where id = ?", (run_id,)).fetchone()
        logs = conn.execute("select * from run_logs where run_id = ? order by id", (run_id,)).fetchall()
    if not task:
        raise HTTPException(status_code=404, detail="run not found")

    report = ""
    screenshots: list[dict[str, str]] = []
    task_dict = dict(task)
    children = list_batch_child_reports(run_id, case_ids=_batch_case_ids(task_dict)) if task_dict.get("mode") == "run-batch" else []
    if task_dict.get("report_path"):
        try:
            report = read_report(run_id)
            screenshots = parse_report(report)["screenshots"]
        except FileNotFoundError:
            report = ""
    if not screenshots and task_dict.get("case_id") and str(task_dict.get("status") or "") not in {"queued", "running"}:
        screenshots = [screenshot_payload(task_dict["case_id"], path) for path in latest_screenshots(task_dict["case_id"])]
    agent_explore = load_agent_explore_artifacts(run_id) if task_dict.get("mode") == "agent-explore" else None
    log_dicts = rows_to_dicts(logs)
    if not log_dicts and task_dict.get("mode") == "agent-explore":
        log_dicts = _synthetic_logs_from_evidence(run_id)
    evidence_lines = evidence_log_lines(run_id)
    return {
        "task": task_dict,
        "summary": summarize_run_task(task_dict),
        "logs": log_dicts,
        "children": children,
        "report": report,
        "screenshots": screenshots,
        "evidence": evidence_summary(run_id),
        "agent_explore": agent_explore,
        "analysis": ai_service.analyze_run_report(report, screenshots, [row["line"] for row in log_dicts] + evidence_lines) if report else None,
    }


def _normalized_mode(mode: str) -> str:
    return "agent" if mode == "agent-explore" else "worker"


def _normalized_status(status: str) -> str:
    if status == "passed":
        return "completed"
    return status or "queued"


def _run_screenshots(run_id: str) -> list[dict[str, str]]:
    run_dir = SCREENSHOTS_RUNS_DIR / run_id
    if not run_dir.exists():
        return []
    return [run_screenshot_payload(run_id, str(path)) for path in sorted(run_dir.glob("*.png"))]


def _report_screenshots(task_dict: dict, report_text: str) -> list[dict[str, str]]:
    screenshots: list[dict[str, str]] = []
    if report_text:
        screenshots = parse_report(report_text)["screenshots"]
    if not screenshots:
        screenshots = _run_screenshots(str(task_dict.get("id") or ""))
    if not screenshots and task_dict.get("case_id") and str(task_dict.get("status") or "") not in {"queued", "running"}:
        screenshots = [screenshot_payload(task_dict["case_id"], path) for path in latest_screenshots(task_dict["case_id"])]
    return screenshots


def _batch_case_ids(task_dict: dict[str, Any]) -> list[str] | None:
    raw = task_dict.get("batch_case_ids")
    if not raw:
        return None
    try:
        values = json.loads(str(raw))
    except json.JSONDecodeError:
        return None
    return values if isinstance(values, list) else None


def _batch_summary_report(run_id: str, children: list[dict[str, Any]], case_ids: list[str] | None = None) -> str:
    completed = [child for child in children if child.get("run_id")]
    passed = sum(1 for child in completed if child.get("status") == "passed")
    failed = sum(1 for child in completed if child.get("status") == "failed")
    lines = [
        f"# {run_id}",
        "",
        f"- case name: {'Custom batch (' + str(len(children)) + ' cases)' if case_ids else 'Batch 001-012'}",
        "- environment: icm-internal",
        f"- status: {'failed' if failed else ('passed' if completed else 'running')}",
        f"- total: {len(completed)}",
        f"- passed: {passed}",
        f"- failed: {failed}",
        "- child reports:",
    ]
    for child in children:
        report_path = child.get("report_path") or "none"
        lines.append(f"  - {child.get('case_id')}: {child.get('status')} / {report_path}")
    lines.append("- screenshot paths:")
    return "\n".join(lines)


def _build_batch_steps(children: list[dict[str, Any]]) -> list[dict[str, Any]]:
    steps: list[dict[str, Any]] = []
    for child in children:
        status = str(child.get("status") or "pending")
        run_id = str(child.get("run_id") or "")
        step_index = int(child.get("order") or len(steps) + 1)
        case_id = str(child.get("case_id") or f"TC-ICM-{step_index:03d}")
        screenshot_count = int(child.get("screenshot_count") or 0)
        report_path = child.get("report_path")
        summary = "尚未开始"
        if status == "passed":
            summary = f"已通过，截图 {screenshot_count} 张"
        elif status == "failed":
            summary = f"执行失败，截图 {screenshot_count} 张"
        elif status == "running":
            summary = "正在执行"
        steps.append(
            {
                "step_index": step_index,
                "step_code": run_id or f"batch-{case_id.lower()}",
                "title": f"{case_id} - {child.get('case_name') or case_id}",
                "status": _normalized_status(status),
                "summary": summary,
                "ai_analysis": "",
                "error_message": "子用例失败" if status == "failed" else "",
                "screenshot_url": "",
                "screenshot_name": "",
                "network_logs": [],
                "command_output": [f"[run_id] {run_id or '--'}", f"[report] {report_path or '--'}"],
            }
        )
    return steps


_CASE_STEP_STOPWORDS = {
    "点击",
    "单击",
    "填写",
    "输入",
    "选择",
    "确认",
    "验证",
    "打开",
    "访问",
    "页面",
    "页面中",
    "页面内",
    "按钮",
    "下拉",
    "弹窗",
    "弹出",
    "依次",
    "当前",
    "指定",
    "然后",
    "使用",
    "系统",
    "列表",
    "信息",
    "成功",
    "新增",
    "设备",
}


def _clean_case_step_text(value: str) -> str:
    text = str(value or "").strip()
    text = re.sub(r"^\s*[-*]?\s*\d+[.、:\-]?\s*", "", text)
    return text.strip()


def _extract_case_step_keywords(step_text: str) -> list[str]:
    tokens = re.findall(r"[\u4e00-\u9fffA-Za-z0-9#/_-]{2,}", step_text)
    keywords: list[str] = []
    seen: set[str] = set()
    for token in tokens:
        normalized = token.strip()
        if not normalized or normalized in seen or normalized in _CASE_STEP_STOPWORDS:
            continue
        seen.add(normalized)
        keywords.append(normalized)
    return keywords


def _agent_target_text(decision: dict, observation: dict) -> str:
    ref = str(decision.get("ref") or "").strip()
    interactives = observation.get("interactives") or []
    for item in interactives:
        if not isinstance(item, dict) or str(item.get("ref") or "") != ref:
            continue
        for key in ("text", "ariaLabel", "placeholder", "selector"):
            value = str(item.get(key) or "").strip()
            if value:
                return value
    return ""


def _agent_step_action_summary(decision: dict, execution: dict, observation: dict) -> str:
    action = str(decision.get("action") or "").strip().lower()
    reason = str(decision.get("reason") or "").strip()
    value = str(decision.get("value") or "").strip()
    key = str(decision.get("key") or "").strip()
    target = _agent_target_text(decision, observation)
    if action == "fill":
        if target:
            return f"填写 {target}" + (f"：{value}" if value else "")
        return f"填写测试数据" + (f"：{value}" if value else "")
    if action == "click":
        return f"点击 {target}" if target else "点击页面控件"
    if action == "wait":
        lowered = reason.lower()
        if "redirect" in lowered and "login" in lowered:
            return "等待登录完成并跳转到目标页面"
        return "等待页面状态更新"
    if action == "press":
        return f"按下 {key}" if key else "按下指定按键"
    if action == "goto":
        url = str(decision.get("url") or "").strip()
        return f"打开 {url}" if url else "打开目标页面"
    if action == "finish":
        return reason or "完成探索验证"
    if action == "fail":
        lowered = reason.lower()
        if "instead of the expected" in lowered and "redirect" in lowered:
            return "登录后跳转地址与 redirect 参数预期不一致"
        return reason or str(execution.get("error") or "探索失败")
    return reason or str(execution.get("result") or action or "执行步骤")


def _load_agent_case_payload(run_id: str, task_dict: dict, trace: dict) -> dict:
    try:
        yaml_text, _ = _load_case_source_for_agent_run(run_id, task_dict, trace)
    except HTTPException:
        return {}
    try:
        payload = yaml.safe_load(yaml_text) or {}
    except yaml.YAMLError:
        return {}
    return payload if isinstance(payload, dict) else {}


def _load_agent_case_steps(run_id: str, task_dict: dict, trace: dict) -> list[str]:
    payload = _load_agent_case_payload(run_id, task_dict, trace)
    steps = payload.get("steps")
    if not isinstance(steps, list):
        return []
    return [_clean_case_step_text(step) for step in steps if _clean_case_step_text(step)]


def _load_agent_expected_results(run_id: str, task_dict: dict, trace: dict) -> list[str]:
    payload = _load_agent_case_payload(run_id, task_dict, trace)
    expected = payload.get("expected_results")
    if expected is None:
        expected = payload.get("expected")
    if not isinstance(expected, list):
        return []
    return [str(item).strip() for item in expected if _clean_case_step_text(item)]


def _match_case_step_index(case_steps: list[str], decision: dict, observation: dict, current_index: int) -> int | None:
    if not case_steps:
        return None
    haystack_parts = [
        str(decision.get("action") or ""),
        str(decision.get("reason") or ""),
        str(decision.get("value") or ""),
        str(decision.get("url") or ""),
        _agent_target_text(decision, observation),
        " ".join(str(item) for item in (observation.get("visibleText") or [])),
        str(observation.get("url") or ""),
    ]
    haystack = " ".join(part for part in haystack_parts if part).lower()
    action = str(decision.get("action") or "").lower()
    target = _agent_target_text(decision, observation)
    if action == "fill":
        if any(word in target for word in ("账号", "用户名", "密码")) or any(word in haystack for word in ("test", "123456", "password", "username", "账号", "密码")):
            for index, step_text in enumerate(case_steps):
                if any(word in step_text for word in ("输入", "填写", "账号", "密码")):
                    return index
    if action == "click":
        if any(word in haystack for word in ("登录", "submit login")):
            for index, step_text in enumerate(case_steps):
                if "点击" in step_text and "登录" in step_text:
                    return index
    if action == "wait" and current_index >= 0:
        return min(current_index, len(case_steps) - 1)

    best_index = current_index if 0 <= current_index < len(case_steps) else 0
    best_score = -1.0
    for index, step_text in enumerate(case_steps):
        score = 0.0
        if index == current_index:
            score += 0.8
        elif index == current_index + 1:
            score += 1.2
        for keyword in _extract_case_step_keywords(step_text):
            if keyword.lower() in haystack:
                score += 2.0
        if action == "fill" and any(word in step_text for word in ("填写", "输入")):
            score += 3.0
        if action == "click" and "点击" in step_text:
            score += 2.5
        if action == "wait" and "等待" in step_text:
            score += 2.0
        if action == "goto" and any(word in step_text for word in ("打开", "访问")):
            score += 3.0
        if "登录" in step_text and action == "goto" and any(word in haystack for word in ("login", "#/login")):
            score += 3.0
        if "icm" in step_text.lower() and "icm" in haystack:
            score += 3.0
        if "设备信息" in step_text and any(word in haystack for word in ("设备信息", "hubble/device")):
            score += 3.0
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _agent_stage_source_map(trace: dict) -> dict[str, list[int]]:
    plan = trace.get("plan") if isinstance(trace, dict) else {}
    plan = plan if isinstance(plan, dict) else {}
    stages = plan.get("stages") if isinstance(plan.get("stages"), list) else []
    source_map: dict[str, list[int]] = {}
    for stage in stages:
        if not isinstance(stage, dict):
            continue
        stage_id = str(stage.get("stage_id") or "").strip()
        if not stage_id:
            continue
        indexes: list[int] = []
        for raw_index in stage.get("source_steps") or []:
            try:
                index = int(raw_index) - 1
            except (TypeError, ValueError):
                continue
            if index >= 0:
                indexes.append(index)
        if indexes:
            source_map[stage_id] = indexes
    return source_map


def _agent_stage_meta_map(trace: dict) -> dict[str, dict]:
    plan = trace.get("plan") if isinstance(trace, dict) else {}
    plan = plan if isinstance(plan, dict) else {}
    stages = plan.get("stages") if isinstance(plan.get("stages"), list) else []
    meta_map: dict[str, dict] = {}
    for stage in stages:
        if not isinstance(stage, dict):
            continue
        stage_id = str(stage.get("stage_id") or "").strip()
        if stage_id:
            meta_map[stage_id] = stage
    return meta_map


def _aggregate_agent_stage_status(step_statuses: list[str]) -> str:
    normalized = [str(status or "").strip().lower() for status in step_statuses if str(status or "").strip()]
    if not normalized:
        return "queued"
    if any(status == "failed" for status in normalized):
        return "failed"
    if any(status == "running" for status in normalized):
        return "running"
    if any(status == "queued" for status in normalized):
        return "queued"
    if all(status == "completed" for status in normalized):
        return "completed"
    return normalized[-1]


def _reconcile_agent_stage_runs(trace: dict, steps: list[dict], stage_runs: list[dict]) -> list[dict]:
    source_map = _agent_stage_source_map(trace)
    if not source_map or not stage_runs:
        return stage_runs
    reconciled: list[dict] = []
    for raw_stage_run in stage_runs:
        stage_run = dict(raw_stage_run or {})
        stage_id = str(stage_run.get("stage_id") or "").strip()
        source_indexes = source_map.get(stage_id) or []
        if source_indexes:
            step_statuses = [
                str((steps[index] or {}).get("status") or "queued")
                for index in source_indexes
                if 0 <= index < len(steps)
            ]
            if step_statuses:
                stage_run["status"] = _aggregate_agent_stage_status(step_statuses)
        reconciled.append(stage_run)
    return reconciled


def _resolve_agent_assertion_title(
    item: dict,
    case_step_title: str,
    expected_results: list[str],
    stage_meta_map: dict[str, dict],
) -> tuple[str, str]:
    if not isinstance(item, dict):
        return case_step_title, ""
    decision = item.get("decision") if isinstance(item.get("decision"), dict) else {}
    action = str(decision.get("action") or "").strip().lower()
    stage_id = str(item.get("stage_id") or decision.get("stage_id") or "").strip()
    strategy = str((stage_meta_map.get(stage_id) or {}).get("strategy") or "").strip().lower()
    if action not in {"assert_text", "finish"} and strategy != "detail_assert":
        return case_step_title, ""
    try:
        local_step = int(item.get("stage_local_step") or 0)
    except (TypeError, ValueError):
        local_step = 0
    if local_step <= 0 or local_step > len(expected_results):
        return case_step_title, ""
    expected_text = expected_results[local_step - 1].strip()
    if not expected_text:
        return case_step_title, ""
    return f"预期结果 {local_step} - {expected_text}", expected_text


def _match_stage_case_step_index(item: dict, source_map: dict[str, list[int]], case_steps: list[str], current_index: int) -> int | None:
    if not source_map or not case_steps or not isinstance(item, dict):
        return None
    decision = item.get("decision") if isinstance(item.get("decision"), dict) else {}
    stage_id = str(item.get("stage_id") or decision.get("stage_id") or "").strip()
    candidates = [index for index in source_map.get(stage_id, []) if 0 <= index < len(case_steps)]
    if not candidates:
        return None
    try:
        local_step = int(item.get("stage_local_step") or 0)
    except (TypeError, ValueError):
        local_step = 0
    if local_step > 0:
        return candidates[min(local_step - 1, len(candidates) - 1)]
    for index in candidates:
        if index > current_index:
            return index
    return candidates[-1]


def _agent_step_screenshot(item: dict, screenshots: list[dict[str, str]], index: int) -> dict[str, str] | None:
    if not screenshots:
        return None
    execution = item.get("execution") if isinstance(item.get("execution"), dict) else {}
    filename = str(item.get("screenshot_name") or execution.get("screenshot_name") or "").strip()
    if filename:
        for screenshot in screenshots:
            if screenshot.get("filename") == filename:
                return screenshot
    return screenshots[min(index - 1, len(screenshots) - 1)]


def _agent_step_command_output(item: dict, case_step_text: str = "") -> list[str]:
    if not isinstance(item, dict):
        return []
    decision = item.get("decision") if isinstance(item.get("decision"), dict) else {}
    execution = item.get("execution") if isinstance(item.get("execution"), dict) else {}
    output: list[str] = []
    action = str(decision.get("action") or "").strip()
    reason = str(decision.get("reason") or "").strip()
    action_text = case_step_text.strip() or reason
    result = str(execution.get("result") or "").strip()
    if action:
        output.append(f"[action] {action_text or action}")
    if result:
        output.append(f"[result] {result}")
    for key in ("value", "url"):
        value = str(decision.get(key) or "").strip()
        if value:
            output.append(f"[{key}] {value}")
    screenshot_name = str(item.get("screenshot_name") or execution.get("screenshot_name") or "").strip()
    if screenshot_name:
        output.append(f"[screenshot] {screenshot_name}")
    error = str(execution.get("error") or "").strip()
    if error:
        output.append(f"[error] {error}")
    return output



def _parse_expected_result_bindings(expected_results: list[str], step_count: int) -> list[dict]:
    bindings: list[dict] = []
    for ordinal, raw_text in enumerate(expected_results, start=1):
        raw_line = str(raw_text or "").strip()
        normalized = _clean_case_step_text(raw_line)
        start_step = ordinal
        end_step = ordinal
        display_text = normalized
        match = re.match(r"^\s*(\d+)(?:\s*-\s*(\d+))?\.\s*(.+?)\s*$", raw_line)
        if match:
            start_step = int(match.group(1))
            end_step = int(match.group(2) or match.group(1))
            display_text = _clean_case_step_text(match.group(3))
        if step_count > 0:
            start_step = min(max(start_step, 1), step_count)
            end_step = min(max(end_step, start_step), step_count)
        bindings.append(
            {
                "ordinal": ordinal,
                "start_step": start_step,
                "end_step": end_step,
                "text": display_text or normalized,
            }
        )
    return bindings


def _agent_assertion_binding(item: dict, expected_bindings: list[dict], stage_meta_map: dict[str, dict]) -> dict | None:
    if not isinstance(item, dict) or not expected_bindings:
        return None
    decision = item.get("decision") if isinstance(item.get("decision"), dict) else {}
    action = str(decision.get("action") or "").strip().lower()
    stage_id = str(item.get("stage_id") or decision.get("stage_id") or "").strip()
    strategy = str((stage_meta_map.get(stage_id) or {}).get("strategy") or "").strip().lower()
    if action not in {"assert_text", "finish"} and strategy != "detail_assert":
        return None
    try:
        local_step = int(item.get("stage_local_step") or 0)
    except (TypeError, ValueError):
        return None
    if local_step <= 0 or local_step > len(expected_bindings):
        return None
    return expected_bindings[local_step - 1]


def _agent_actual_result_text(item: dict, trace: dict) -> str:
    if not isinstance(item, dict):
        return ""
    execution = item.get("execution") if isinstance(item.get("execution"), dict) else {}
    observation = item.get("observation") if isinstance(item.get("observation"), dict) else {}
    error_text = str(execution.get("error") or "").strip()
    if error_text:
        return error_text
    visible_texts: list[str] = []
    for raw_text in observation.get("visibleText") or []:
        text = _clean_case_step_text(raw_text)
        if text and text not in visible_texts:
            visible_texts.append(text)
        if len(visible_texts) >= 3:
            break
    if visible_texts:
        return "；".join(visible_texts)
    interactive_texts: list[str] = []
    for raw_item in observation.get("interactives") or []:
        if not isinstance(raw_item, dict):
            continue
        text = _clean_case_step_text(str(raw_item.get("text") or ""))
        if text and text not in interactive_texts:
            interactive_texts.append(text)
        if len(interactive_texts) >= 2:
            break
    if interactive_texts:
        return "；".join(interactive_texts)
    result_text = str(execution.get("result") or "").strip()
    if result_text:
        return result_text
    final_url = str(observation.get("url") or trace.get("finalUrl") or trace.get("final_url") or "").strip()
    if final_url:
        return final_url
    return ""


def _assertion_status_rank(status: str) -> int:
    if status == "failed":
        return 3
    if status == "completed":
        return 2
    if status == "running":
        return 1
    return 0


def _extract_expected_value_pairs(text: str) -> list[tuple[str, str]]:
    pairs: list[tuple[str, str]] = []
    for match in re.finditer(
        r"(设备名称|IP|状态|用户名|账号|端口|连接类型|设备类型|服务器名称|标题|名称)(?:为|是|显示为|默认为)([^，。；,;]+)",
        text,
    ):
        label = _clean_case_step_text(match.group(1))
        value = _clean_case_step_text(match.group(2))
        if label and value:
            pairs.append((label, value))
    return pairs


def _expected_text_fragments(text: str) -> list[str]:
    fragments: list[str] = []
    for match in re.finditer(r"[\"'“‘【](.+?)[\"'”’】]", text):
        fragment = _clean_case_step_text(match.group(1))
        if fragment and fragment not in fragments:
            fragments.append(fragment)
    if fragments:
        return fragments
    normalized = _clean_case_step_text(text)
    for piece in re.split(r"[，。；,;]|并且|并|且", normalized):
        fragment = piece.strip()
        if len(fragment) >= 2 and fragment not in fragments:
            fragments.append(fragment)
    return fragments[:4]


def _cjk_bigrams(text: str) -> list[str]:
    result: list[str] = []
    cleaned = re.sub(r"[^\u4e00-\u9fff]", "", text)
    for i in range(len(cleaned) - 1):
        result.append(cleaned[i:i + 2])
    return result


def _loose_text_match(expected: str, facts: dict) -> tuple[bool, str]:
    exp = expected.lower().strip()
    haystack = facts["haystack"]
    if not exp:
        return False, ""
    if exp in haystack:
        return True, "strict"
    for vt in facts["visible_texts"] + facts["interactive_texts"]:
        v = vt.lower().strip()
        if v and len(v) >= 2 and v in exp:
            return True, "loose_substring"
    tokens = _cjk_bigrams(expected)
    if tokens:
        hit = sum(1 for t in tokens if t in haystack)
        if hit >= 2 and hit >= len(tokens) * 0.5:
            return True, "loose_overlap"
    return False, ""


def _should_try_ai_fallback(checks: list[dict]) -> bool:
    if not checks:
        return True
    return all(str(c.get("type") or "") == "text_contains" for c in checks)


def _build_page_context(item: dict) -> dict | None:
    facts = _agent_observation_facts(item, {})
    if len(facts["visible_texts"]) < 3 and not facts["url"]:
        return None
    return {
        "url": facts["url"],
        "visible_texts": facts["visible_texts"][:30],
        "interactive_texts": facts["interactive_texts"][:15],
    }


def _context_signature(page_context: dict | None) -> str:
    if not page_context:
        return ""
    return sha256(json.dumps(
        {"url": page_context.get("url", ""),
         "vt": page_context.get("visible_texts", [])[:30]},
        ensure_ascii=False, sort_keys=True
    ).encode("utf-8")).hexdigest()


def _should_try_context_refinement(checks: list[dict]) -> bool:
    if not checks:
        return False
    return all(
        str(c.get("source") or "") == "ai" or str(c.get("type") or "") == "text_contains"
        for c in checks
    )


def _ai_assertion_fallback(expected_text: str, page_context: dict | None = None) -> list[dict]:
    """AI 断言兜底。任何失败都返回 []，由调用方降级为规则解析结果。"""
    try:
        settings = get_ai_settings(mask_key=False)
    except Exception:
        return []
    ctx_sig = _context_signature(page_context)
    try:
        cached = load_cached_assertion_analysis(expected_text, settings, context_sig=ctx_sig)
        if cached is not None:
            return cached
    except Exception:
        cached = None
    try:
        ai_checks = ai_service.parse_assertions_with_ai(expected_text, settings, page_context=page_context)
    except Exception:
        return []
    if not ai_checks:
        return []
    try:
        save_assertion_analysis(expected_text, settings, ai_checks, context_sig=ctx_sig)
    except Exception:
        pass
    return ai_checks


def _build_assertion_checks(expected_text: str, ai_fallback: bool = False, page_context: dict | None = None) -> list[dict]:
    text = _clean_case_step_text(expected_text)
    if not text:
        return []
    checks: list[dict] = []
    seen: set[tuple[str, str]] = set()

    def push(check_type: str, expected: str, label: str = "", extra: dict | None = None) -> None:
        normalized_expected = _clean_case_step_text(expected)
        if not normalized_expected:
            return
        key = (check_type, normalized_expected)
        if key in seen:
            return
        seen.add(key)
        payload = {
            "type": check_type,
            "expected": normalized_expected,
            "label": label or check_type,
            "status": "queued",
            "actual": "",
            "evidence_source": "",
            "reason": "",
        }
        if extra:
            payload.update(extra)
        checks.append(payload)

    for url_match in re.finditer(r"https?://[^\s,;，。；]+|#/[^\s,;，。；]+", text):
        push("url_contains", url_match.group(0), "URL")

    if ((any(word in text for word in ("复选框", "勾选框")) and any(word in text for word in ("选中", "已选中", "勾选"))) or "选中状态" in text):
        push("checkbox_checked", text, "复选框选中")
    if re.search(r"(回到|返回|跳转到).*(登录页|登录页面)|退出登录", text):
        push("url_contains", "#/login", "登录页")
    if "登录成功" in text:
        if "首页" in text:
            push("url_contains", "#/index", "首页")
        push("login_success", text, "登录成功")

    if re.search(r"弹窗.*(打开|显示|出现)|对话框.*(打开|显示|出现)", text):
        target = ""
        for fragment in _expected_text_fragments(text):
            if any(word in fragment for word in ("弹窗", "对话框", "页面", "提示")):
                target = fragment
                break
        push("dialog_visible", target or text, "弹窗打开")
    if re.search(r"弹窗.*(关闭|消失|收起)|对话框.*(关闭|消失|收起)", text):
        push("dialog_hidden", text, "弹窗关闭")
    if re.search(r"列表.*(新增|出现|显示).*(记录|一行)|表格.*(新增|出现|显示).*(记录|一行)", text):
        terms = [value for _label, value in _extract_expected_value_pairs(text)]
        push("table_row_exists", text, "列表记录", {"terms": terms})

    value_pairs = _extract_expected_value_pairs(text)
    for label, value in value_pairs:
        check_type = "status_tag" if label == "状态" else "field_value"
        push(check_type, value, label, {"field": label})

    fragments = _expected_text_fragments(text)
    for fragment in fragments:
        if re.search(r"https?://|#/", fragment):
            continue
        if "列表" in fragment and any(word in fragment for word in ("新增", "记录", "一行")):
            continue
        if any(word in fragment for word in ("弹窗关闭", "弹窗打开", "对话框关闭", "对话框打开")):
            continue
        if "登录成功" in fragment:
            continue
        if ((any(word in fragment for word in ("复选框", "勾选框")) and any(word in fragment for word in ("选中", "已选中", "勾选"))) or "选中状态" in fragment):
            continue
        if re.search(r"(回到|返回|跳转到).*(登录页|登录页面)|退出登录", fragment):
            continue
        push("text_contains", fragment, "页面文本")

    if ai_fallback and _should_try_ai_fallback(checks):
        ai_checks = _ai_assertion_fallback(text, page_context=page_context)
        if ai_checks:
            return ai_checks

    if not checks:
        push("text_contains", text, "页面文本")
    return checks


def _agent_observation_facts(item: dict, trace: dict) -> dict:
    decision = item.get("decision") if isinstance(item.get("decision"), dict) else {}
    execution = item.get("execution") if isinstance(item.get("execution"), dict) else {}
    observation = item.get("observation") if isinstance(item.get("observation"), dict) else {}
    visible_texts: list[str] = []
    for raw_text in observation.get("visibleText") or []:
        text = _clean_case_step_text(raw_text)
        if text and text not in visible_texts:
            visible_texts.append(text)
    interactive_texts: list[str] = []
    for raw_item in observation.get("interactives") or []:
        if not isinstance(raw_item, dict):
            continue
        text = _clean_case_step_text(str(raw_item.get("text") or ""))
        if text and text not in interactive_texts:
            interactive_texts.append(text)
    url = str(observation.get("url") or trace.get("finalUrl") or trace.get("final_url") or "").strip()
    result = str(execution.get("result") or "").strip()
    error = str(execution.get("error") or "").strip()
    decision_value = _clean_case_step_text(str(decision.get("value") or ""))
    haystack_parts = visible_texts + interactive_texts + [url, result, str(decision.get("reason") or ""), decision_value]
    haystack = " ".join(part for part in haystack_parts if part).lower()
    return {
        "visible_texts": visible_texts,
        "interactive_texts": interactive_texts,
        "url": url,
        "result": result,
        "error": error,
        "decision_value": decision_value,
        "haystack": haystack,
    }


def _evaluate_assertion_check(check: dict, item: dict, trace: dict) -> dict:
    evaluated = dict(check)
    facts = _agent_observation_facts(item, trace)
    expected = str(check.get("expected") or "").strip()
    check_type = str(check.get("type") or "").strip()
    visible_text = "；".join(facts["visible_texts"][:3])
    if facts["error"] and facts["result"] not in {"duplicate_action_blocked"}:
        evaluated.update(
            {
                "status": "failed",
                "actual": facts["error"],
                "evidence_source": "execution.error",
                "reason": "执行阶段已返回错误",
            }
        )
        return evaluated

    if check_type == "url_contains":
        url = facts["url"]
        if not url:
            evaluated.update({"status": "queued", "reason": "缺少 URL 证据"})
        elif expected.lower() in url.lower():
            evaluated.update({"status": "completed", "actual": url, "evidence_source": "observation.url", "reason": "当前 URL 命中预期"})
        elif expected == "#/login" and facts["result"] == "account_switch_passed":
            evaluated.update({"status": "completed", "actual": facts["url"] or facts["result"], "evidence_source": "execution.result", "reason": "账号切换已通过，URL 校验通过"})
        else:
            evaluated.update({"status": "failed", "actual": url, "evidence_source": "observation.url", "reason": "当前 URL 未命中预期"})
        return evaluated

    if check_type == "dialog_visible":
        dialog_signal = expected.lower()
        result = facts["result"].lower()
        haystack = facts["haystack"]
        if "dialog_opened" in result or "drawer_opened" in result:
            evaluated.update({"status": "completed", "actual": facts["result"], "evidence_source": "execution.result", "reason": "执行结果显示弹窗已打开"})
        elif dialog_signal and dialog_signal in haystack:
            evaluated.update({"status": "completed", "actual": visible_text or facts["url"], "evidence_source": "observation.visibleText", "reason": "页面可见文本命中弹窗内容"})
        elif facts["visible_texts"]:
            evaluated.update({"status": "failed", "actual": visible_text, "evidence_source": "observation.visibleText", "reason": "未观察到预期弹窗内容"})
        else:
            evaluated.update({"status": "queued", "reason": "缺少足够的弹窗证据"})
        return evaluated

    if check_type == "dialog_hidden":
        dialog_signal = expected.lower()
        result = facts["result"].lower()
        haystack = facts["haystack"]
        if dialog_signal and dialog_signal in haystack:
            evaluated.update({"status": "failed", "actual": visible_text, "evidence_source": "observation.visibleText", "reason": "弹窗相关文本仍然可见"})
        elif result in {"detail_assert_passed", "dialog_form_fill_passed", "finished"}:
            evaluated.update({"status": "completed", "actual": facts["result"], "evidence_source": "execution.result", "reason": "执行结果表明弹窗关闭校验已通过"})
        elif facts["visible_texts"] or facts["url"]:
            evaluated.update({"status": "queued", "actual": visible_text or facts["url"], "evidence_source": "observation.visibleText", "reason": "未再看到弹窗文本，但证据不足以直接判定关闭"})
        else:
            evaluated.update({"status": "queued", "reason": "缺少弹窗关闭证据"})
        return evaluated

    if check_type == "table_row_exists":
        terms = [str(item).strip() for item in (check.get("terms") or []) if str(item).strip()]
        if not facts["visible_texts"]:
            evaluated.update({"status": "queued", "reason": "缺少列表文本证据"})
        elif terms and all(term.lower() in facts["haystack"] for term in terms):
            evaluated.update({"status": "completed", "actual": visible_text, "evidence_source": "observation.visibleText", "reason": "列表中已匹配到目标记录关键字段"})
        elif not terms and expected.lower() in facts["haystack"]:
            evaluated.update({"status": "completed", "actual": visible_text, "evidence_source": "observation.visibleText", "reason": "列表文本命中预期描述"})
        else:
            evaluated.update({"status": "failed", "actual": visible_text, "evidence_source": "observation.visibleText", "reason": "列表中未找到目标记录"})
        return evaluated

    if check_type in {"field_value", "status_tag", "text_contains"}:
        if not facts["visible_texts"] and not facts["interactive_texts"]:
            if facts["result"] == "detail_assert_passed" and facts["decision_value"] and (
                facts["decision_value"].lower() in expected.lower() or expected.lower() in facts["decision_value"].lower()
            ):
                evaluated.update({"status": "completed", "actual": facts["decision_value"], "evidence_source": "decision.value", "reason": "断言动作已通过且匹配到目标文本", "match_strength": "strict"})
            else:
                evaluated.update({"status": "queued", "reason": "缺少页面文本证据"})
        else:
            matched, strength = _loose_text_match(expected, facts)
            if matched:
                source = "observation.visibleText" if expected.lower() in " ".join(facts["visible_texts"]).lower() else "observation.interactives"
                evaluated.update({"status": "completed", "actual": visible_text or "；".join(facts["interactive_texts"][:2]), "evidence_source": source, "reason": f"页面文本命中预期（{strength}）", "match_strength": strength})
            else:
                evaluated.update({"status": "failed", "actual": visible_text or "；".join(facts["interactive_texts"][:2]), "evidence_source": "observation.visibleText", "reason": "页面文本未命中预期"})
        return evaluated

    if check_type == "checkbox_checked":
        if facts["result"] == "user_device_bound":
            evaluated.update({
                "status": "completed",
                "actual": facts["decision_value"] or facts["result"],
                "evidence_source": "execution.result",
                "reason": "账号绑定动作已生效，复选框已勾选",
                "match_strength": "strict",
            })
        elif "is-checked" in facts["haystack"] or "已勾选" in facts["haystack"]:
            evaluated.update({
                "status": "completed",
                "actual": visible_text or "?".join(facts["interactive_texts"][:2]),
                "evidence_source": "observation.visibleText",
                "reason": "页面观察到复选框已勾选",
                "match_strength": "loose",
            })
        elif facts["visible_texts"] or facts["interactive_texts"]:
            evaluated.update({
                "status": "failed",
                "actual": visible_text or "?".join(facts["interactive_texts"][:2]),
                "evidence_source": "observation.visibleText",
                "reason": "页面未观察到复选框已勾选",
            })
        else:
            evaluated.update({"status": "queued", "reason": "缺少足够的复选框证据"})
        return evaluated

    if check_type == "login_success":
        url = facts["url"].lower()
        result = facts["result"]
        if result in {"login_guard_passed", "account_switch_passed"}:
            evaluated.update({
                "status": "completed",
                "actual": facts["url"] or result,
                "evidence_source": "execution.result",
                "reason": "登录阶段执行结果已通过",
                "match_strength": "strict",
            })
        elif url and "#/login" not in url:
            evaluated.update({
                "status": "completed",
                "actual": facts["url"],
                "evidence_source": "observation.url",
                "reason": "当前 URL 已离开登录页",
                "match_strength": "strict",
            })
        elif url:
            evaluated.update({
                "status": "failed",
                "actual": facts["url"],
                "evidence_source": "observation.url",
                "reason": "当前仍停留在登录页",
            })
        else:
            evaluated.update({"status": "queued", "reason": "缺少登录结果证据"})
        return evaluated

    match_mode = str(check.get("match_mode") or "contains").strip().lower()
    if match_mode not in {"contains", "equals", "not_contains", "regex"}:
        match_mode = "contains"
    if not expected:
        evaluated.update({"status": "queued", "reason": "断言缺少 expected 文本"})
        return evaluated
    if not facts["visible_texts"] and not facts["interactive_texts"]:
        if facts["result"] == "detail_assert_passed" and facts["decision_value"] and (
            facts["decision_value"].lower() in expected.lower() or expected.lower() in facts["decision_value"].lower()
        ):
            evaluated.update({"status": "completed", "actual": facts["decision_value"], "evidence_source": "decision.value", "reason": "断言动作已通过且匹配到目标文本", "match_strength": "strict"})
        else:
            evaluated.update({"status": "queued", "reason": "缺少页面文本证据"})
        return evaluated
    haystack = facts["haystack"]
    try:
        if match_mode == "contains":
            matched, strength = _loose_text_match(expected, facts)
        elif match_mode == "equals":
            matched = expected.lower() in [t.lower() for t in facts["visible_texts"] + facts["interactive_texts"]]
            strength = "strict" if matched else ""
        elif match_mode == "not_contains":
            matched = expected.lower() not in haystack
            strength = "strict"
        elif match_mode == "regex":
            matched = re.search(expected, haystack, re.IGNORECASE) is not None
            strength = "strict" if matched else ""
        else:
            matched, strength = _loose_text_match(expected, facts)
    except re.error:
        evaluated.update({"status": "queued", "reason": "断言 regex 模式非法"})
        return evaluated
    if matched:
        evaluated.update({"status": "completed", "actual": visible_text or "；".join(facts["interactive_texts"][:2]), "evidence_source": "observation.visibleText", "reason": f"页面文本命中预期（{match_mode}）", "match_strength": strength})
    else:
        evaluated.update({"status": "failed", "actual": visible_text or "；".join(facts["interactive_texts"][:2]), "evidence_source": "observation.visibleText", "reason": f"页面文本未命中预期（{match_mode}）", "match_strength": strength})
    return evaluated


def _aggregate_assertion_status(checks: list[dict]) -> str:
    final_status = "queued"
    for check in checks:
        status = str(check.get("status") or "queued")
        if _assertion_status_rank(status) > _assertion_status_rank(final_status):
            final_status = status
    return final_status


def _assertion_resolution_rank(status: str) -> int:
    if status == "completed":
        return 2
    if status == "failed":
        return 1
    return 0


def _history_event_index(history: list[dict], target: dict) -> int:
    for index, item in enumerate(history):
        if item is target:
            return index
    for index, item in enumerate(history):
        if item == target:
            return index
    return -1


def _assertion_evidence_candidates(step_entry: dict, history: list[dict], page_segments: list[dict]) -> list[dict]:
    candidates: list[dict] = []
    seen_ids: set[int] = set()

    def push(item: dict | None) -> None:
        if not isinstance(item, dict):
            return
        marker = id(item)
        if marker in seen_ids:
            return
        seen_ids.add(marker)
        candidates.append(item)

    step_events = [event for event in (step_entry.get("events") or []) if isinstance(event, dict)]
    for event in step_events:
        push(event)
    if step_events:
        last_index = _history_event_index(history, step_events[-1])
        if last_index >= 0:
            for offset in (1, 2):
                next_index = last_index + offset
                if next_index < len(history):
                    push(history[next_index])
    elif page_segments:
        try:
            step_index = max(int(step_entry.get("step_index") or 1), 1)
        except (TypeError, ValueError):
            step_index = 1
        segment_index = min(step_index - 1, len(page_segments) - 1)
        push(page_segments[segment_index]["item"])
    return candidates


def _assertion_actual_summary(checks: list[dict]) -> str:
    parts: list[str] = []
    for check in checks:
        actual = str(check.get("actual") or "").strip()
        if actual and actual not in parts:
            parts.append(actual)
    return "；".join(parts[:3])

def _build_agent_steps(agent_explore: dict | None, logs: list[dict], screenshots: list[dict[str, str]], run_id: str, task_dict: dict, ai_fallback: bool = False) -> list[dict]:
    trace = (agent_explore or {}).get("trace") or {}
    history = trace.get("history") or []
    case_steps = _load_agent_case_steps(run_id, task_dict, trace)
    expected_results = _load_agent_expected_results(run_id, task_dict, trace)
    expected_bindings = _parse_expected_result_bindings(expected_results, len(case_steps))
    stage_source_map = _agent_stage_source_map(trace)
    stage_meta_map = _agent_stage_meta_map(trace)
    trace_status = str(trace.get("status") or "").strip().lower()
    trace_failed = bool(trace.get("error")) or trace_status == "failed"
    trace_completed = bool(trace.get("ok")) or trace_status in {"passed", "completed", "success"}
    trace_active = not trace_failed and not trace_completed
    if not case_steps:
        steps: list[dict] = []
        current_case_step_index = -1
        for index, item in enumerate(history, start=1):
            decision = item.get("decision") if isinstance(item, dict) else {}
            execution = item.get("execution") if isinstance(item, dict) else {}
            observation = item.get("observation") if isinstance(item, dict) else {}
            decision = decision if isinstance(decision, dict) else {}
            execution = execution if isinstance(execution, dict) else {}
            observation = observation if isinstance(observation, dict) else {}
            item_dict = item if isinstance(item, dict) else {}
            screenshot = _agent_step_screenshot(item_dict, screenshots, index)
            matched_case_step = _match_stage_case_step_index(item_dict, stage_source_map, case_steps, current_case_step_index)
            if matched_case_step is None:
                matched_case_step = _match_case_step_index(case_steps, decision, observation, current_case_step_index)
            if matched_case_step is not None:
                current_case_step_index = matched_case_step
            action_summary = _agent_step_action_summary(decision, execution, observation)
            case_step_title = ""
            case_step_text = ""
            if matched_case_step is not None and 0 <= matched_case_step < len(case_steps):
                case_step_text = case_steps[matched_case_step]
                case_step_title = f"用例步骤 {matched_case_step + 1} - {case_step_text}"
            assertion_title, assertion_text = _resolve_agent_assertion_title(
                item_dict,
                case_step_title,
                expected_results,
                stage_meta_map,
            )
            action_name = str(decision.get("action") or "").strip().lower()
            display_title = assertion_title or case_step_title
            display_action_text = assertion_text or case_step_text
            summary_text = (
                assertion_text
                or (case_step_text if action_name in {"fill", "click", "goto"} else "")
                or (action_summary if action_name == "wait" else "")
                or str(decision.get("reason") or "").strip()
                or action_summary
                or str(execution.get("result") or "")
            )
            step_status = "completed"
            if execution.get("error"):
                step_status = "failed"
            elif index == len(history) and trace_failed:
                step_status = "failed"
            elif not trace_completed and index == len(history):
                step_status = "running"
            steps.append(
                {
                    "step_index": int(item.get("step", index)) if isinstance(item, dict) else index,
                    "step_code": f"agent_{index:02d}",
                    "title": (
                        display_title
                        or str(decision.get("reason") or "").strip()
                        or action_summary
                        or str(decision.get("action") or f"Step {index}")
                    ),
                    "status": step_status,
                    "started_at": None,
                    "finished_at": None,
                    "duration_seconds": None,
                    "summary": summary_text,
                    "error_message": str(execution.get("error") or ""),
                    "screenshot_url": screenshot["url"] if screenshot else "",
                    "ai_analysis": str(decision.get("reason") or action_summary or ""),
                    "final_url": str(trace.get("finalUrl") or trace.get("final_url") or ""),
                    "command_output": _agent_step_command_output(item_dict, display_action_text),
                    "selectors": [],
                    "inputs": [],
                    "console_logs": [],
                    "network_logs": [],
                    "dom_snapshot_url": "",
                    "events": [item] if isinstance(item, dict) else [],
                    "expected_result": "",
                    "expected_result_status": "queued",
                    "actual_result": "",
                    "assertion_checks": [],
                }
            )
        return steps
    steps: list[dict] = [
        {
            "step_index": step_index,
            "step_code": f"agent_step_{step_index:02d}",
            "title": f"用例步骤 {step_index} - {case_step_text}",
            "status": "queued",
            "started_at": None,
            "finished_at": None,
            "duration_seconds": None,
            "summary": case_step_text,
            "error_message": "",
            "screenshot_url": "",
            "ai_analysis": "",
            "final_url": str(trace.get("finalUrl") or trace.get("final_url") or ""),
            "command_output": [],
            "selectors": [],
            "inputs": [],
            "console_logs": [],
            "network_logs": [],
            "dom_snapshot_url": "",
            "events": [],
            "expected_result": "",
            "expected_result_status": "queued",
            "actual_result": "",
            "assertion_checks": [],
        }
        for step_index, case_step_text in enumerate(case_steps, start=1)
    ]
    for binding in expected_bindings:
        checks = _build_assertion_checks(binding["text"], ai_fallback=ai_fallback)
        for step_index in range(binding["start_step"], binding["end_step"] + 1):
            step_entry = steps[step_index - 1]
            if step_entry["expected_result"]:
                continue
            step_entry["expected_result"] = binding["text"]
            step_entry["expected_result_status"] = "queued"
            step_entry["assertion_checks"] = [dict(check) for check in checks]
            if step_index < binding["end_step"]:
                step_entry["actual_result"] = f"该预期结果将在步骤 {binding['end_step']} 完成后校验"
    current_case_step_index = -1
    for history_index, item in enumerate(history, start=1):
        decision = item.get("decision") if isinstance(item, dict) else {}
        execution = item.get("execution") if isinstance(item, dict) else {}
        observation = item.get("observation") if isinstance(item, dict) else {}
        decision = decision if isinstance(decision, dict) else {}
        execution = execution if isinstance(execution, dict) else {}
        observation = observation if isinstance(observation, dict) else {}
        item_dict = item if isinstance(item, dict) else {}
        screenshot = _agent_step_screenshot(item_dict, screenshots, history_index)
        matched_case_step = _match_stage_case_step_index(item_dict, stage_source_map, case_steps, current_case_step_index)
        if matched_case_step is None:
            matched_case_step = _match_case_step_index(case_steps, decision, observation, current_case_step_index)
        if matched_case_step is not None:
            current_case_step_index = matched_case_step
        if matched_case_step is None or not (0 <= matched_case_step < len(case_steps)):
            continue
        step_entry = steps[matched_case_step]
        case_step_text = case_steps[matched_case_step]
        action_summary = _agent_step_action_summary(decision, execution, observation)
        action_name = str(decision.get("action") or "").strip().lower()
        summary_text = (
            (case_step_text if action_name in {"fill", "click", "goto"} else "")
            or (action_summary if action_name == "wait" else "")
            or str(decision.get("reason") or "").strip()
            or action_summary
            or str(execution.get("result") or "")
        )
        if summary_text:
            step_entry["summary"] = summary_text
        step_entry["ai_analysis"] = str(decision.get("reason") or action_summary or step_entry["ai_analysis"])
        if screenshot:
            step_entry["screenshot_url"] = screenshot["url"]
        step_entry["events"].append(item)
        step_entry["command_output"].extend(_agent_step_command_output(item_dict, case_step_text))
        if execution.get("error"):
            step_entry["status"] = "failed"
            step_entry["error_message"] = str(execution.get("error") or "")
        elif step_entry["status"] != "failed":
            step_entry["status"] = "completed"
        assertion_binding = _agent_assertion_binding(item_dict, expected_bindings, stage_meta_map)
        if assertion_binding:
            bound_step = steps[assertion_binding["end_step"] - 1]
            bound_step["expected_result"] = assertion_binding["text"]
            current_checks = bound_step["assertion_checks"] or _build_assertion_checks(assertion_binding["text"], ai_fallback=ai_fallback)
            evaluated_checks = [_evaluate_assertion_check(check, item_dict, trace) for check in current_checks]
            bound_step["assertion_checks"] = evaluated_checks
            bound_step["actual_result"] = _assertion_actual_summary(evaluated_checks) or _agent_actual_result_text(item_dict, trace) or summary_text or case_step_text
            if execution.get("error"):
                bound_step["expected_result_status"] = "failed"
                bound_step["status"] = "failed"
                bound_step["error_message"] = str(execution.get("error") or "")
            else:
                aggregated_status = _aggregate_assertion_status(evaluated_checks)
                if trace_active and aggregated_status == "failed":
                    evaluated_checks = [
                        {**check, "status": "queued", "reason": "waiting for final evidence"}
                        if str(check.get("status") or "") == "failed"
                        else check
                        for check in evaluated_checks
                    ]
                    bound_step["assertion_checks"] = evaluated_checks
                    aggregated_status = "queued"
                bound_step["expected_result_status"] = aggregated_status
                if bound_step["expected_result_status"] == "failed":
                    bound_step["status"] = "failed"
                elif bound_step["expected_result_status"] == "completed" and bound_step["status"] != "failed":
                    bound_step["status"] = "completed"
    if trace_failed:
        for step_entry in reversed(steps):
            if step_entry["status"] == "completed" and step_entry["expected_result_status"] == "queued":
                step_entry["status"] = "failed"
                step_entry["actual_result"] = str(trace.get("error") or step_entry["actual_result"] or step_entry["summary"])
                step_entry["error_message"] = str(trace.get("error") or step_entry["error_message"])
                break
    elif not trace_completed:
        for step_entry in reversed(steps):
            if step_entry["status"] == "completed":
                step_entry["status"] = "running"
                break
    elif trace_completed:
        _finalize_queued_assertions_with_last_observation(steps, history, ai_fallback=ai_fallback)
    for step_entry in steps:
        if step_entry.get("expected_result_status") == "completed" and step_entry.get("status") == "failed":
            step_entry["status"] = "completed"
            step_entry["error_message"] = ""
    return steps


def _finalize_queued_assertions_with_last_observation(steps: list[dict], history: list[dict], ai_fallback: bool = False) -> None:
    """trace 成功时，对"未明确通过"的断言用"该步骤匹配的 history observation"兜底求值。

    解决三类问题：
    1. stage_local_step 与 expected_binding 错位：agent 在某 stage 内自主做了
       多次 assert_text，绑定会错位，导致用错误的 observation 求值。
    2. duplicate_action_blocked 被当执行错误：这些步骤的 observation 其实是正常的。
    3. 用最终页面 observation 求值早期步骤：早期步骤的预期应基于早期页面状态。

    策略：按 URL 变化点把 history 分段，每段代表"一个页面状态"。
    按步骤序号选对应的 page_segment，用那个 observation 重新求值。
    只重求 status 为 queued 或 failed 的步骤；completed 的保持不变。
    ai_fallback=True 时，对仍 failed 的步骤用 page_context 调 AI 再解析。
    """
    if not history or not steps:
        return
    page_segments: list[dict] = []
    for item in history:
        if not isinstance(item, dict):
            continue
        obs = item.get("observation") or {}
        url = str(obs.get("url") or "").strip()
        if not url or not (obs.get("visibleText")):
            continue
        if not page_segments or page_segments[-1]["url"] != url:
            page_segments.append({"url": url, "item": item})
    if not page_segments:
        return
    trace_ref: dict = {}
    for step_index, step_entry in enumerate(steps):
        current_status = step_entry.get("expected_result_status")
        if current_status not in {"queued", "failed"}:
            continue
        checks = step_entry.get("assertion_checks") or []
        if not checks:
            continue
        candidate_items = _assertion_evidence_candidates(step_entry, history, page_segments)
        if not candidate_items:
            continue
        if ai_fallback and _should_try_context_refinement(checks):
            page_context = _build_page_context(candidate_items[-1])
            if page_context:
                refined = _ai_assertion_fallback(step_entry.get("expected_result", ""), page_context=page_context)
                if refined:
                    checks = refined
                    step_entry["assertion_checks"] = checks
        best_item = candidate_items[0]
        evaluated_checks = [_evaluate_assertion_check(check, best_item, trace_ref) for check in checks]
        best_status = _aggregate_assertion_status(evaluated_checks)
        for candidate in candidate_items[1:]:
            candidate_checks = [_evaluate_assertion_check(check, candidate, trace_ref) for check in checks]
            candidate_status = _aggregate_assertion_status(candidate_checks)
            if _assertion_resolution_rank(candidate_status) > _assertion_resolution_rank(best_status):
                best_item = candidate
                evaluated_checks = candidate_checks
                best_status = candidate_status
        step_entry["assertion_checks"] = evaluated_checks
        aggregated = best_status
        step_entry["expected_result_status"] = aggregated
        failed_count = sum(1 for c in evaluated_checks if c.get("status") == "failed")
        completed_count = sum(1 for c in evaluated_checks if c.get("status") == "completed")
        step_entry["partial_pass"] = completed_count > 0 and failed_count > 0
        step_entry["failed_checks_count"] = failed_count
        step_entry["completed_checks_count"] = completed_count
        if not step_entry.get("actual_result") or current_status == "failed":
            step_entry["actual_result"] = _assertion_actual_summary(evaluated_checks) or _agent_actual_result_text(best_item, trace_ref)
        if aggregated == "completed":
            step_entry["status"] = "completed"
        elif aggregated == "failed":
            step_entry["status"] = "failed"


def _case_title_for_run(run_id: str, task_dict: dict, agent_explore: dict | None) -> str:
    trace = (agent_explore or {}).get("trace") or {}
    try:
        yaml_text, _ = _load_case_source_for_agent_run(run_id, task_dict, trace)
        payload = yaml.safe_load(yaml_text) or {}
    except (HTTPException, OSError, yaml.YAMLError):
        return ""
    if not isinstance(payload, dict):
        return ""
    title = str(payload.get("title") or "").strip()
    case_id = str(payload.get("id") or task_dict.get("case_id") or "").strip()
    return "" if title == case_id else title


def _build_unified_run_detail(run_id: str) -> dict:
    with connect() as conn:
        task = conn.execute("select * from run_tasks where id = ?", (run_id,)).fetchone()
        logs = conn.execute("select * from run_logs where run_id = ? order by id", (run_id,)).fetchall()
    if not task:
        raise HTTPException(status_code=404, detail="run not found")
    task_dict = dict(task)
    log_dicts = rows_to_dicts(logs)
    if not log_dicts and task_dict.get("mode") == "agent-explore":
        log_dicts = _synthetic_logs_from_evidence(run_id)
    report_text = ""
    if task_dict.get("report_path") and Path(str(task_dict["report_path"])).suffix.lower() == ".md":
        try:
            report_text = read_report(run_id)
        except FileNotFoundError:
            report_text = ""
    batch_children = list_batch_child_reports(run_id, case_ids=_batch_case_ids(task_dict)) if task_dict.get("mode") == "run-batch" else []
    if task_dict.get("mode") == "run-batch" and not report_text and batch_children:
        report_text = _batch_summary_report(run_id, batch_children, _batch_case_ids(task_dict))
    screenshots = _report_screenshots(task_dict, report_text)
    analysis = ai_service.analyze_run_report(report_text, screenshots, [row["line"] for row in log_dicts] + evidence_log_lines(run_id)) if report_text else None
    agent_explore = load_agent_explore_artifacts(run_id) if task_dict.get("mode") == "agent-explore" else None
    step_detail_payload = load_step_details(run_id)
    if step_detail_payload:
        steps = step_detail_payload.get("steps") or []
        final_url = step_detail_payload.get("final_url") or ""
        summary = step_detail_payload.get("summary") or {}
    elif task_dict.get("mode") == "run-batch":
        steps = _build_batch_steps(batch_children)
        final_url = ""
        summary = {
            "title": f"Custom batch ({len(batch_children)} cases)" if _batch_case_ids(task_dict) else "Batch 001-012",
            "conclusion": f"Batch 已执行 {sum(1 for child in batch_children if child.get('run_id'))} 条用例。",
            "failure_reason": task_dict.get("error") or "",
            "ai_analysis": "",
        }
    else:
        steps = _build_agent_steps(agent_explore, log_dicts, screenshots, run_id, task_dict, ai_fallback=True)
        trace = (agent_explore or {}).get("trace") or {}
        final_url = str(trace.get("finalUrl") or trace.get("final_url") or "")
        summary = {
            "title": task_dict.get("case_id") or run_id,
            "conclusion": str(trace.get("summary") or ""),
            "failure_reason": str(trace.get("error") or task_dict.get("error") or ""),
            "ai_analysis": str(trace.get("summary") or ""),
        }
    case_name = _case_title_for_run(run_id, task_dict, agent_explore) or task_dict.get("case_id") or run_id
    if report_text:
        report_meta = parse_report(report_text)
        case_name = report_meta.get("case_name") or case_name
    trace = (agent_explore or {}).get("trace") or {}
    agent_stage_runs = _reconcile_agent_stage_runs(trace, steps, (trace.get("stage_runs") or []))
    return {
        "run_id": run_id,
        "case_id": task_dict.get("case_id"),
        "case_name": case_name,
        "mode": _normalized_mode(str(task_dict.get("mode") or "")),
        "trigger": str(task_dict.get("trigger") or "manual"),
        "parent_run_id": task_dict.get("parent_run_id"),
        "status": _normalized_status(str(task_dict.get("status") or "")),
        "operator": "admin",
        "started_at": task_dict.get("started_at") or task_dict.get("created_at"),
        "finished_at": task_dict.get("finished_at"),
        "duration_seconds": task_dict.get("summary", {}).get("duration_seconds") if isinstance(task_dict.get("summary"), dict) else None,
        "final_url": final_url,
        "summary": {
            "title": summary.get("title") or case_name,
            "conclusion": summary.get("conclusion") or "",
            "failure_reason": summary.get("failure_reason") or "",
            "ai_analysis": analysis["conclusion"] if analysis else (summary.get("ai_analysis") or ""),
        },
        "steps": steps,
        "artifacts": {
            "report_markdown_url": f"/api/reports/{run_id}" if report_text else "",
            "observed_asset_path": (step_detail_payload or {}).get("artifacts", {}).get("observed_asset_path", ""),
            "observed_asset_merge_url": f"/api/runs/{run_id}/merge-observed-asset",
            "trace_download_url": f"/api/runs/{run_id}/evidence/trace" if evidence_summary(run_id).get("trace", {}).get("exists") else "",
            "candidate_flow_url": f"/api/runs/{run_id}/agent-explore/candidate-flow" if agent_explore and agent_explore.get("candidate_flow_path") else "",
        },
        "raw_report": report_text,
        "logs": log_dicts,
        "screenshots": screenshots,
        "evidence": evidence_summary(run_id),
        "agent_explore": agent_explore,
        "analysis": analysis,
        "healing_hint": str((trace.get("healing_hint") or "")),
        "agent_plan": trace.get("plan") or {},
        "agent_stage_runs": agent_stage_runs,
        "current_stage_id": str((trace.get("current_stage_id") or "")),
        "current_stage_name": str((trace.get("current_stage_name") or "")),
        "current_strategy": str((trace.get("current_strategy") or "")),
    }


@app.get("/api/runs/{run_id}/detail")
def run_detail_view(run_id: str) -> dict:
    return _build_unified_run_detail(run_id)


def load_agent_explore_artifacts(run_id: str) -> dict | None:
    root = ROOT / "reports" / "agent-explore" / run_id
    trace_path = root / "trace.json"
    if not trace_path.exists():
        return None
    candidate_path = root / "candidate_flow.py"
    try:
        trace = json.loads(trace_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        trace = {"ok": False, "error": "invalid agent trace json"}
    return {
        "trace_path": str(trace_path.relative_to(ROOT)),
        "candidate_flow_path": str(candidate_path.relative_to(ROOT)) if candidate_path.exists() else "",
        "trace": trace,
    }


def _load_case_source_for_agent_run(run_id: str, task_dict: dict, trace: dict) -> tuple[str, str]:
    draft_path = DRAFT_RUN_DIR / run_id / "case.yaml"
    if draft_path.exists():
        return draft_path.read_text(encoding="utf-8"), str(draft_path)
    case_arg = str(trace.get("case_arg") or "").strip()
    if case_arg:
        path = Path(case_arg)
        if path.exists():
            return path.read_text(encoding="utf-8"), str(path)
    case_id = str(task_dict.get("case_id") or "").strip()
    if case_id:
        path = find_case_yaml(case_id)
        return path.read_text(encoding="utf-8"), str(path)
    raise HTTPException(status_code=404, detail="case source not found for self heal")


def _classify_self_heal_failure(trace: dict, events: list[dict]) -> dict:
    """Classify the failure into one of 4 self-heal categories.

    Categories: locator_drift | timing | logic_understanding | unrecoverable | unknown
    """
    error = str(trace.get("error") or "").strip()
    if not error:
        history = trace.get("history") or []
        for item in reversed(history):
            if isinstance(item, dict):
                exec_result = item.get("execution") if isinstance(item.get("execution"), dict) else {}
                err = exec_result.get("error")
                if err:
                    error = str(err)
                    break

    evidence = error or "(no explicit error captured)"

    normalized = error.lower()
    if "unknown ref" in error or "no assertion signal matched" in error:
        return {"category": "locator_drift", "evidence": evidence}
    if "timeout" in normalized or "timed out" in normalized:
        return {"category": "timing", "evidence": evidence}
    if "login" in normalized or "credential" in normalized or "password" in normalized:
        return {"category": "logic_understanding", "evidence": evidence}
    if "agent_action_failed" in error or "action failed" in normalized:
        return {"category": "unrecoverable", "evidence": evidence}
    return {"category": "unknown", "evidence": evidence}


def _self_heal_hint(trace: dict, events: list[dict]) -> str:
    """Build the V2 self-heal hint text shown in the AI测试 self-heal info panel.

    Returns a structured text block with Failure Diagnosis, Recovery Strategy, and Stop Conditions.
    """
    diagnosis = _classify_self_heal_failure(trace, events if isinstance(events, list) else [])
    category = str(diagnosis.get("category") or "unknown")
    evidence = str(diagnosis.get("evidence") or "")

    strategy_lines = [
        "- locator_drift（定位漂移）：重新观察页面，按邻近 label 文本或 placeholder 选新 ref；不要复用之前的 ref。",
        "- timing（时序问题）：等待 1-3 秒后再 assert 或 click；弹窗/路由切换后不要立即操作，先等 DOM 稳定。",
        "- logic_understanding（业务理解偏差）：重读用例步骤和预期结果；不要在不理解业务的情况下继续操作。",
        "- unrecoverable（不可恢复）：不要重试，立刻 finish，reason 写明 'unrecoverable: <具体原因>'。",
    ]

    lines = [
        "失败诊断：",
        f"- Category: {category}",
        f"- 证据: {evidence}",
        "",
        "恢复策略（按 Category 选一条）：",
    ]
    lines.extend(strategy_lines)
    lines.extend(
        [
            "",
            "停止条件（任一命中立刻 finish）：",
            "- 已是第 3 次重试。",
            "- 页面 visibleText 已包含任意一条 Expected results。",
            "- 失败信号与上一轮 Category 相同。",
            "- 剩余工作无法仅靠浏览器操作完成（需后端/账号/数据准备）。",
        ]
    )
    return "\n".join(lines)


def _build_self_heal_context(run_id: str, task_dict: dict, trace: dict) -> dict:
    events = evidence_summary(run_id).get("events", {}).get("latest", [])
    history = trace.get("history") or []
    events_list = events if isinstance(events, list) else []
    diagnosis = _classify_self_heal_failure(trace, events_list)
    return {
        "parent_run_id": run_id,
        "trigger": "self_heal",
        "failure_summary": str(trace.get("error") or task_dict.get("error") or trace.get("summary") or "").strip(),
        "healing_hint": _self_heal_hint(trace, events_list),
        "diagnosis": diagnosis,
        "attempt_index": 1,
        "max_attempts": 3,
        "last_history": history[-5:] if isinstance(history, list) else [],
    }



@app.get("/api/runs/{run_id}/agent-explore/trace")
def agent_explore_trace_file(run_id: str) -> FileResponse:
    path = ROOT / "reports" / "agent-explore" / run_id / "trace.json"
    if not path.exists():
        raise HTTPException(status_code=404, detail="agent trace not found")
    return FileResponse(path, media_type="application/json", filename=f"{run_id}-agent-trace.json")


@app.get("/api/runs/{run_id}/agent-explore/candidate-flow")
def agent_explore_candidate_flow_file(run_id: str) -> FileResponse:
    path = ROOT / "reports" / "agent-explore" / run_id / "candidate_flow.py"
    if not path.exists():
        raise HTTPException(status_code=404, detail="agent candidate flow not found")
    return FileResponse(path, media_type="text/x-python", filename=f"{run_id}-candidate-flow.py")


def _next_agent_case_id() -> str:
    numbers = [
        int(match.group(1))
        for path in TEST_CASE_DIR.glob("TC-ICM-*.yaml")
        for match in [re.search(r"TC-ICM-(\d+)", path.name, flags=re.IGNORECASE)]
        if match
    ]
    return f"TC-ICM-{(max(numbers) if numbers else 12) + 1:03d}"


def _draft_for_agent_run(run_id: str) -> tuple[dict, Path]:
    draft_path = DRAFT_RUN_DIR / run_id / "case.yaml"
    if not draft_path.exists():
        raise HTTPException(status_code=400, detail="only draft Agent Explore runs can be promoted")
    try:
        data = yaml.safe_load(draft_path.read_text(encoding="utf-8")) or {}
    except yaml.YAMLError as exc:
        raise HTTPException(status_code=400, detail=f"draft case YAML is invalid: {exc}") from exc
    if not isinstance(data, dict):
        raise HTTPException(status_code=400, detail="draft case YAML root must be an object")
    return data, draft_path


def _formal_case_id_for_draft(data: dict) -> str:
    raw_case_id = str(data.get("id") or "").strip().upper()
    if re.fullmatch(r"TC-ICM-[0-9A-Z-]+", raw_case_id):
        try:
            find_case_yaml(raw_case_id)
        except HTTPException:
            return raw_case_id
    return _next_agent_case_id()


def _draft_source_case_id(data: dict) -> str:
    return str(data.get("id") or "").strip().upper()


def _existing_promoted_case_for_draft(draft_path: Path) -> str | None:
    yaml_text = draft_path.read_text(encoding="utf-8")
    with connect() as conn:
        row = conn.execute(
            """
            select promoted_case_id from case_drafts
            where yaml = ? and promoted_case_id is not null and promoted_case_id != ''
            order by id desc limit 1
            """,
            (yaml_text,),
        ).fetchone()
    return str(row["promoted_case_id"]) if row else None


def _existing_promoted_case_for_draft_identity(draft_data: dict) -> tuple[str, Path] | None:
    source_case_id = _draft_source_case_id(draft_data)
    if not source_case_id or source_case_id.startswith("TC-ICM-"):
        return None
    matches: list[tuple[str, Path]] = []
    for path in sorted(TEST_CASE_DIR.glob("*.yaml")):
        try:
            data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        except (OSError, yaml.YAMLError):
            continue
        if not isinstance(data, dict) or str(data.get("source_draft_case_id") or "").strip().upper() != source_case_id:
            continue
        case_id = str(data.get("id") or "").strip()
        if case_id:
            matches.append((case_id, path))
    return matches[0] if matches else None


def _existing_promoted_case_for_run(run_id: str) -> tuple[str, Path] | None:
    matches: list[tuple[str, Path]] = []
    for path in sorted(TEST_CASE_DIR.glob("*.yaml")):
        try:
            data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        except (OSError, yaml.YAMLError):
            continue
        if not isinstance(data, dict) or str(data.get("source_run_id") or "") != run_id:
            continue
        case_id = str(data.get("id") or "").strip()
        if case_id:
            matches.append((case_id, path))
    return matches[0] if matches else None


def _mark_matching_draft_promoted(draft_path: Path, case_id: str, target: Path) -> int | None:
    yaml_text = draft_path.read_text(encoding="utf-8")
    now = utc_now()
    with connect() as conn:
        row = conn.execute("select id from case_drafts where yaml = ? order by id desc limit 1", (yaml_text,)).fetchone()
        if not row:
            data = yaml.safe_load(yaml_text) or {}
            row = conn.execute(
                "select id from case_drafts where title = ? order by id desc limit 1",
                (str(data.get("title") or ""),),
            ).fetchone()
        if not row:
            return None
        draft_id = int(row["id"])
        conn.execute(
            """
            update case_drafts
            set status = 'promoted', promoted_case_id = ?, promoted_path = ?, updated_at = ?
            where id = ?
            """,
            (case_id, str(target), now, draft_id),
        )
        return draft_id


def _agent_trace_asset(trace: dict, draft_data: dict) -> dict:
    history = trace.get("history") or []
    operation_steps: list[str] = []
    selectors: dict[str, str] = {}
    input_values: dict[str, str] = {}
    for index, item in enumerate(history, start=1):
        decision = item.get("decision") if isinstance(item, dict) else {}
        execution = item.get("execution") if isinstance(item, dict) else {}
        decision = decision if isinstance(decision, dict) else {}
        execution = execution if isinstance(execution, dict) else {}
        action = str(decision.get("action") or f"step_{index}")
        selector = str(execution.get("selector") or "").strip()
        value = str(decision.get("value") or "").strip()
        operation_steps.append(str(decision.get("reason") or action))
        if selector:
            selectors[f"{action}_{index}"] = selector
        if value:
            input_values[f"{action}_{index}"] = value
    expected = draft_data.get("expected_results") or draft_data.get("expected") or []
    assertions = [str(item) for item in expected] if isinstance(expected, list) else [str(expected)] if expected else []
    return {
        "status": "verified",
        "source": "agent-explore",
        "operation_steps": operation_steps or [str(item) for item in (draft_data.get("steps") or [])],
        "selectors": selectors or {"agent_candidate": "candidate_flow.py"},
        "input_values": input_values,
        "assertions": assertions or [str(trace.get("summary") or "Agent Explore passed")],
    }


def _normalize_agent_draft_for_regression(draft_data: dict, case_id: str, run_id: str, trace: dict) -> dict:
    data = dict(draft_data)
    if "expected_results" not in data and "expected" in data:
        data["expected_results"] = data.get("expected")
    if "preconditions" not in data and "precondition" in data:
        data["preconditions"] = [str(data.get("precondition"))]
    data["id"] = case_id
    data["status"] = "formal"
    data["source"] = "agent-explore"
    data["source_run_id"] = run_id
    data["source_draft_case_id"] = _draft_source_case_id(draft_data)
    data["automation"] = "Yes"
    data["automation_asset"] = merge_automation_asset(data.get("automation_asset") if isinstance(data.get("automation_asset"), dict) else {}, _agent_trace_asset(trace, data))
    return data


@app.post("/api/runs/{run_id}/agent-explore/promote-regression")
def promote_agent_explore_regression(run_id: str) -> dict:
    with connect() as conn:
        task = conn.execute("select * from run_tasks where id = ?", (run_id,)).fetchone()
    if not task:
        raise HTTPException(status_code=404, detail="run not found")
    task_dict = dict(task)
    if task_dict.get("mode") != "agent-explore":
        raise HTTPException(status_code=400, detail="only Agent Explore runs can be promoted")
    if _normalized_status(str(task_dict.get("status") or "")) != "completed":
        raise HTTPException(status_code=400, detail="only passed Agent Explore runs can be promoted")

    agent_explore = load_agent_explore_artifacts(run_id)
    trace = (agent_explore or {}).get("trace") or {}
    if trace.get("ok") is False or str(trace.get("status") or "").lower() in {"failed", "error"}:
        raise HTTPException(status_code=400, detail="only passed Agent Explore traces can be promoted")

    candidate_path = ROOT / "reports" / "agent-explore" / run_id / "candidate_flow.py"
    if not candidate_path.exists():
        raise HTTPException(status_code=404, detail=f"candidate flow not found for run_id={run_id}")
    candidate_text = candidate_path.read_text(encoding="utf-8")
    if not candidate_text.strip():
        raise HTTPException(status_code=400, detail="candidate_flow.py is empty")
    try:
        ast.parse(candidate_text)
    except SyntaxError as exc:
        raise HTTPException(status_code=400, detail=f"candidate_flow.py syntax error: {exc}") from exc

    draft_data, draft_path = _draft_for_agent_run(run_id)
    existing_by_run = _existing_promoted_case_for_run(run_id)
    if existing_by_run:
        existing_case_id, existing_case_path = existing_by_run
        existing_flow_path = compute_target_path(existing_case_id)
        return {
            "case_id": existing_case_id,
            "case_path": str(existing_case_path),
            "flow_path": str(existing_flow_path),
            "draft_id": _mark_matching_draft_promoted(draft_path, existing_case_id, existing_case_path),
            "status": "promoted",
        }
    existing_by_identity = _existing_promoted_case_for_draft_identity(draft_data)
    case_id = (existing_by_identity[0] if existing_by_identity else None) or _existing_promoted_case_for_draft(draft_path) or _formal_case_id_for_draft(draft_data)
    case_filename = normalize_case_filename(None, case_id)
    case_target = TEST_CASE_DIR / case_filename
    flow_target = compute_target_path(case_id)

    draft_data = _normalize_agent_draft_for_regression(draft_data, case_id, run_id, trace)
    yaml_text = yaml.safe_dump(draft_data, allow_unicode=True, sort_keys=False)
    validation = validate_case_yaml(yaml_text)
    if not validation["valid"]:
        raise HTTPException(status_code=400, detail=f"YAML validation failed: {'; '.join(validation['errors'])}")

    TEST_CASE_DIR.mkdir(parents=True, exist_ok=True)
    flow_target.parent.mkdir(parents=True, exist_ok=True)
    backup_case_path = _backup_yaml_before_write(case_id, case_target) if case_target.exists() else None
    backup_flow_path = _backup_flow_before_write(flow_target, case_id) if flow_target.exists() else None
    try:
        case_target.write_text(yaml_text, encoding="utf-8")
        flow_target.write_text(candidate_text, encoding="utf-8")
        py_compile.compile(str(flow_target), doraise=True)
    except Exception as exc:
        if backup_case_path is not None and backup_case_path.exists():
            case_target.write_text(backup_case_path.read_text(encoding="utf-8"), encoding="utf-8")
        else:
            case_target.unlink(missing_ok=True)
        if backup_flow_path is not None and backup_flow_path.exists():
            flow_target.write_text(backup_flow_path.read_text(encoding="utf-8"), encoding="utf-8")
        else:
            flow_target.unlink(missing_ok=True)
        raise HTTPException(status_code=500, detail=f"failed to promote regression case: {exc}") from exc

    draft_id = _mark_matching_draft_promoted(draft_path, case_id, case_target)
    return {
        "case_id": case_id,
        "case_path": str(case_target),
        "flow_path": str(flow_target),
        "draft_id": draft_id,
        "status": "promoted",
    }


@app.post("/api/runs/{run_id}/agent-explore/self-heal")
def self_heal_agent_explore(run_id: str) -> dict:
    with connect() as conn:
        task = conn.execute("select * from run_tasks where id = ?", (run_id,)).fetchone()
    if not task:
        raise HTTPException(status_code=404, detail="run not found")
    task_dict = dict(task)
    if task_dict.get("mode") != "agent-explore":
        raise HTTPException(status_code=400, detail="only Agent Explore runs can self heal")
    agent_explore = load_agent_explore_artifacts(run_id)
    trace = (agent_explore or {}).get("trace") or {}
    if not trace:
        raise HTTPException(status_code=404, detail="trace not found for self heal")
    case_yaml, _ = _load_case_source_for_agent_run(run_id, task_dict, trace)
    context = _build_self_heal_context(run_id, task_dict, trace)
    return worker.enqueue_agent_self_heal(run_id, case_yaml, context, case_id=task_dict.get("case_id"))


@app.post("/api/runs/{run_id}/agent-explore/promote-candidate")
def promote_agent_explore_candidate(run_id: str) -> dict:
    """资产流通闭环（路线 D · 增量）：把 agent-explore 生成的 candidate_flow.py
    提升为 case_drafts 草稿。

    - 读取 ``reports/agent-explore/{run_id}/candidate_flow.py``
    - 解析为 case_draft YAML 顶层骨架（标题默认用 run_id；status=draft；template=spec）
    - 复制 candidate_flow.py 全文到 ``operation_steps`` 注释区，方便后续手工补全
    - 失败：文件不存在 → 404；内容不是 UTF-8 → 400
    - 成功：返回新建 case_draft 详情（同 ``case_draft_detail``）
    """
    candidate_path = ROOT / "reports" / "agent-explore" / run_id / "candidate_flow.py"
    if not candidate_path.exists():
        raise HTTPException(
            status_code=404,
            detail=f"candidate flow not found for run_id={run_id}",
        )
    try:
        candidate_text = candidate_path.read_text(encoding="utf-8")
    except (UnicodeDecodeError, OSError) as exc:
        raise HTTPException(status_code=400, detail=f"candidate_flow.py is not readable: {exc}") from exc
    if not candidate_text.strip():
        raise HTTPException(status_code=400, detail="candidate_flow.py is empty")

    # 骨架：top-level 字段 + 占位步骤骨架 + 注释区贴候选脚本原文
    # 避免 Python 文件里出现非 ASCII YAML 转义问题，统一 allow_unicode
    draft_payload: dict = {
        "title": run_id,
        "template": "spec",
        "status": "draft",
        "priority": "P2",
        "type": "智能探索采纳",
        "source": "agent-explore",
        "source_run_id": run_id,
        "steps": [
            f"待补充：从 {run_id} 的 candidate_flow 拷贝逻辑",
        ],
        "_candidate_flow_excerpt": candidate_text,
    }
    yaml_text = yaml.safe_dump(draft_payload, allow_unicode=True, sort_keys=False)

    now = utc_now()
    title = run_id[:200]
    requirement_id = ensure_manual_requirement()
    with connect() as conn:
        cur = conn.execute(
            """
            insert into case_drafts(requirement_id, title, yaml, status, created_at, updated_at, template)
            values (?, ?, ?, 'draft', ?, ?, 'spec')
            """,
            (requirement_id, title, yaml_text, now, now),
        )
        new_id = int(cur.lastrowid)
    return case_draft_detail(new_id)


def evidence_log_lines(run_id: str) -> list[str]:
    summary = evidence_summary(run_id)
    lines: list[str] = []
    for item in summary.get("events", {}).get("latest", []):
        lines.append(f"[evidence:event] {item.get('kind', '')} {item.get('message', '')} {item.get('url', '')}")
    for item in summary.get("console", {}).get("latest", []):
        lines.append(f"[evidence:console] {item.get('level', '')} {item.get('text', '')}")
    for item in summary.get("network", {}).get("latest", []):
        lines.append(f"[evidence:network] {item.get('method', '')} {item.get('status', '')} {item.get('url', '')}")
    dom_count = summary.get("dom", {}).get("count", 0)
    if summary.get("trace", {}).get("exists"):
        lines.append("[evidence:trace] trace.zip generated")
    if dom_count:
        lines.append(f"[evidence:dom] {dom_count} DOM snapshots generated")
    return lines


def _synthetic_logs_from_evidence(run_id: str) -> list[dict]:
    evidence_root = EVIDENCE_ROOT / run_id
    events = read_jsonl(evidence_root / "events.jsonl", limit=100000)
    console = read_jsonl(evidence_root / "console.jsonl", limit=100000)
    rows: list[dict] = []
    index = 1
    for item in events:
        line = f"[{item.get('kind', '')}] {item.get('message', '')}".strip()
        if item.get("value") not in (None, ""):
            line = f"{line} value={item.get('value')}"
        if item.get("url"):
            line = f"{line} url={item.get('url')}"
        rows.append(
            {
                "id": index,
                "run_id": run_id,
                "stream": "evidence",
                "line": line,
                "created_at": item.get("created_at") or "",
            }
        )
        index += 1
    for item in console:
        rows.append(
            {
                "id": index,
                "run_id": run_id,
                "stream": "console",
                "line": f"[console:{item.get('level', '')}] {item.get('text', '')}".strip(),
                "created_at": item.get("created_at") or "",
            }
        )
        index += 1
    return rows


@app.get("/api/runs/{run_id}/evidence/trace")
def run_evidence_trace(run_id: str) -> FileResponse:
    path = TRACE_ROOT / run_id / "trace.zip"
    if not path.exists():
        raise HTTPException(status_code=404, detail="trace not found")
    return FileResponse(path, media_type="application/zip", filename=f"{run_id}-trace.zip")


@app.get("/api/runs/{run_id}/evidence/{kind}")
def run_evidence_jsonl(run_id: str, kind: Literal["events", "console", "network"]) -> FileResponse:
    path = EVIDENCE_ROOT / run_id / f"{kind}.jsonl"
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"{kind} evidence not found")
    return FileResponse(path, media_type="application/jsonl", filename=f"{run_id}-{kind}.jsonl")


@app.get("/api/runs/{run_id}/evidence/dom/{filename}")
def run_evidence_dom(run_id: str, filename: str) -> FileResponse:
    path = EVIDENCE_ROOT / run_id / "dom" / filename
    if not path.exists() or path.suffix.lower() != ".html":
        raise HTTPException(status_code=404, detail="dom snapshot not found")
    return FileResponse(path, media_type="text/html", filename=filename)


@app.get("/api/runs/{run_id}/observed-asset")
def run_observed_asset(run_id: str) -> dict:
    return load_observed_asset_for_run(run_id)


@app.post("/api/runs/{run_id}/merge-observed-asset")
def merge_run_observed_asset(run_id: str) -> dict:
    with connect() as conn:
        task = conn.execute("select * from run_tasks where id = ?", (run_id,)).fetchone()
    if not task:
        raise HTTPException(status_code=404, detail="run not found")
    if task["status"] != "passed":
        raise HTTPException(status_code=400, detail="only passed runs can merge observed automation assets")
    if not task["case_id"]:
        raise HTTPException(status_code=400, detail="batch parent runs cannot merge a single case asset")

    observed = load_observed_asset_for_run(run_id)
    case_path = find_case_yaml(str(task["case_id"]))
    data = yaml.safe_load(case_path.read_text(encoding="utf-8")) or {}
    existing = data.get("automation_asset") if isinstance(data.get("automation_asset"), dict) else {}
    data["automation_asset"] = merge_automation_asset(existing, observed)
    validation = validate_case_yaml(yaml.safe_dump(data, allow_unicode=True, sort_keys=False))
    if not validation["valid"]:
        raise HTTPException(status_code=400, detail=f"merged YAML validation failed: {'; '.join(validation['errors'])}")
    case_path.write_text(yaml.safe_dump(data, allow_unicode=True, sort_keys=False), encoding="utf-8")
    return {"case_id": task["case_id"], "path": str(case_path), "automation_asset": data["automation_asset"]}


# -----------------------------------------------------------------------------
# 路线 A · 资产采纳主流程化（diff 预览 + adoptions 持久化）
# -----------------------------------------------------------------------------


def _find_latest_passed_run_for_case(case_id: str) -> dict | None:
    """从 run_tasks 查指定 case 的最新 passed run（优先用 case_runs，否则回退 run_tasks.status='passed'）"""
    with connect() as conn:
        # 优先 case_runs（更精确，记录每次执行）
        try:
            row = conn.execute(
                """
                select cr.run_id, rt.finished_at, cr.started_at
                from case_runs cr
                left join run_tasks rt on rt.id = cr.run_id
                where cr.case_id = ? and cr.passed = 1
                order by cr.started_at desc
                limit 1
                """,
                (case_id,),
            ).fetchone()
            if row:
                return {"run_id": row["run_id"], "finished_at": row["finished_at"], "started_at": row["started_at"]}
        except sqlite3.OperationalError:
            pass
        # 回退 run_tasks（case_runs 暂未落库的兼容路径）
        row = conn.execute(
            """
            select id as run_id, finished_at, started_at
            from run_tasks
            where case_id = ? and status = 'passed'
            order by coalesce(finished_at, started_at, created_at) desc
            limit 1
            """,
            (case_id,),
        ).fetchone()
        if row:
            return {"run_id": row["run_id"], "finished_at": row["finished_at"], "started_at": row["started_at"]}
    return None


def compute_observed_asset_diff(existing: dict, observed: dict) -> dict:
    """纯函数：计算 observed vs existing automation_asset 的三段 diff。

    - kept：existing 中非空且与 observed 等价的字段子集（conservative 语义）
    - added：observed 独有但 existing 缺失的字段
    - missing：YAML 验证要求必须有的字段，且 existing / observed 都缺失

    不写库，便于单测。
    """
    existing = existing if isinstance(existing, dict) else {}
    observed = observed if isinstance(observed, dict) else {}

    def _is_non_empty(value: object) -> bool:
        if value is None:
            return False
        if isinstance(value, (list, dict, str)) and len(value) == 0:
            return False
        return True

    def _equivalent(a: object, b: object) -> bool:
        """比较两个值是否等价（list 转 set 排序后再比，dict 比较 key 集合）"""
        if a is None or b is None:
            return a is None and b is None
        if isinstance(a, dict) and isinstance(b, dict):
            keys = set(a) | set(b)
            return all(_equivalent(a.get(k), b.get(k)) for k in keys)
        if isinstance(a, list) and isinstance(b, list):
            try:
                return sorted([json.dumps(x, ensure_ascii=False, sort_keys=True) for x in a]) == sorted(
                    [json.dumps(x, ensure_ascii=False, sort_keys=True) for x in b]
                )
            except TypeError:
                return a == b
        return a == b

    track_fields = ("operation_steps", "selectors", "input_values", "assertions")

    kept: dict = {}
    added: dict = {}
    for field in track_fields:
        ex_value = existing.get(field)
        ob_value = observed.get(field)
        if _is_non_empty(ex_value) and _is_non_empty(ob_value) and _equivalent(ex_value, ob_value):
            kept[field] = ex_value
        elif _is_non_empty(ob_value) and not _is_non_empty(ex_value):
            added[field] = ob_value
        elif _is_non_empty(ob_value) and _is_non_empty(ex_value):
            # 双方都有但不完全等价：保守视为 kept(existing)，并把差异的 observed 计入 added（提示用户）
            kept[field] = ex_value
            if not _equivalent(ex_value, ob_value):
                added[field] = ob_value
        # ex 非空 + ob 为空 → kept（保守不覆盖）
        elif _is_non_empty(ex_value):
            kept[field] = ex_value

    # missing：YAML 验证要求必须有，但两边都缺
    missing: list[str] = []
    if not _is_non_empty(existing.get("operation_steps")) and not _is_non_empty(observed.get("operation_steps")):
        missing.append("automation_asset.operation_steps")
    if not _is_non_empty(existing.get("assertions")) and not _is_non_empty(observed.get("assertions")):
        missing.append("automation_asset.assertions")
    if not _is_non_empty(existing.get("selectors")) and not _is_non_empty(observed.get("selectors")):
        missing.append("automation_asset.selectors")
    if not isinstance(existing.get("input_values"), dict):
        if not isinstance(observed.get("input_values"), dict):
            missing.append("automation_asset.input_values")

    return {"kept": kept, "added": added, "missing": missing}


@app.get("/api/cases/{case_id}/observed-asset-diff")
def get_observed_asset_diff(case_id: str) -> dict:
    """路线 A · T2：返回该 case 最新 passed run 的 observed vs 现有 YAML 的三段 diff。"""
    case_id_norm = normalize_case_id(case_id)
    latest = _find_latest_passed_run_for_case(case_id_norm)
    if not latest:
        raise HTTPException(status_code=404, detail="no passed run for case")
    run_id = latest["run_id"]

    try:
        observed = load_observed_asset_for_run(run_id)
    except HTTPException:
        raise
    case_path = find_case_yaml(case_id_norm)
    try:
        data = yaml.safe_load(case_path.read_text(encoding="utf-8")) or {}
    except yaml.YAMLError as exc:
        raise HTTPException(status_code=400, detail=f"case YAML parse failed: {exc}") from exc
    existing = data.get("automation_asset") if isinstance(data.get("automation_asset"), dict) else {}

    diff = compute_observed_asset_diff(existing, observed)
    return {
        "case_id": case_id_norm,
        "run_id": run_id,
        "diff": diff,
        "observed_at": observed.get("observed_at"),
    }


class AdoptionRequest(BaseModel):
    run_id: str
    mode: Literal["accept", "reject"]
    adopted_by: str | None = None


YAML_BACKUP_DIR = ROOT / ".codex-tmp" / "yaml-backup"


def _backup_yaml_before_write(case_id: str, case_path: Path) -> Path:
    """在覆盖 YAML 前备份到 .codex-tmp/yaml-backup/{case_id}-{ts}.yaml"""
    YAML_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    target = YAML_BACKUP_DIR / f"{case_id}-{ts}.yaml"
    target.write_text(case_path.read_text(encoding="utf-8"), encoding="utf-8")
    return target


@app.post("/api/cases/{case_id}/adoptions")
def post_adoption(case_id: str, payload: AdoptionRequest) -> dict:
    """路线 A · T3：accept → 合并 + 备份 + 落盘；reject → 仅写 asset_adoptions"""
    case_id_norm = normalize_case_id(case_id)
    run_id = payload.run_id
    if not run_id:
        raise HTTPException(status_code=400, detail="run_id required")

    with connect() as conn:
        task = conn.execute("select * from run_tasks where id = ?", (run_id,)).fetchone()
        if not task:
            raise HTTPException(status_code=404, detail="run not found")
        if task["status"] != "passed":
            raise HTTPException(status_code=400, detail="only passed runs can be adopted")

    case_path = find_case_yaml(case_id_norm)
    try:
        original_text = case_path.read_text(encoding="utf-8")
        data = yaml.safe_load(original_text) or {}
    except yaml.YAMLError as exc:
        raise HTTPException(status_code=400, detail=f"case YAML parse failed: {exc}") from exc
    existing = data.get("automation_asset") if isinstance(data.get("automation_asset"), dict) else {}

    diff_summary: dict = {"kept": 0, "added": 0, "missing": 0}

    if payload.mode == "accept":
        try:
            observed = load_observed_asset_for_run(run_id)
        except HTTPException:
            raise
        diff = compute_observed_asset_diff(existing, observed)
        diff_summary = {
            "kept": len(diff.get("kept", {})),
            "added": len(diff.get("added", {})),
            "missing": len(diff.get("missing", [])),
        }

        # 备份原文件
        backup_path = _backup_yaml_before_write(case_id_norm, case_path)
        try:
            data["automation_asset"] = merge_automation_asset(existing, observed)
            new_text = yaml.safe_dump(data, allow_unicode=True, sort_keys=False)
            validation = validate_case_yaml(new_text)
            if not validation["valid"]:
                # 回滚
                backup_text = backup_path.read_text(encoding="utf-8")
                case_path.write_text(backup_text, encoding="utf-8")
                raise HTTPException(
                    status_code=400,
                    detail=f"merged YAML validation failed: {'; '.join(validation['errors'])}",
                )
            case_path.write_text(new_text, encoding="utf-8")
        except Exception:
            # 任何异常都回滚
            if backup_path.exists():
                backup_text = backup_path.read_text(encoding="utf-8")
                case_path.write_text(backup_text, encoding="utf-8")
            raise
    else:
        # reject：YAML 不写，diff_summary 留空
        diff_summary = {"kept": 0, "added": 0, "missing": 0, "rejected": True}

    with connect() as conn:
        cur = conn.execute(
            """
            insert into asset_adoptions(case_id, run_id, mode, diff_summary_json, adopted_by, adopted_at)
            values (?, ?, ?, ?, ?, ?)
            """,
            (
                case_id_norm,
                run_id,
                payload.mode,
                json.dumps(diff_summary, ensure_ascii=False),
                payload.adopted_by,
                utc_now(),
            ),
        )
        adoption_id = cur.lastrowid

    result = {
        "case_id": case_id_norm,
        "run_id": run_id,
        "mode": payload.mode,
        "asset_adoption_id": adoption_id,
        "diff_summary": diff_summary,
    }
    if payload.mode == "accept":
        result["yaml_path"] = str(case_path)
    return result


@app.get("/api/cases/{case_id}/adoptions")
def get_adoptions(case_id: str, limit: int = Query(default=10, ge=1, le=50)) -> list[dict]:
    """路线 A · T3：按 adopted_at DESC 返回该 case 最近 N 条采纳历史"""
    case_id_norm = normalize_case_id(case_id)
    with connect() as conn:
        rows = conn.execute(
            """
            select id, case_id, run_id, mode, diff_summary_json, adopted_by, adopted_at
            from asset_adoptions
            where case_id = ?
            order by adopted_at desc
            limit ?
            """,
            (case_id_norm, limit),
        ).fetchall()
    results: list[dict] = []
    for row in rows:
        item = dict(row)
        try:
            item["diff_summary"] = json.loads(item.pop("diff_summary_json") or "null")
        except (TypeError, json.JSONDecodeError):
            item["diff_summary"] = None
        results.append(item)
    return results


# 路线 C · T11 / T12：YAML → Python codegen 端点
import py_compile  # 内置模块；用 import 形式方便 monkey-patch 测试
import ast

CODEGEN_TEMPLATE_REL = Path("runner") / "flows" / "templates" / "icm_case.py.j2"
FLOW_BACKUP_DIR = ROOT / ".codex-tmp" / "flow-backup"
CODEGEN_TEMPLATE_DIR = ROOT / "runner" / "flows" / "templates"

# 关键词正则清单（架构 §2 C2）：每个 tuple = (regex, dispatch_kind)
# dispatch_kind 仅用于 validate_operation_steps 的命中判定；模板内部用相同正则。
CODEGEN_KEYWORD_RULES: tuple[tuple[re.Pattern[str], str], ...] = (
    (re.compile(r"登录|prepare|已登录|确保登录|login"), "prepare_session"),
    (re.compile(r"打开|open|navigate|访问|进入|跳到|跳至"), "open_route"),
    (re.compile(r"搜索|查询|query|keyword|关键字"), "search"),
    (re.compile(r"输入|填写|fill|enter|键入"), "fill"),
    (re.compile(r"点击|click|按下|触发|submit|提交"), "click"),
    (re.compile(r"断言|assert|verify|确认|验证|可见"), "assert"),
    (re.compile(r"settle|等待|wait", re.IGNORECASE), "settle"),
)


class CodegenRequest(BaseModel):
    write: bool = False
    template: Literal["functional", "negative", "regression"] = "functional"


def compute_target_path(case_id: str) -> Path:
    """架构 §2 C3 命名规则：runner/flows/icm_case_XXX.py

    XXX 默认从 case_id 提取 TC-ICM- 后面的数字段。
    提取不到时回退到末段数字；都没有则用 case_id 全小写。
    """
    match = re.search(r"TC-ICM-(\d+)", case_id, flags=re.IGNORECASE)
    if match:
        suffix = match.group(1)
    else:
        tail = re.search(r"(\d+)$", case_id)
        suffix = tail.group(1) if tail else (case_id.lower() or "case")
    return ROOT / "runner" / "flows" / f"icm_case_{suffix}.py"


def validate_operation_steps(steps: list[str] | None) -> list[str]:
    """架构 §2 C2：未命中关键词的步骤进 errors（不允许落盘）。

    返回错误列表；空 list 表示全部命中。
    """
    if not steps:
        return ["missing operation_steps"]
    errors: list[str] = []
    for idx, raw in enumerate(steps):
        text = (raw or "").strip() if isinstance(raw, str) else ""
        if not text:
            errors.append(f"operation_steps[{idx}] is empty")
            continue
        if not any(rule.search(text) for rule, _ in CODEGEN_KEYWORD_RULES):
            errors.append(f"unsupported step kind: {text}")
    return errors


def _render_codegen_template(
    case_id: str,
    title: str,
    system_id: str,
    operation_steps: list[str],
    selectors: dict,
    input_values: dict,
    assertions: list[str],
    template_name: str,
) -> str:
    """用 Jinja2 渲染模板到内存字符串（架构 §2 C3 步骤 4）。"""
    from jinja2 import Environment, FileSystemLoader, StrictUndefined

    env = Environment(
        loader=FileSystemLoader(str(CODEGEN_TEMPLATE_DIR)),
        undefined=StrictUndefined,
        keep_trailing_newline=True,
        autoescape=False,
    )
    tpl = env.get_template("icm_case.py.j2")
    return tpl.render(
        case_id=case_id,
        title=title,
        system_id=system_id,
        operation_steps=list(operation_steps or []),
        selectors=dict(selectors or {}),
        input_values=dict(input_values or {}),
        assertions=list(assertions or []),
        template=template_name,
    )


def _append_codegen_log(case_id: str, message: str) -> None:
    """所有 codegen dry-run / write 进 run_logs，关键字前缀 codegen（架构 §8）。

    run_id 使用 'codegen-{case_id}-{ts}' 避免与真实 run 冲突。
    """
    try:
        ts = datetime.utcnow().strftime("%Y%m%d%H%M%S")
        with connect() as conn:
            conn.execute(
                """
                insert into run_logs(run_id, stream, line, created_at)
                values (?, ?, ?, ?)
                """,
                (
                    f"codegen-{case_id}-{ts}",
                    "codegen",
                    message,
                    utc_now(),
                ),
            )
    except Exception:  # pragma: no cover - 防御
        pass


def _backup_flow_before_write(target: Path, case_id: str) -> Path | None:
    """落盘前备份：.codex-tmp/flow-backup/{case_id}-{ts}.py（首次生成时跳过）。"""
    if not target.exists():
        return None
    FLOW_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    backup = FLOW_BACKUP_DIR / f"{case_id}-{ts}.py"
    backup.write_text(target.read_text(encoding="utf-8"), encoding="utf-8")
    return backup


@app.post("/api/cases/{case_id}/codegen")
def post_codegen(case_id: str, payload: CodegenRequest | None = None) -> dict:
    """路线 C · T11 / T12：dry-run / 落盘 + py_compile 自检。

    dry-run (write=false): 渲染模板到内存，py_compile 自检，**不**落盘。
    write=true: 备份旧 .py → 写盘 → 再次 py_compile → 失败回滚。提示人工检查 cases.py。
    """
    case_id_norm = normalize_case_id(case_id)
    payload = payload or CodegenRequest()
    target_path = compute_target_path(case_id_norm)

    # 1) 加载 case YAML
    try:
        case_path = find_case_yaml(case_id_norm)
    except HTTPException:
        return {
            "ok": False,
            "code": "",
            "target_path": str(target_path),
            "errors": [f"case YAML not found: {case_id_norm}"],
            "warnings": [],
        }

    try:
        data = yaml.safe_load(case_path.read_text(encoding="utf-8")) or {}
    except yaml.YAMLError as exc:
        return {
            "ok": False,
            "code": "",
            "target_path": str(target_path),
            "errors": [f"case YAML parse failed: {exc}"],
            "warnings": [],
        }

    asset = data.get("automation_asset") or {}
    operation_steps: list[str] = [str(s) for s in (asset.get("operation_steps") or [])]
    selectors: dict = dict(asset.get("selectors") or {})
    input_values: dict = dict(asset.get("input_values") or {})
    assertions: list[str] = [str(a) for a in (asset.get("assertions") or [])]
    system_id = str(data.get("system") or "icm-internal")
    title = str(data.get("title") or case_id_norm)

    # 2) 优先拦截空 operation_steps（架构 §10 验收 #11 + PRD §4 REQ-C-05 细分错误码）
    if not operation_steps:
        _append_codegen_log(case_id_norm, "codegen missing operation_steps")
        return {
            "ok": False,
            "code": "",
            "target_path": str(target_path),
            "errors": ["missing operation_steps"],
            "warnings": [],
        }

    # 3) validate_case_yaml（基础字段 / selectors / assertions 等）
    validation = validate_case_yaml(yaml.safe_dump(data, allow_unicode=True, sort_keys=False))
    if not validation["valid"]:
        return {
            "ok": False,
            "code": "",
            "target_path": str(target_path),
            "errors": list(validation.get("errors", [])),
            "warnings": list(validation.get("warnings", [])),
        }

    # 4) validate_operation_steps（关键词派发校验；架构 §2 C2）
    step_errors = validate_operation_steps(operation_steps)

    # 5) 渲染模板到内存
    try:
        code = _render_codegen_template(
            case_id=case_id_norm,
            title=title,
            system_id=system_id,
            operation_steps=operation_steps,
            selectors=selectors,
            input_values=input_values,
            assertions=assertions,
            template_name=payload.template,
        )
    except Exception as exc:
        _append_codegen_log(case_id_norm, f"codegen render failed: {exc}")
        return {
            "ok": False,
            "code": "",
            "target_path": str(target_path),
            "errors": [f"template render failed: {exc}"],
            "warnings": list(validation.get("warnings", [])),
        }

    # 6) 内存源码语法自检（满足 PRD §8 安全沙箱）
    try:
        ast.parse(code)
    except SyntaxError as exc:
        _append_codegen_log(case_id_norm, f"codegen syntax check failed (memory): {exc}")
        return {
            "ok": False,
            "code": code,
            "target_path": str(target_path),
            "errors": [f"py_compile failed: {exc}"],
            "warnings": list(validation.get("warnings", [])),
        }

    result: dict = {
        "ok": not step_errors,
        "code": code,
        "target_path": str(target_path),
        "errors": step_errors,
        "warnings": list(validation.get("warnings", [])),
    }

    # 6) 落盘流程
    if not payload.write:
        result["written"] = False
        _append_codegen_log(
            case_id_norm,
            f"codegen dry-run ok ({len(operation_steps)} steps, errors={len(step_errors)})",
        )
        return result

    # write=true：必须通过关键词校验才允许落盘
    if step_errors:
        result["ok"] = False
        result["written"] = False
        result["errors"] = step_errors + ["refused: operation_steps validation failed"]
        _append_codegen_log(case_id_norm, f"codegen write refused: {step_errors}")
        return result

    # 备份旧文件
    backup_path = _backup_flow_before_write(target_path, case_id_norm)
    wrote_existed_before = target_path.exists()

    # 写盘
    try:
        target_path.parent.mkdir(parents=True, exist_ok=True)
        target_path.write_text(code, encoding="utf-8")
    except Exception as exc:
        _append_codegen_log(case_id_norm, f"codegen write failed: {exc}")
        raise HTTPException(status_code=500, detail=f"failed to write flow: {exc}") from exc

    # 再次 py_compile 落盘文件
    try:
        py_compile.compile(str(target_path), doraise=True)
    except py_compile.PyCompileError as exc:
        # 回滚：若之前存在则恢复；否则删除（满足 P1 mock 覆盖目标）
        try:
            if backup_path is not None and backup_path.exists():
                target_path.write_text(backup_path.read_text(encoding="utf-8"), encoding="utf-8")
                _append_codegen_log(case_id_norm, f"codegen rollback: restored from {backup_path}")
            elif not wrote_existed_before:
                target_path.unlink(missing_ok=True)
                _append_codegen_log(case_id_norm, "codegen rollback: removed newly created file (no prior content)")
            else:
                # 理论上不会进此分支（wrote_existed_before==True 时 backup_path 必非 None），
                # 但保留防御性清理。
                target_path.unlink(missing_ok=True)
                _append_codegen_log(case_id_norm, "codegen rollback: removed inconsistent file (defensive)")
        except Exception as rollback_exc:  # pragma: no cover - 防御
            _append_codegen_log(
                case_id_norm,
                f"codegen rollback FAILED: {rollback_exc}; manual cleanup may be required for {target_path}",
            )
            raise HTTPException(
                status_code=500,
                detail=f"py_compile failed and rollback also failed: {rollback_exc}",
            ) from exc
        _append_codegen_log(case_id_norm, f"codegen write py_compile failed, rolled back: {exc}")
        return {
            "ok": False,
            "code": code,
            "target_path": str(target_path),
            "errors": [f"py_compile on disk failed, rolled back: {exc}"],
            "warnings": [],
        }

    _append_codegen_log(case_id_norm, f"codegen write ok: {target_path}")
    result["written"] = True
    result["backup_path"] = str(backup_path) if backup_path is not None else None
    result["message"] = "已写入，请人工检查 runner/cases.py:CASE_RUNNERS 是否需要追加"
    return result


# 路线 B · T7 / T8：稳定性计算 + scan 端点
STABILITY_THRESHOLD_DEFAULT = 0.95
STABILITY_UNSTABLE_THRESHOLD_DEFAULT = 0.80
STABILITY_INSUFFICIENT_THRESHOLD = 5  # 样本 < 5 视为 insufficient
SCAN_LOG_DIR = ROOT / "platform-data" / "runner-logs"


def _utc_now_iso() -> str:
    return utc_now()


def _compute_stability(case_id: str, window: int = 20) -> dict:
    """从 case_runs 派生稳定分（纯函数，不写库）。"""
    window = max(1, min(int(window or 20), 200))
    with connect() as conn:
        rows = conn.execute(
            """
            select passed, started_at, finished_at
            from case_runs
            where case_id = ?
            order by coalesce(started_at, finished_at) desc, id desc
            limit ?
            """,
            (case_id, window),
        ).fetchall()

    total = len(rows)
    passed = sum(1 for r in rows if int(r["passed"] or 0) == 1)
    pass_rate = (passed / total) if total else 0.0

    last_passed_at: str | None = None
    last_failed_at: str | None = None
    for r in rows:
        ts = r["finished_at"] or r["started_at"]
        if last_passed_at is None and int(r["passed"] or 0) == 1:
            last_passed_at = ts
        elif last_failed_at is None and int(r["passed"] or 0) == 0:
            last_failed_at = ts
        if last_passed_at and last_failed_at:
            break

    if total < STABILITY_INSUFFICIENT_THRESHOLD:
        status = "insufficient"
    elif pass_rate >= STABILITY_THRESHOLD_DEFAULT:
        status = "stable"
    elif pass_rate >= STABILITY_UNSTABLE_THRESHOLD_DEFAULT:
        status = "flaky"
    else:
        status = "unstable"

    return {
        "case_id": case_id,
        "total": total,
        "passed": passed,
        "pass_rate": round(pass_rate, 4),
        "status": status,
        "last_passed_at": last_passed_at,
        "last_failed_at": last_failed_at,
        "thresholds": {
            "stable": STABILITY_THRESHOLD_DEFAULT,
            "unstable": STABILITY_UNSTABLE_THRESHOLD_DEFAULT,
        },
        "insufficient_threshold": STABILITY_INSUFFICIENT_THRESHOLD,
        "window": window,
    }


@app.get("/api/cases/{case_id}/stability")
def get_case_stability(case_id: str, window: int = Query(default=20, ge=1, le=200)) -> dict:
    """路线 B · T7：从 case_runs 派生稳定分（默认窗口 20）。"""
    case_id_norm = normalize_case_id(case_id)
    return _compute_stability(case_id_norm, window=window)


def _append_scan_log(scan_id: str, message: str) -> None:
    try:
        SCAN_LOG_DIR.mkdir(parents=True, exist_ok=True)
        log_path = SCAN_LOG_DIR / f"{scan_id}.log"
        with log_path.open("a", encoding="utf-8") as fp:
            fp.write(f"[{_utc_now_iso()}] {message}\n")
    except Exception:  # pragma: no cover - 防御
        pass


def _run_stability_scan_thread(scan_id: str, case_id: str, times: int) -> None:
    """线程入口：在 stability_scans 中跑 N 次 case，写 case_runs，更新进度。"""
    _append_scan_log(scan_id, f"start scan_id={scan_id} case_id={case_id} times={times}")
    completed = 0
    passed_count = 0
    try:
        with connect() as conn:
            conn.execute("update stability_scans set status = 'running' where id = ?", (scan_id,))

        for i in range(1, times + 1):
            run_id = f"{scan_id}-{i:02d}"
            _append_scan_log(scan_id, f"running attempt {i}/{times} run_id={run_id}")
            try:
                proc = subprocess.run(
                    [sys.executable, "-m", "runner.main", "run-case", case_id, run_id, "--retry", "0"],
                    cwd=str(ROOT),
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=600,
                    check=False,
                )
                # 从 case_runs 查最新一行（runner 每次 attempt 都会写）
                with connect() as conn:
                    row = conn.execute(
                        """
                        select passed from case_runs
                        where case_id = ? and run_id = ?
                        order by id desc limit 1
                        """,
                        (case_id, run_id),
                    ).fetchone()
                if row is None:
                    _append_scan_log(scan_id, f"attempt {i} no case_runs row (rc={proc.returncode})")
                else:
                    if int(row["passed"] or 0) == 1:
                        passed_count += 1
                    _append_scan_log(
                        scan_id, f"attempt {i} passed={int(row['passed'] or 0)} rc={proc.returncode}"
                    )
            except subprocess.TimeoutExpired:
                _append_scan_log(scan_id, f"attempt {i} timed out")
            except Exception as exc:
                _append_scan_log(scan_id, f"attempt {i} error: {exc}")

            completed = i
            try:
                with connect() as conn:
                    conn.execute(
                        "update stability_scans set completed = ?, passed = ? where id = ?",
                        (completed, passed_count, scan_id),
                    )
            except Exception:
                pass

        final_status = "done" if completed == times else "failed"
        with connect() as conn:
            conn.execute(
                """
                update stability_scans
                set status = ?, completed = ?, passed = ?, finished_at = ?
                where id = ?
                """,
                (final_status, completed, passed_count, _utc_now_iso(), scan_id),
            )
        _append_scan_log(
            scan_id, f"scan {final_status} completed={completed} passed={passed_count}/{times}"
        )
    except Exception as exc:
        try:
            with connect() as conn:
                conn.execute(
                    """
                    update stability_scans
                    set status = 'failed', finished_at = ?
                    where id = ?
                    """,
                    (_utc_now_iso(), scan_id),
                )
        except Exception:
            pass
        _append_scan_log(scan_id, f"scan crashed: {exc}")


def _start_stability_scan(case_id: str, times: int) -> dict:
    """辅助函数（T7/T8 共享）：创建 stability_scans 行 + 启线程，返回 scan_id。"""
    if times < 1:
        raise HTTPException(status_code=400, detail="times must be >= 1")
    if times > 50:
        raise HTTPException(status_code=400, detail="times must be <= 50")

    scan_id = f"scan-{uuid.uuid4().hex[:12]}"
    now = _utc_now_iso()
    with connect() as conn:
        conn.execute(
            """
            insert into stability_scans(id, case_id, status, times, completed, passed, created_at)
            values (?, ?, 'queued', ?, 0, 0, ?)
            """,
            (scan_id, case_id, times, now),
        )
    thread = threading.Thread(
        target=_run_stability_scan_thread,
        args=(scan_id, case_id, times),
        name=f"icm-stability-scan-{scan_id}",
        daemon=True,
    )
    thread.start()
    return {
        "scan_id": scan_id,
        "case_id": case_id,
        "status": "queued",
        "times": times,
        "started_at": now,
    }


class StabilityScanRequest(BaseModel):
    times: int = 10


@app.post("/api/cases/{case_id}/stability-scan")
def post_stability_scan(case_id: str, payload: StabilityScanRequest | None = None) -> dict:
    """路线 B · T7：触发 stability scan。times 默认 10。返回 scan_id（异步执行不阻塞）。"""
    case_id_norm = normalize_case_id(case_id)
    times = (payload.times if payload else 10) or 10
    return _start_stability_scan(case_id_norm, times)


@app.post("/api/cases/{case_id}/recompute-stability")
def post_recompute_stability(case_id: str) -> dict:
    """路线 B · T8：固定 N=10 次重跑（与 PRD US-B4 "跑 10 次" 对齐）。"""
    case_id_norm = normalize_case_id(case_id)
    return _start_stability_scan(case_id_norm, times=10)


@app.get("/api/stability-scans/{scan_id}")
def get_stability_scan(scan_id: str) -> dict:
    with connect() as conn:
        row = conn.execute(
            """
            select id, case_id, status, times, completed, passed, created_at, finished_at
            from stability_scans
            where id = ?
            """,
            (scan_id,),
        ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="scan not found")
    return dict(row)


def _element_knowledge_payload(refresh_summary: dict | None = None) -> dict:
    library = load_json_file(ELEMENT_LIBRARY_PATH)
    summary = refresh_summary or load_json_file(ELEMENT_SUMMARY_PATH)
    model = build_report_model(library, summary)
    return {
        "summary": model,
        "elements": library.get("elements") or [],
        "hotspots": model.get("hotspots") or [],
        "report_paths": {
            "markdown": summary.get("markdown_report_path") or str(ROOT / "reports" / "element-library" / "refresh-report.md"),
            "html": summary.get("html_report_path") or str(ROOT / "reports" / "element-library" / "refresh-report.html"),
        },
        "source_paths": {
            "library": str(ELEMENT_LIBRARY_PATH),
            "summary": str(ELEMENT_SUMMARY_PATH),
        },
        "exists": {
            "library": ELEMENT_LIBRARY_PATH.exists(),
            "summary": ELEMENT_SUMMARY_PATH.exists(),
        },
    }


@app.get("/api/element-knowledge")
def get_element_knowledge() -> dict:
    return _element_knowledge_payload()


def _append_element_refresh_log(task_id: str, message: str) -> None:
    with connect() as conn:
        conn.execute(
            "insert into run_logs(run_id, stream, line, created_at) values (?, ?, ?, ?)",
            (task_id, "element-knowledge", message, utc_now()),
        )


def _append_element_refresh_progress(task_id: str, progress: dict) -> None:
    payload = {"kind": "progress", **progress, "updated_at": utc_now()}
    _append_element_refresh_log(task_id, "progress " + json.dumps(payload, ensure_ascii=False, sort_keys=True))


def _latest_element_refresh_progress(logs: list[dict]) -> dict | None:
    for log in reversed(logs):
        line = str(log.get("line") or "")
        if not line.startswith("progress "):
            continue
        try:
            payload = json.loads(line[len("progress ") :])
        except Exception:
            continue
        if isinstance(payload, dict):
            return payload
    return None


def _element_refresh_task_payload(task_id: str) -> dict:
    with connect() as conn:
        task = conn.execute("select * from run_tasks where id = ?", (task_id,)).fetchone()
        logs = conn.execute("select * from run_logs where run_id = ? order by id", (task_id,)).fetchall()
    if not task:
        raise HTTPException(status_code=404, detail="element knowledge refresh task not found")
    task_dict = dict(task)
    log_dicts = rows_to_dicts(logs)
    return {
        **task_dict,
        "logs": log_dicts,
        "progress": _latest_element_refresh_progress(log_dicts),
        "snapshot": _element_knowledge_payload() if task_dict.get("status") in {"done", "passed"} else None,
    }


def _cdp_version_url(endpoint: str) -> str:
    parsed = urlparse(str(endpoint or "").strip())
    if parsed.scheme not in {"http", "https"} or parsed.hostname not in {"127.0.0.1", "localhost", "::1"}:
        raise ValueError("cdp_endpoint must use a loopback host such as http://127.0.0.1:9222")
    try:
        port = parsed.port or _DEDICATED_CDP_DEFAULT_PORT
    except ValueError as exc:
        raise ValueError("cdp_endpoint has an invalid port") from exc
    return f"{parsed.scheme}://{parsed.hostname}:{port}/json/version"


def _cdp_browser_ready(endpoint: str) -> bool:
    try:
        with urlopen(_cdp_version_url(endpoint), timeout=0.8) as response:
            payload = json.loads(response.read().decode("utf-8"))
        return bool(payload.get("webSocketDebuggerUrl"))
    except (OSError, URLError, ValueError, json.JSONDecodeError):
        return False


def _find_chrome_executable() -> Path | None:
    candidates = [
        Path("C:/Program Files/Google/Chrome/Application/chrome.exe"),
        Path("C:/Program Files (x86)/Google/Chrome/Application/chrome.exe"),
    ]
    return next((path for path in candidates if path.exists()), None)


def _ensure_dedicated_cdp_browser(endpoint: str) -> str:
    """Start the element-library Chrome only when its local CDP endpoint is absent."""
    if _cdp_browser_ready(endpoint):
        return "reused"
    version_url = _cdp_version_url(endpoint)
    port = urlparse(version_url).port or _DEDICATED_CDP_DEFAULT_PORT
    chrome_path = _find_chrome_executable()
    if not chrome_path:
        raise RuntimeError("Google Chrome was not found; install Chrome before starting the dedicated CDP browser")
    profile_path = _DEDICATED_CDP_PROFILE_ROOT if port == _DEDICATED_CDP_DEFAULT_PORT else ROOT / "platform-data" / f"chrome-element-scan-{port}"
    profile_path.mkdir(parents=True, exist_ok=True)
    subprocess.Popen(
        [
            str(chrome_path),
            "--remote-debugging-address=127.0.0.1",
            f"--remote-debugging-port={port}",
            f"--user-data-dir={profile_path}",
        ],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    for _ in range(40):
        if _cdp_browser_ready(endpoint):
            return "started"
        time.sleep(0.25)
    raise RuntimeError(f"Dedicated CDP Chrome did not become ready at {version_url}; close stale dedicated Chrome windows and retry")


async def _scan_element_knowledge_async(
    *,
    base_url: str,
    include_states: bool,
    headless: bool,
    min_healing_failures: int,
    environment_id: str | None = None,
    target_url: str | None = None,
    target_page_id: str | None = None,
    target_name: str | None = None,
    progress_callback=None,
) -> dict:
    system_profile = None
    state_scan_max_states = 8
    state_scan_max_per_kind = 2
    if environment_id:
        system_profile = resolve_scan_settings(environment_id)
        system_profile = with_account_credentials(system_profile)
        base_url = str(system_profile.get("base_url") or base_url)
        headless = bool(system_profile.get("headless", headless))
        state_scan_max_states = max(1, int(system_profile.get("state_scan_max_states") or state_scan_max_states))
        state_scan_max_per_kind = max(1, int(system_profile.get("state_scan_max_per_kind") or state_scan_max_per_kind))
        if progress_callback:
            progress_callback({"stage": "environment_loaded", "environment_id": environment_id, "base_url": base_url, "storage_state": system_profile.get("storage_state", "")})
    auth_mode = str((system_profile or {}).get("auth_mode") or "auto_login")
    if auth_mode == "cdp_attach":
        cdp_endpoint = str((system_profile or {}).get("cdp_endpoint") or "")
        if progress_callback:
            progress_callback({"stage": "starting_cdp_browser", "base_url": base_url, "environment_id": environment_id, "cdp_endpoint": cdp_endpoint})
        cdp_browser_status = await asyncio.to_thread(_ensure_dedicated_cdp_browser, cdp_endpoint)
        if progress_callback:
            progress_callback({"stage": "cdp_browser_ready", "cdp_endpoint": cdp_endpoint, "browser_status": cdp_browser_status})
            progress_callback({"stage": "attaching_cdp_browser", "base_url": base_url, "environment_id": environment_id, "cdp_endpoint": cdp_endpoint})
        session = await attach_browser_over_cdp(cdp_endpoint)
    else:
        if progress_callback:
            progress_callback({"stage": "launching_browser", "base_url": base_url, "headless": headless, "environment_id": environment_id})
        session = await launch_browser(headless=headless, system=system_profile, reuse_storage_state=True)
    try:
        if progress_callback:
            progress_callback({"stage": "browser_attached" if auth_mode == "cdp_attach" else "browser_launched", "base_url": base_url, "headless": False if auth_mode == "cdp_attach" else headless})
        if auth_mode == "cdp_attach":
            if not await is_login_state_valid(session.page, system_profile or {}):
                raise RuntimeError("CDP Chrome is not authenticated; log in to the dedicated Chrome profile and retry the refresh")
            if progress_callback:
                progress_callback({"stage": "cdp_login_verified", "cdp_endpoint": str((system_profile or {}).get("cdp_endpoint") or "")})
        elif system_profile and (not target_url or not is_login_url(str(target_url))):
            await ensure_storage_state_for_profile(session, system_profile, progress_callback=progress_callback)
        targets = None
        if target_url:
            targets = [
                {
                    "page_id": str(target_page_id or "login").strip() or "login",
                    "name": str(target_name or target_page_id or "被测系统页面").strip() or "被测系统页面",
                    "route": str(target_url).strip(),
                    "url": str(target_url).strip(),
                }
            ]
            if progress_callback:
                progress_callback({"stage": "explicit_target_loaded", "target_url": str(target_url).strip(), "target_page_id": targets[0]["page_id"], "target_name": targets[0]["name"]})
        elif system_profile:
            targets = build_scan_targets_from_profile(system_profile)
            if not targets:
                raise ValueError("environment page list is empty; configure tested-system pages in configs/environments/*.json before scanning")
            if system_profile.get("auto_discover_routes"):
                await session.page.goto(str(targets[0]["url"]), wait_until="domcontentloaded", timeout=8000)
                discovered = await discover_routes(
                    session.page,
                    base_url=base_url,
                    max_actions=max(1, int(system_profile.get("route_discovery_max_actions") or 24)),
                    max_pages=max(1, int(system_profile.get("route_discovery_max_pages") or 20)),
                    progress_callback=progress_callback,
                )
                known_urls = {item.get("url") for item in targets}
                targets.extend(item for item in discovered if item.get("url") not in known_urls)
                if progress_callback:
                    progress_callback({"stage": "route_discovery_completed", "discovered_page_count": len(discovered), "page_total": len(targets)})
        return await refresh_element_knowledge(
            page=session.page,
            base_url=base_url,
            targets=targets,
            include_states=include_states,
            state_scan_max_states=state_scan_max_states,
            state_scan_max_per_kind=state_scan_max_per_kind,
            scan=True,
            min_healing_failures=max(1, int(min_healing_failures or 1)),
            preserve_unscanned_pages=bool(target_url or system_profile),
            progress_callback=progress_callback,
        )
    finally:
        if progress_callback:
            progress_callback({"stage": "closing_browser"})
        await close_browser(session)


def _run_element_refresh_task(task_id: str, *, no_scan: bool, min_healing_failures: int, base_url: str = "", environment_id: str | None = None, target_url: str | None = None, target_page_id: str | None = None, target_name: str | None = None, include_states: bool = False, headless: bool = True) -> None:
    try:
        with connect() as conn:
            conn.execute("update run_tasks set status = 'running', started_at = ? where id = ?", (utc_now(), task_id))
        started_monotonic = time.perf_counter()
        _append_element_refresh_log(task_id, "refresh started")
        _append_element_refresh_progress(task_id, {"stage": "refresh_started", "no_scan": no_scan, "include_states": include_states, "headless": headless})
        if no_scan:
            _append_element_refresh_log(task_id, "loading existing library and feedback")
            _append_element_refresh_progress(task_id, {"stage": "loading_existing_library"})
            summary = refresh_library_file(min_healing_failures=max(1, int(min_healing_failures or 1)))
        else:
            safe_base_url = str(base_url or "").strip()
            safe_target_url = str(target_url or "").strip()
            if not safe_base_url and not environment_id and not safe_target_url:
                raise ValueError("base_url, environment_id or target_url is required for browser scan refresh")
            _append_element_refresh_log(task_id, f"browser scan started: base_url={safe_base_url or '[environment]'} environment_id={environment_id or ''} target_url={safe_target_url} include_states={include_states} headless={headless}")
            summary = asyncio.run(
                _scan_element_knowledge_async(
                    base_url=safe_base_url,
                    include_states=include_states,
                    headless=headless,
                    min_healing_failures=min_healing_failures,
                    environment_id=environment_id,
                    target_url=safe_target_url or None,
                    target_page_id=target_page_id,
                    target_name=target_name,
                    progress_callback=lambda progress: _append_element_refresh_progress(task_id, progress),
                )
            )
        duration_ms = int((time.perf_counter() - started_monotonic) * 1000)
        report_path = str(summary.get("html_report_path") or summary.get("output_path") or "")
        page_count = int(summary.get("page_count") or 0)
        _append_element_refresh_progress(
            task_id,
            {
                "stage": "refresh_completed",
                "duration_ms": duration_ms,
                "page_index": page_count,
                "page_total": page_count,
                "scanned_page_count": page_count,
                "element_count": summary.get("element_count", 0),
                "healing_suggestion_count": summary.get("healing_suggestion_count", 0),
                "report_path": report_path,
            },
        )
        _append_element_refresh_log(task_id, f"refresh completed: elements={summary.get('element_count', 0)} healing={summary.get('healing_suggestion_count', 0)} duration_ms={duration_ms}")
        with connect() as conn:
            conn.execute(
                """
                update run_tasks
                set status = 'done', finished_at = ?, return_code = 0, report_path = ?, error = null
                where id = ?
                """,
                (utc_now(), report_path, task_id),
            )
    except Exception as exc:
        try:
            _append_element_refresh_log(task_id, f"refresh failed: {exc}")
            with connect() as conn:
                conn.execute(
                    """
                    update run_tasks
                    set status = 'failed', finished_at = ?, return_code = 1, error = ?
                    where id = ?
                    """,
                    (utc_now(), str(exc), task_id),
                )
        except Exception:
            pass


def _start_element_refresh_task(*, no_scan: bool = True, min_healing_failures: int = 1, base_url: str = "", environment_id: str | None = None, target_url: str | None = None, target_page_id: str | None = None, target_name: str | None = None, include_states: bool = False, headless: bool = True) -> dict:
    task_id = f"ekr-{uuid.uuid4().hex[:12]}"
    if no_scan:
        mode_text = "--no-scan"
    elif target_url:
        mode_text = f"--scan --target-url {target_url} --target-page-id {target_page_id or 'login'} --target-name {target_name or ''} --include-states={include_states} --headless={headless}"
    else:
        mode_text = f"--scan --base-url {base_url} --include-states={include_states} --headless={headless}"
    command = f"internal:element-knowledge-refresh {mode_text} --min-healing-failures {max(1, int(min_healing_failures or 1))}"
    with connect() as conn:
        conn.execute(
            """
            insert into run_tasks(id, mode, case_id, status, command, created_at)
            values (?, 'element-knowledge-refresh', null, 'queued', ?, ?)
            """,
            (task_id, command, utc_now()),
        )
    _append_element_refresh_log(task_id, "refresh queued")
    thread = threading.Thread(
        target=_run_element_refresh_task,
        kwargs={
            "task_id": task_id,
            "no_scan": no_scan,
            "min_healing_failures": max(1, int(min_healing_failures or 1)),
            "base_url": str(base_url or "").strip(),
            "environment_id": environment_id,
            "target_url": str(target_url or "").strip() or None,
            "target_page_id": str(target_page_id or "").strip() or None,
            "target_name": str(target_name or "").strip() or None,
            "include_states": bool(include_states),
            "headless": bool(headless),
        },
        name=f"element-knowledge-refresh-{task_id}",
        daemon=True,
    )
    thread.start()
    return _element_refresh_task_payload(task_id)


async def _validate_element_library_async(task_id: str, environment_id: str) -> dict:
    profile = with_account_credentials(resolve_scan_settings(environment_id))
    auth_mode = str(profile.get("auth_mode") or "auto_login")
    if auth_mode == "cdp_attach":
        cdp_endpoint = str(profile.get("cdp_endpoint") or "")
        _append_element_refresh_progress(task_id, {"stage": "starting_cdp_browser", "environment_id": environment_id, "cdp_endpoint": cdp_endpoint})
        cdp_browser_status = await asyncio.to_thread(_ensure_dedicated_cdp_browser, cdp_endpoint)
        _append_element_refresh_progress(task_id, {"stage": "cdp_browser_ready", "environment_id": environment_id, "cdp_endpoint": cdp_endpoint, "browser_status": cdp_browser_status})
        _append_element_refresh_progress(task_id, {"stage": "attaching_cdp_browser", "environment_id": environment_id})
        session = await attach_browser_over_cdp(cdp_endpoint)
    else:
        headless = bool(profile.get("headless", True))
        _append_element_refresh_progress(task_id, {"stage": "launching_browser", "environment_id": environment_id, "headless": headless})
        session = await launch_browser(headless=headless, system=profile, reuse_storage_state=True)
    try:
        if auth_mode == "cdp_attach":
            if not await is_login_state_valid(session.page, profile):
                raise RuntimeError("CDP Chrome is not authenticated; log in to the dedicated Chrome profile and retry validation")
            _append_element_refresh_progress(task_id, {"stage": "cdp_login_verified", "environment_id": environment_id})
        else:
            await ensure_storage_state_for_profile(session, profile, progress_callback=lambda progress: _append_element_refresh_progress(task_id, progress))
        library = load_json_file(ELEMENT_LIBRARY_PATH)
        if not library.get("elements"):
            raise ValueError("element library is empty; refresh it before validation")
        readiness = {str(target.get("page_id") or ""): target for target in build_scan_targets_from_profile(profile)}
        return await validate_element_library(session.page, library, page_readiness=readiness, progress_callback=lambda progress: _append_element_refresh_progress(task_id, progress))
    finally:
        await close_browser(session)


def _run_element_validation_task(task_id: str, environment_id: str) -> None:
    try:
        with connect() as conn:
            conn.execute("update run_tasks set status = 'running', started_at = ? where id = ?", (utc_now(), task_id))
        report = asyncio.run(_validate_element_library_async(task_id, environment_id))
        _append_element_refresh_progress(task_id, {"stage": "validation_completed", "page_total": report["page_count"], "page_index": report["page_count"], "element_count": report["element_count"], "valid_count": report["summary"]["valid"], "invalid_count": report["summary"]["invalid"], "needs_review_count": report["summary"]["needs_review"], "report_path": report["output_path"]})
        with connect() as conn:
            conn.execute("update run_tasks set status = 'done', finished_at = ?, return_code = 0, report_path = ?, error = null where id = ?", (utc_now(), report["output_path"], task_id))
    except Exception as exc:
        _append_element_refresh_log(task_id, f"validation failed: {exc}")
        with connect() as conn:
            conn.execute("update run_tasks set status = 'failed', finished_at = ?, return_code = 1, error = ? where id = ?", (utc_now(), str(exc), task_id))


def _start_element_validation_task(environment_id: str) -> dict:
    task_id = f"ekv-{uuid.uuid4().hex[:12]}"
    with connect() as conn:
        conn.execute("insert into run_tasks(id, mode, case_id, status, command, created_at) values (?, 'element-knowledge-validation', null, 'queued', ?, ?)", (task_id, f"internal:element-knowledge-validate --environment-id {environment_id}", utc_now()))
    _append_element_refresh_log(task_id, "validation queued")
    threading.Thread(target=_run_element_validation_task, args=(task_id, environment_id), name=f"element-knowledge-validation-{task_id}", daemon=True).start()
    return _element_refresh_task_payload(task_id)


@app.post("/api/element-knowledge/refresh")
def post_element_knowledge_refresh(payload: ElementKnowledgeRefreshRequest | None = None) -> dict:
    request_payload = payload or ElementKnowledgeRefreshRequest()
    return _start_element_refresh_task(
        no_scan=bool(request_payload.no_scan),
        min_healing_failures=max(1, int(request_payload.min_healing_failures or 1)),
        base_url=str(request_payload.base_url or ""),
        environment_id=request_payload.environment_id,
        target_url=str(request_payload.target_url or ""),
        target_page_id=str(request_payload.target_page_id or ""),
        target_name=str(request_payload.target_name or ""),
        include_states=bool(request_payload.include_states),
        headless=bool(request_payload.headless),
    )


@app.post("/api/element-knowledge/validation-tasks")
def post_element_knowledge_validation_task(payload: ElementKnowledgeValidationTaskRequest) -> dict:
    return _start_element_validation_task(payload.environment_id)


def _element_environment_preview(profile: dict) -> dict:
    storage_path = resolve_storage_state_path(profile)
    existing_path = existing_storage_state_for_profile(profile)
    visible_path = existing_path or storage_path
    return {
        **profile,
        "login_configured": bool((profile.get("login") or {}).get("url")),
        "page_count": len(profile.get("pages") or []),
        "storage_state_path": str(visible_path) if visible_path else "",
        "storage_state_exists": bool(existing_path and existing_path.exists()),
        "storage_state_updated_at": datetime.fromtimestamp(existing_path.stat().st_mtime).isoformat() if existing_path and existing_path.exists() else None,
    }


@app.get("/api/element-knowledge/environments")
def get_element_knowledge_environments() -> list[dict]:
    return [
        profile
        for profile in list_environment_profiles()
        if profile.get("element_knowledge_scan_enabled") and profile.get("pages")
    ]


@app.get("/api/element-knowledge/environment-preview")
def get_element_knowledge_environment_preview() -> list[dict]:
    return [_element_environment_preview(profile) for profile in list_environment_profiles()]


@app.get("/api/element-knowledge/environments/{environment_id}")
def get_element_knowledge_environment(environment_id: str) -> dict:
    try:
        return resolve_scan_settings(environment_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.post("/api/element-knowledge/refresh-tasks")
def post_element_knowledge_refresh_task(payload: ElementKnowledgeRefreshTaskRequest | None = None) -> dict:
    request_payload = payload or ElementKnowledgeRefreshTaskRequest()
    return _start_element_refresh_task(
        no_scan=bool(request_payload.no_scan),
        min_healing_failures=max(1, int(request_payload.min_healing_failures or 1)),
        base_url=str(request_payload.base_url or ""),
        environment_id=request_payload.environment_id,
        target_url=str(request_payload.target_url or ""),
        target_page_id=str(request_payload.target_page_id or ""),
        target_name=str(request_payload.target_name or ""),
        include_states=bool(request_payload.include_states),
        headless=bool(request_payload.headless),
    )


@app.get("/api/element-knowledge/refresh-tasks/{task_id}")
def get_element_knowledge_refresh_task(task_id: str) -> dict:
    return _element_refresh_task_payload(task_id)


@app.get("/api/element-knowledge/refresh-tasks")
def list_element_knowledge_refresh_tasks() -> list[dict]:
    with connect() as conn:
        rows = conn.execute(
            """
            select * from run_tasks
            where mode = 'element-knowledge-refresh'
            order by created_at desc
            limit 20
            """
        ).fetchall()
    return rows_to_dicts(rows)


@app.get("/api/reports")
def reports() -> list[dict]:
    with connect() as conn:
        rows = conn.execute(
            """
            select *
            from run_tasks
            where coalesce(report_deleted_at, '') = ''
            order by created_at desc
            limit 500
            """
        ).fetchall()
    tasks = rows_to_dicts(rows)
    report_meta = {item["run_id"]: item for item in list_reports(limit=500)}
    items: list[dict] = []
    for task in tasks:
        detail = load_step_details(str(task["id"])) or {}
        report = report_meta.get(str(task["id"]), {})
        items.append(
            {
                "id": str(task["id"]),
                "run_id": str(task["id"]),
                "case_id": task.get("case_id"),
                "case_name": detail.get("case_name") or report.get("case_name") or task.get("case_id") or str(task["id"]),
                "mode": _normalized_mode(str(task.get("mode") or "")),
                "status": _normalized_status(str(task.get("status") or "")),
                "operator": "admin",
                "started_at": task.get("started_at") or task.get("created_at"),
                "finished_at": task.get("finished_at"),
                "has_report": bool(task.get("report_path")),
                "has_evidence": bool(evidence_summary(str(task["id"])).get("root")),
            }
        )
    items.sort(key=lambda item: (0 if item["status"] == "failed" else 1, item.get("finished_at") or item.get("started_at") or ""), reverse=False)
    return items


def _get_report_task_or_404(run_id: str) -> dict:
    with connect() as conn:
        row = conn.execute("select * from run_tasks where id = ?", (run_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="report not found")
    return dict(row)


def _ensure_report_available(run_id: str) -> dict:
    task = _get_report_task_or_404(run_id)
    if task.get("report_deleted_at"):
        raise HTTPException(status_code=404, detail="report not found")
    return task


@app.delete("/api/reports/{run_id}")
def delete_report(run_id: str) -> dict:
    task = _ensure_report_available(run_id)
    now = utc_now()
    with connect() as conn:
        conn.execute("update run_tasks set report_deleted_at = ? where id = ?", (now, run_id))
    return {"ok": True, "run_id": run_id, "deleted_at": now, "mode": _normalized_mode(str(task.get("mode") or ""))}


@app.post("/api/reports/batch-delete")
def batch_delete_reports(payload: ReportBatchDeleteRequest) -> dict:
    run_ids = [str(item).strip() for item in payload.run_ids if str(item).strip()]
    if not run_ids:
        raise HTTPException(status_code=400, detail="run_ids is required")
    placeholders = ",".join("?" for _ in run_ids)
    now = utc_now()
    with connect() as conn:
        rows = conn.execute(
            f"select id from run_tasks where coalesce(report_deleted_at, '') = '' and id in ({placeholders})",
            run_ids,
        ).fetchall()
        existing_ids = [str(row["id"]) for row in rows]
        if existing_ids:
            update_placeholders = ",".join("?" for _ in existing_ids)
            conn.execute(
                f"update run_tasks set report_deleted_at = ? where id in ({update_placeholders})",
                [now, *existing_ids],
            )
    return {"ok": True, "deleted_count": len(existing_ids), "run_ids": existing_ids, "deleted_at": now}


@app.get("/api/reports/{run_id}")
def report_detail(run_id: str) -> dict:
    _ensure_report_available(run_id)
    report, meta, screenshots = load_report_or_run_detail_report(run_id)
    analysis = load_cached_report_analysis(run_id, report) or ai_service.analyze_run_report(report, screenshots, evidence_log_lines(run_id))
    return {
        "run_id": run_id,
        "metadata": meta,
        "markdown": report,
        "screenshots": screenshots,
        "evidence": evidence_summary(run_id),
        "analysis": analysis,
    }


@app.post("/api/reports/{run_id}/analyze")
def analyze_report_with_ai(run_id: str, payload: AnalyzeReportRequest | None = None) -> dict:
    _ensure_report_available(run_id)
    report, _meta, screenshots = load_report_or_run_detail_report(run_id)
    settings = get_ai_settings(mask_key=False)
    force = bool(payload.force) if payload else False
    cached = None if force else load_cached_report_analysis(run_id, report, settings)
    if cached:
        return {**cached, "cached": True}
    with connect() as conn:
        rows = conn.execute("select line from run_logs where run_id = ? order by id", (run_id,)).fetchall()
    try:
        analysis = ai_service.analyze_run_report_with_ai(
            report,
            screenshots,
            [row["line"] for row in rows] + evidence_log_lines(run_id),
            settings,
        )
        save_report_analysis(run_id, report, settings, analysis)
        return {**analysis, "cached": False}
    except AIConfigurationError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except AIProviderError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@app.get("/api/reports/{run_id}/analyses")
def report_analysis_versions(run_id: str) -> list[dict]:
    _ensure_report_available(run_id)
    report, _meta, _screenshots = load_report_or_run_detail_report(run_id)
    return load_report_analysis_versions(run_id, report)


def load_report_or_run_detail_report(run_id: str) -> tuple[str, dict, list[dict]]:
    try:
        report = read_report(run_id)
        meta = parse_report(report)
        screenshots = meta["screenshots"]
        if not screenshots and meta["case_id"]:
            screenshots = [screenshot_payload(meta["case_id"], path) for path in latest_screenshots(meta["case_id"])]
        return report, meta, screenshots
    except FileNotFoundError:
        detail = _build_unified_run_detail(run_id)
        report = build_markdown_from_run_detail(detail)
        meta = {
            "case_id": detail.get("case_id") or "",
            "case_name": detail.get("case_name") or detail.get("case_id") or run_id,
            "status": detail.get("status") or "unknown",
            "screenshots": detail.get("screenshots") or [],
            "observed_asset_path": "",
        }
        return report, meta, list(detail.get("screenshots") or [])


def build_markdown_from_run_detail(detail: dict) -> str:
    summary = detail.get("summary") if isinstance(detail.get("summary"), dict) else {}
    steps = detail.get("steps") if isinstance(detail.get("steps"), list) else []
    lines = [
        "# 自动化执行报告",
        "",
        f"- run id: {detail.get('run_id') or ''}",
        f"- case id: {detail.get('case_id') or ''}",
        f"- case name: {detail.get('case_name') or detail.get('case_id') or ''}",
        f"- mode: {detail.get('mode') or ''}",
        f"- status: {detail.get('status') or ''}",
        f"- started at: {detail.get('started_at') or ''}",
        f"- finished at: {detail.get('finished_at') or ''}",
        f"- final url: {detail.get('final_url') or ''}",
        "",
        "## 执行结论",
        str(summary.get("conclusion") or summary.get("ai_analysis") or ""),
        "",
        "## 故障描述",
        str(summary.get("failure_reason") or ""),
        "",
        "## 执行步骤",
    ]
    for step in steps:
        if not isinstance(step, dict):
            continue
        index = step.get("step_index") or ""
        title = step.get("title") or step.get("step_code") or ""
        status = step.get("status") or ""
        text = step.get("summary") or step.get("ai_analysis") or step.get("error_message") or ""
        lines.extend([f"- Step {index}: {title} [{status}]", f"  - {text}"])
    return "\n".join(lines).strip()


@app.get("/api/screenshots/latest/{case_id}/{filename}")
def screenshot_file(case_id: str, filename: str) -> FileResponse:
    path = SCREENSHOTS_LATEST_DIR / case_id / filename
    if not path.exists() or path.suffix.lower() != ".png":
        raise HTTPException(status_code=404, detail="screenshot not found")
    return FileResponse(path)


@app.get("/api/screenshots/runs/{run_id}/{filename}")
def run_screenshot_file(run_id: str, filename: str) -> FileResponse:
    path = SCREENSHOTS_RUNS_DIR / run_id / filename
    if not path.exists() or path.suffix.lower() != ".png":
        raise HTTPException(status_code=404, detail="screenshot not found")
    return FileResponse(path)


def screenshot_payload(case_id: str, path: str) -> dict[str, str]:
    filename = path.replace("\\", "/").rsplit("/", 1)[-1]
    return {
        "case_id": case_id,
        "filename": filename,
        "path": path,
        "url": f"/api/screenshots/latest/{case_id}/{filename}",
    }


def run_screenshot_payload(run_id: str, path: str) -> dict[str, str]:
    filename = path.replace("\\", "/").rsplit("/", 1)[-1]
    return {
        "case_id": run_id,
        "filename": filename,
        "path": path,
        "url": f"/api/screenshots/runs/{run_id}/{filename}",
    }


def path_health(path: Path) -> dict:
    exists = path.exists()
    stat = path.stat() if exists else None
    return {
        "path": str(path),
        "exists": exists,
        "is_dir": path.is_dir() if exists else False,
        "updated_at": datetime_from_mtime(stat.st_mtime) if stat else None,
    }


def sqlite_health() -> dict:
    exists = DB_PATH.exists()
    stat = DB_PATH.stat() if exists else None
    return {
        "path": str(DB_PATH),
        "exists": exists,
        "size_bytes": stat.st_size if stat else 0,
        "updated_at": datetime_from_mtime(stat.st_mtime) if stat else None,
    }


def datetime_from_mtime(mtime: float) -> str:
    return datetime.fromtimestamp(mtime).isoformat(timespec="seconds")


def report_hash(report: str) -> str:
    return sha256(report.encode("utf-8")).hexdigest()


def load_cached_report_analysis(run_id: str, report: str, settings: dict | None = None) -> dict | None:
    digest = report_hash(report)
    provider = settings.get("provider") if settings else None
    model = settings.get("model") if settings else None
    with connect() as conn:
        if provider and model:
            row = conn.execute(
                """
                select * from report_analyses
                where run_id = ? and report_hash = ? and provider = ? and model = ?
                order by updated_at desc
                limit 1
                """,
                (run_id, digest, provider, model),
            ).fetchone()
        else:
            row = conn.execute(
                """
                select * from report_analyses
                where run_id = ? and report_hash = ?
                order by updated_at desc
                limit 1
                """,
                (run_id, digest),
            ).fetchone()
    if not row:
        return None
    try:
        analysis = json.loads(row["analysis_json"])
    except json.JSONDecodeError:
        return None
    return {**analysis, "cached": True, "cached_at": row["updated_at"]}


def save_report_analysis(run_id: str, report: str, settings: dict, analysis: dict) -> None:
    now = utc_now()
    provider = str(settings.get("provider", ""))
    model = str(settings.get("model", ""))
    payload = json.dumps(analysis, ensure_ascii=False)
    digest = report_hash(report)
    with connect() as conn:
        conn.execute(
            """
            insert into report_analysis_versions(run_id, report_hash, provider, model, analysis_json, created_at)
            values (?, ?, ?, ?, ?, ?)
            """,
            (run_id, digest, provider, model, payload, now),
        )
        conn.execute(
            """
            insert into report_analyses(run_id, report_hash, provider, model, analysis_json, created_at, updated_at)
            values (?, ?, ?, ?, ?, ?, ?)
            on conflict(run_id, report_hash, provider, model) do update set
              analysis_json = excluded.analysis_json,
              updated_at = excluded.updated_at
            """,
            (run_id, digest, provider, model, payload, now, now),
        )


def load_report_analysis_versions(run_id: str, report: str) -> list[dict]:
    digest = report_hash(report)
    with connect() as conn:
        rows = conn.execute(
            """
            select * from report_analysis_versions
            where run_id = ? and report_hash = ?
            order by created_at desc, id desc
            limit 20
            """,
            (run_id, digest),
        ).fetchall()
    versions: list[dict] = []
    for row in rows:
        try:
            analysis = json.loads(row["analysis_json"])
        except json.JSONDecodeError:
            continue
        versions.append(
            {
                "id": row["id"],
                "run_id": row["run_id"],
                "provider": row["provider"],
                "model": row["model"],
                "created_at": row["created_at"],
                "analysis": {**analysis, "cached": True, "cached_at": row["created_at"]},
            }
        )
    return versions


def expected_text_hash(expected_text: str, context_sig: str = "") -> str:
    if not context_sig:
        return sha256(expected_text.encode("utf-8")).hexdigest()
    return sha256(f"{expected_text}\n{context_sig}".encode("utf-8")).hexdigest()


def load_cached_assertion_analysis(expected_text: str, settings: dict[str, Any], context_sig: str = "") -> list[dict] | None:
    digest = expected_text_hash(expected_text, context_sig)
    provider = str(settings.get("provider", ""))
    model = str(settings.get("model", ""))
    with connect() as conn:
        row = conn.execute(
            """
            select assertions_json, updated_at from assertion_analyses
            where expected_hash = ? and provider = ? and model = ?
            order by updated_at desc
            limit 1
            """,
            (digest, provider, model),
        ).fetchone()
    if not row:
        return None
    try:
        assertions = json.loads(row["assertions_json"])
    except json.JSONDecodeError:
        return None
    if not isinstance(assertions, list):
        return None
    for item in assertions:
        if isinstance(item, dict):
            item["status"] = "queued"
            item["actual"] = ""
            item["evidence_source"] = ""
            item["reason"] = ""
            item["cached"] = True
    return assertions


def save_assertion_analysis(expected_text: str, settings: dict[str, Any], assertions: list[dict], context_sig: str = "") -> None:
    now = utc_now()
    provider = str(settings.get("provider", ""))
    model = str(settings.get("model", ""))
    digest = expected_text_hash(expected_text, context_sig)
    snapshot = [
        {k: v for k, v in item.items() if k not in {"status", "actual", "evidence_source", "reason", "cached"}}
        for item in assertions
        if isinstance(item, dict)
    ]
    payload = json.dumps(snapshot, ensure_ascii=False)
    with connect() as conn:
        conn.execute(
            """
            insert into assertion_analyses(expected_hash, provider, model, assertions_json, created_at, updated_at)
            values (?, ?, ?, ?, ?, ?)
            on conflict(expected_hash, provider, model) do update set
              assertions_json = excluded.assertions_json,
              updated_at = excluded.updated_at
            """,
            (digest, provider, model, payload, now, now),
        )


def is_confirmed_status(status: str) -> bool:
    normalized = status.strip().lower()
    return normalized in {"已确认", "confirmed", "passed", "通过"} or "确认" in status


def load_requirement_title(requirement_id: int | None) -> str:
    if not requirement_id:
        return ""
    with connect() as conn:
        row = conn.execute("select title from requirements where id = ?", (requirement_id,)).fetchone()
    return str(row["title"]) if row else ""


def next_test_point_sort_order(parent_id: int | None) -> int:
    with connect() as conn:
        if parent_id is None:
            row = conn.execute("select coalesce(max(sort_order), 0) + 1 as next_order from test_points where parent_id is null").fetchone()
        else:
            row = conn.execute(
                "select coalesce(max(sort_order), 0) + 1 as next_order from test_points where parent_id = ?",
                (parent_id,),
            ).fetchone()
    return int(row["next_order"] if row else 1)


def ensure_manual_requirement() -> int:
    title = "测试点思维导图手工维护"
    project_id = resolve_requirement_project_id()
    with connect() as conn:
        row = conn.execute("select id from requirements where title = ? order by id limit 1", (title,)).fetchone()
        if row:
            conn.execute("update requirements set project_id = coalesce(nullif(project_id, ''), ?) where id = ?", (project_id, int(row["id"])))
            return int(row["id"])
        now = utc_now()
        cur = conn.execute(
            """
            insert into requirements(title, document, status, analysis_summary, risk_summary, case_count, project_id, created_at, updated_at)
            values (?, ?, 'manual', '', '', 0, 'proj-icm-default', ?, ?)
            """,
            (title, "测试点菜单中手工新增的测试点。", now, now),
        )
        return int(cur.lastrowid)


def ensure_generated_requirement(points: list[dict]) -> int:
    title = "测试点思维导图选中生成"
    document = "\n".join(str(point.get("name", "")) for point in points)
    now = utc_now()
    with connect() as conn:
        cur = conn.execute(
            """
            insert into requirements(title, document, status, analysis_summary, risk_summary, case_count, project_id, created_at, updated_at)
            values (?, ?, 'draft', '', '', ?, 'proj-icm-default', ?, ?)
            """,
            (title, document, len(points), now, now),
        )
        return int(cur.lastrowid)


def load_case_draft(draft_id: int) -> dict | None:
    with connect() as conn:
        row = conn.execute(
            """
            select cd.*, r.title as requirement_title
            from case_drafts cd
            left join requirements r on r.id = cd.requirement_id
            where cd.id = ?
            """,
            (draft_id,),
        ).fetchone()
    return normalize_case_draft(dict(row)) if row else None


def normalize_case_draft(row: dict) -> dict:
    raw_ids = row.get("source_test_point_ids") or "[]"
    try:
        source_ids = json.loads(raw_ids)
    except (TypeError, json.JSONDecodeError):
        source_ids = []
    return {**row, "source_test_point_ids": source_ids if isinstance(source_ids, list) else []}


def draft_title_for_template(template: str) -> str:
    return {
        "functional": "测试点生成 - 功能用例草稿",
        "negative": "测试点生成 - 异常用例草稿",
        "regression": "测试点生成 - 回归用例草稿",
        "e2e": "测试点生成 - 端到端链路草稿",
    }.get(template, "测试点生成 - YAML 草稿")


def normalize_case_id(case_id: str) -> str:
    value = case_id.strip().upper()
    if not re.fullmatch(r"TC-ICM-[0-9A-Z-]+", value):
        raise HTTPException(status_code=400, detail="case_id must look like TC-ICM-013")
    return value


def normalize_case_filename(filename: str | None, case_id: str) -> str:
    value = (filename or f"{case_id.lower()}-generated.yaml").strip()
    if not value.endswith((".yaml", ".yml")):
        value = f"{value}.yaml"
    if "/" in value or "\\" in value or ".." in value:
        raise HTTPException(status_code=400, detail="filename must be a plain yaml filename")
    if not value.lower().startswith(case_id.lower()):
        value = f"{case_id.lower()}-{value}"
    return value


def replace_yaml_case_id(yaml_text: str, case_id: str) -> str:
    if re.search(r"^id:\s*.+$", yaml_text, flags=re.MULTILINE):
        return re.sub(r"^id:\s*.+$", f"id: {case_id}", yaml_text, count=1, flags=re.MULTILINE)
    return f"id: {case_id}\n{yaml_text}"


def load_observed_asset_for_run(run_id: str) -> dict:
    path = OBSERVED_ASSET_DIR / f"{run_id}.json"
    if not path.exists():
        raise HTTPException(status_code=404, detail="observed asset not found")
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="observed asset is not valid JSON") from exc
    if not isinstance(data, dict):
        raise HTTPException(status_code=400, detail="observed asset root must be an object")
    return data


def find_case_yaml(case_id: str) -> Path:
    for path in sorted(TEST_CASE_DIR.glob(f"{case_id}*.yaml")):
        data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        if data.get("id") == case_id:
            return path
    raise HTTPException(status_code=404, detail=f"case YAML not found: {case_id}")


def merge_automation_asset(existing: dict, observed: dict) -> dict:
    merged = {
        "status": "verified",
        "source": observed.get("source", "playwright_observed"),
        "observed_at": observed.get("observed_at"),
        "evidence": observed.get("evidence", {}),
        "operation_steps": existing.get("operation_steps") or observed.get("operation_steps") or [],
        "selectors": existing.get("selectors") or observed.get("selectors") or {},
        "input_values": existing.get("input_values") or observed.get("input_values") or {},
        "assertions": existing.get("assertions") or observed.get("assertions") or [],
    }
    return {key: value for key, value in merged.items() if value is not None}


def validate_case_yaml(yaml_text: str) -> dict:
    errors: list[str] = []
    warnings: list[str] = []
    try:
        data = yaml.safe_load(yaml_text)
    except yaml.YAMLError as exc:
        return {"valid": False, "errors": [f"YAML syntax error: {exc}"], "warnings": [], "parsed_id": None}

    if not isinstance(data, dict):
        return {"valid": False, "errors": ["YAML root must be a mapping/object"], "warnings": [], "parsed_id": None}

    for field in ("id", "title", "status"):
        if not is_non_empty_scalar(data.get(field)):
            errors.append(f"missing or empty required field: {field}")

    for field in ("steps", "expected_results"):
        if not is_non_empty_list(data.get(field)):
            errors.append(f"missing or empty required list: {field}")

    asset = data.get("automation_asset")
    if not isinstance(asset, dict):
        errors.append("missing or invalid required mapping: automation_asset")
    else:
        for field in ("operation_steps", "assertions"):
            if not is_non_empty_list(asset.get(field)):
                errors.append(f"automation_asset.{field} must be a non-empty list")
        if not is_non_empty_selectors(asset.get("selectors")):
            errors.append("automation_asset.selectors must be a non-empty list or mapping")
        if "input_values" not in asset or not isinstance(asset.get("input_values"), dict):
            errors.append("automation_asset.input_values must be a mapping")
        elif not asset.get("input_values"):
            warnings.append("automation_asset.input_values is empty; confirm this case really has no inputs")

    parsed_id = str(data.get("id")) if data.get("id") is not None else None
    if parsed_id == "TC-ICM-DRAFT":
        warnings.append("case id is still TC-ICM-DRAFT; promote will replace it with the formal case id")
    return {"valid": not errors, "errors": errors, "warnings": warnings, "parsed_id": parsed_id}


def is_non_empty_scalar(value: object) -> bool:
    return isinstance(value, (str, int, float)) and str(value).strip() != ""


def is_non_empty_list(value: object) -> bool:
    return isinstance(value, list) and len(value) > 0


def is_non_empty_selectors(value: object) -> bool:
    return (isinstance(value, list) and len(value) > 0) or (isinstance(value, dict) and len(value) > 0)


def load_requirement_detail(requirement_id: int) -> dict | None:
    with connect() as conn:
        requirement = conn.execute("select * from requirements where id = ?", (requirement_id,)).fetchone()
        if not requirement:
            return None
        points = conn.execute(
            "select * from test_points where requirement_id = ? order by coalesce(sort_order, id), id",
            (requirement_id,),
        ).fetchall()
        drafts = conn.execute("select * from case_drafts where requirement_id = ? order by created_at desc", (requirement_id,)).fetchall()
    return {
        "requirement": dict(requirement),
        "test_points": rows_to_dicts(points),
        "drafts": rows_to_dicts(drafts),
    }
